import os
from dateutil import parser
from openpyxl.styles import Font
from pypdf import PdfReader
import re
import shutil
import io
import openpyxl
from openpyxl.styles import Alignment
import time
import pandas as pd
import numpy as np
import PyPDF2
from PyPDF2 import PdfReader, PdfWriter
import pdfplumber
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import regex as re
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
import datefinder
from calendar import monthrange
import calendar
from openpyxl.styles import Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
from openpyxl.styles import Border, Side
from openpyxl.styles import Font
import logging
import openpyxl
from openpyxl.styles import Alignment

bold_font = Font(bold=True)
pd.options.display.float_format = "{:,.2f}".format
pd.set_option("display.max_columns", None)
pd.set_option("display.max_rows", None)
pd.set_option("display.width", None)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

from utils.common_functions import CommonFunctions
from utils.model_loader import model
from collections import deque

class ATSFunctions:
    def __init__(
        self, bank_names, pdf_paths, pdf_passwords, CASE_ID
    ):
        self.writer = None
        self.bank_names = bank_names
        self.pdf_paths = pdf_paths
        self.pdf_passwords = pdf_passwords
        self.account_number = ""
        self.file_name = None
        self.CASE_ID = CASE_ID
        start_date = "1"
        end_date = "2"
        self.commoner = CommonFunctions(
            bank_names, pdf_paths, pdf_passwords, start_date, end_date, CASE_ID
        )

    def single_name_entity_addition(self, df):
        df['Entity'] = df['Entity']
        return df

    def single_analyze_entities(self, df):
        entity_analysis_df = df.groupby('Entity').agg(
            Entity_Name=('Entity', 'first'),
            Total_Debit=('Debit', 'sum'),
            Total_Credit=('Credit', 'sum'),
            No_of_times_occurred=('Entity', 'count')
        ).reset_index(drop=True)
        entity_analysis_df = entity_analysis_df.sort_values(by='No_of_times_occurred', ascending=False).reset_index(
            drop=True)
        return entity_analysis_df

    def single_bidirectional_analysis(self, df):
        # 1. Split Analysis Between Inflows and Outflows
        outflows = df[df['Debit'] > 0]
        inflows = df[df['Credit'] > 0]

        # 2. Category-Level Bidirectional Analysis
        category_summary = df.groupby('Category').agg({
            'Debit': 'sum',
            'Credit': 'sum',
            'Balance': 'last',
            'Value Date': 'count'
        }).rename(columns={'Value Date': 'Transaction Count'}).reset_index()
        category_summary['Net Flow'] = category_summary['Credit'] - category_summary['Debit']

        # 3. Entity-Level Bidirectional Analysis
        entity_summary = df.groupby('Entity').agg({
            'Debit': 'sum',
            'Credit': 'sum',
            'Balance': 'last',
            'Value Date': 'count'
        }).rename(columns={'Value Date': 'Transaction Count'}).reset_index()
        entity_summary['Net Flow'] = entity_summary['Credit'] - entity_summary['Debit']

        # 4. Net Cash Flow Calculation
        net_cash_flow = df['Credit'].sum() - df['Debit'].sum()

        # 5. Trend Analysis Over Time
        df['Value Date'] = pd.to_datetime(df['Value Date'], format='%d-%m-%Y', errors='coerce')
        df['YearMonth'] = df['Value Date'].dt.to_period('M')
        monthly_summary = df.groupby('YearMonth').agg({
            'Debit': 'sum',
            'Credit': 'sum',
            'Balance': 'last'
        }).reset_index()
        monthly_summary['Net Flow'] = monthly_summary['Credit'] - monthly_summary['Debit']

        # 6. Balance Analysis
        df['Daily Net Flow'] = df['Credit'].fillna(0) - df['Debit'].fillna(0)
        df['Cumulative Balance'] = df['Daily Net Flow'].cumsum() + df['Balance'].iloc[0]

        # 7. Extract Useful Insights for Bidirectional Analysis
        top_expense_categories = category_summary.sort_values(by='Debit', ascending=False).head(5)
        top_income_sources = entity_summary.sort_values(by='Credit', ascending=False).head(5)
        high_net_flow_entities = entity_summary.sort_values(by='Net Flow', ascending=False).head(5)

        # Returning all results
        return {
            'outflows': outflows,
            'inflows': inflows,
            'category_summary': category_summary,
            'entity_summary': entity_summary,
            'net_cash_flow': net_cash_flow,
            'monthly_summary': monthly_summary,
            'cumulative_balance': df[['Value Date', 'Cumulative Balance']],
            'top_expense_categories': top_expense_categories,
            'top_income_sources': top_income_sources,
            'high_net_flow_entities': high_net_flow_entities
        }

    def single_fifo_analysis(self, df):
        # Step 1: Separate Credits and Debits
        credits = df[df['Credit'] > 0].copy()
        debits = df[df['Debit'] > 0].copy()
        credits['Remaining'] = credits['Credit']

        # Step 2: Apply FIFO Logic for Allocating Credits to Debits
        allocation_summary = []

        for j, credit in credits.iterrows():
            credit_amount = credit['Remaining']

            for i, debit in debits.iterrows():
                if credit_amount <= 0:
                    break
                if debit['Debit'] > 0:
                    allocation = min(credit_amount, debit['Debit'])

                    # Create detailed allocation record
                    allocation_summary.append({
                        'allocated_date': credit['Value Date'],
                        'allocated_entity': credit['Entity'],
                        'allocated_category': credit['Category'],
                        'allocated_amount': allocation,
                        'allocated_description': credit['Description'],
                        'used_in_date': debit['Value Date'],
                        'used_in_entity': debit['Entity'],
                        'used_in_category': debit['Category'],
                        'used_in_description': debit['Description'],
                        'remaining_amount': credit_amount - allocation,
                        'days_in_between_allocation_and_usage': max((debit['Value Date'] - credit['Value Date']).days,
                                                                    0)
                    })

                    # Update remaining balance of the credit and debit
                    credits.at[j, 'Remaining'] -= allocation
                    debits.at[i, 'Debit'] -= allocation

                    # Decrease credit amount to be allocated
                    credit_amount -= allocation

        allocation_df = pd.DataFrame(allocation_summary)

        # Step 3: Group by Entity and Category for Detailed Insights
        entity_summary = allocation_df.groupby('allocated_entity').agg({
            'allocated_amount': 'sum',
            'used_in_description': 'count'
        }).rename(columns={'allocated_amount': 'Total_Allocated', 'used_in_description': 'Transaction_Count'})

        category_summary = allocation_df.groupby('allocated_category').agg({
            'allocated_amount': 'sum',
            'used_in_description': 'count'
        }).rename(columns={'allocated_amount': 'Total_Allocated', 'used_in_description': 'Transaction_Count'})

        # Step 4: Add Summary Notes
        summary_note = """
        FIFO Analysis Summary:
        - Allocations are made using a FIFO approach where the earliest credits are used to cover debits.
        - The 'allocation_df' DataFrame contains detailed records of all debit-credit allocations.
        """
        print(summary_note)

        return {
            'debits': debits,
            'credits': credits,
            'allocation_df': allocation_df,
            'entity_summary': entity_summary,
            'category_summary': category_summary
        }

    def single_money_trail_analysis(self, df):
        # Step 1: Summarize inflows and outflows for each Entity
        entity_summary = df.groupby('Entity').agg(
            total_inflows=pd.NamedAgg(column='Credit', aggfunc='sum'),
            total_outflows=pd.NamedAgg(column='Debit', aggfunc='sum'),
            transaction_count=pd.NamedAgg(column='Value Date', aggfunc='count')
        ).reset_index()

        print("Entity-wise Summary of Inflows and Outflows:")
        print(entity_summary)

        # Step 2: Identify transaction patterns over time
        df['Value Date'] = pd.to_datetime(df['Value Date'])

        # Deeper analysis - calculating net flow and cumulative balance per entity
        df['Net Flow'] = df['Credit'].fillna(0) - df['Debit'].fillna(0)
        cumulative_balance = df.groupby('Entity').agg(
            total_net_flow=pd.NamedAgg(column='Net Flow', aggfunc='sum'),
            avg_balance=pd.NamedAgg(column='Balance', aggfunc='mean'),
            max_balance=pd.NamedAgg(column='Balance', aggfunc='max'),
            min_balance=pd.NamedAgg(column='Balance', aggfunc='min')
        ).reset_index()

        print("Deeper Analysis - Cumulative Balance and Net Flow per Entity:")
        print(cumulative_balance)

        # Finding entities with significant inflows or outflows
        significant_inflows = df[df['Credit'] > df['Credit'].quantile(0.95)]
        significant_outflows = df[df['Debit'] > df['Debit'].quantile(0.95)]

        print("Entities with Significant Inflows (Top 5%):")
        print(significant_inflows[['Entity', 'Credit', 'Value Date', 'Description']])

        print("Entities with Significant Outflows (Top 5%):")
        print(significant_outflows[['Entity', 'Debit', 'Value Date', 'Description']])

        return df

        # Daily FIFO Analysis Function
    
    def fifo_allocation(self, df, period = 'week', category_wise = 'no', entities_of_interest = []):
        imp_categories = ["UPI-Cr", "UPI-Dr", "Debtors", "Creditors", "Donation", "Loan", "Loan given", "Rent Paid"]

        if category_wise == 'yes':
            df = df[df['Category'].isin(imp_categories)]

        # Convert Value Date to datetime format and sort chronologically
        df['Value Date'] = pd.to_datetime(df['Value Date'])
        df = df.sort_values(by='Value Date').reset_index(drop=True)

        # Define time period mapping
        period_map = {
            "week": pd.Timedelta(weeks=1),  
            "month": pd.Timedelta(days=30),
            "half_year": pd.Timedelta(days=182),
            "year": pd.Timedelta(days=365)
        }

        # Get the period duration based on the user input
        duration = period_map.get(period.lower(), pd.Timedelta(weeks=1))  # Default to "week" if invalid

        # Define the unique people/entities involved
        # Create the `people` list dynamically based on whether `entities_of_interest` is provided
        people = list(df['Name'].unique()) if entities_of_interest is None else list(set(df['Name'].unique()) | set(entities_of_interest))

        # Initialize a dictionary to store the results
        utilization_dict = {}

        # Iterate through each person to find relevant credit transactions
        for person in people:
            # Filter out credit transactions where the person is receiving money from others in the 'people' list
            person_df = df[(df['Name'] == person) & (df['Credit'] > 0) & (df['Entity'].isin(people))]

            # Process each credit transaction for the person
            for idx, credit_row in person_df.iterrows():
                credit_date = credit_row['Value Date']
                credit_amount = credit_row['Credit']
                credit_entity = credit_row['Entity']

                # Find the sender's transactions from a week before the credit transaction
                sender_history = df[
                    (df['Name'] == credit_entity) &
                    (df['Value Date'] >= credit_date - pd.Timedelta(weeks=1)) &
                    (df['Value Date'] < credit_date)
                ].reset_index(drop=True)

                # Filter the recipient's transactions within a week after the credit transaction
                subsequent_transactions = df[
                    (df['Name'] == person) &
                    (df['Value Date'] > credit_date) &
                    (df['Value Date'] <= credit_date + pd.Timedelta(weeks=1))
                ].reset_index(drop=True)

                # Add utilized and remaining credit columns to the DataFrame
                utilization_df = pd.concat([credit_row.to_frame().T, subsequent_transactions]).reset_index(drop=True)
                utilization_df['Utilized Credit'] = 0
                utilization_df['Remaining Credit'] = credit_amount

                # Track utilized and remaining credit for each subsequent transaction
                for i in range(1, len(utilization_df)):
                    row = utilization_df.iloc[i]
                    if row['Debit'] > 0:
                        utilization_df.at[i, 'Utilized Credit'] = utilization_df.at[i - 1, 'Utilized Credit'] + row['Debit']
                        utilization_df.at[i, 'Remaining Credit'] = credit_amount - utilization_df.at[i, 'Utilized Credit']
                    elif row['Credit'] > 0 and i != 0:
                        # If another credit transaction occurs, update the remaining credit
                        credit_amount += row['Credit']
                        utilization_df.at[i, 'Remaining Credit'] = credit_amount

                # Create a key for the utilization dictionary entry
                key = f"Utilization of Credit ({credit_row['Credit']}) received by {person} from {credit_entity} on {credit_date.date()}:"

                # Store both sender's history and utilization data in the dictionary
                utilization_dict[key] = {
                    'LIFO': sender_history[['Value Date', 'Name', 'Description', 'Debit', 'Credit','Category', 'Entity']],
                    'FIFO': utilization_df[['Value Date', 'Name', 'Description', 'Debit', 'Credit', 'Category', 'Entity', 'Utilized Credit', 'Remaining Credit']]
                }
    
        return utilization_dict

    # Analysis Function
    def analyze_period(self, df, freq):
        # Calculate overall period in days
        overall_period = (df["Value Date"].max() - df["Value Date"].min()).days

        if freq == 'M' and overall_period < 30:
            return pd.DataFrame()
        elif freq == 'W' and overall_period < 7:
            return pd.DataFrame()
        elif freq == '6M' and overall_period < 180:
            return pd.DataFrame()
        elif freq == 'Y' and overall_period < 365:
            return pd.DataFrame()
        else:
            return df.resample(freq, on='Value Date').agg(
                {'Debit': 'sum', 'Credit': 'sum', 'Balance': 'last'}).reset_index()

    def cumulative_bidirectional_analysis(self, df, entities_of_interest=[]):

        # Convert 'Value Date' to datetime
        df['Value Date'] = pd.to_datetime(df['Value Date'])

        # Extract unique names from the 'Name' column
        people = df['Name'].unique()

        # Set to keep track of already processed pairs
        processed_pairs = set()

        # List to store bidirectional analysis results
        bidirectional_results = []

        # Helper function to analyze a pair
        def analyze_pair(temp_df, pair):
            # Start and end dates for the analysis
            start_date = temp_df['Value Date'].iloc[0]
            end_date = temp_df['Value Date'].iloc[-1]

            # Calculate totals and net flow
            total_credit_amount = temp_df[temp_df['Credit'] > 0]['Credit'].sum()
            total_debit_amount = temp_df[temp_df['Debit'] > 0]['Debit'].sum()
            total_credit_transactions = len(temp_df[temp_df['Credit'] > 0])
            total_debit_transactions = len(temp_df[temp_df['Debit'] > 0])
            net_exchange = total_credit_amount - total_debit_amount

            # Calculate average amount exchanged
            total_transactions = total_credit_transactions + total_debit_transactions
            average_amount_exchanged = (total_credit_amount + total_debit_amount) / total_transactions if total_transactions > 0 else 0

            # Calculate median and standard deviation of transaction amounts
            all_transactions = temp_df[['Credit', 'Debit']].stack().reset_index(drop=True)
            median_amount_exchanged = all_transactions.median() if not all_transactions.empty else 0
            std_dev_amount_exchanged = all_transactions.std() if not all_transactions.empty else 0

            # Determine highest amount exchanged and its type
            highest_credit = temp_df['Credit'].max() if not temp_df['Credit'].isna().all() else 0
            highest_debit = temp_df['Debit'].max() if not temp_df['Debit'].isna().all() else 0
            if highest_credit >= highest_debit:
                highest_amount_exchanged = highest_credit
                highest_amount_type = 'Credit'
            else:
                highest_amount_exchanged = highest_debit
                highest_amount_type = 'Debit'

            # Construct the result dictionary
            return {
                'Entity One': pair[0],
                'Entity Two': pair[1],
                'Total Credit Transactions': total_credit_transactions,
                'Total Debit Transactions': total_debit_transactions,
                'Total Credit Amount': total_credit_amount,
                'Total Debit Amount': total_debit_amount,
                'Net Exchange (Credit - Debit)': net_exchange,
                'Average Amount Exchanged': average_amount_exchanged,
                'Median Amount Exchanged': median_amount_exchanged,
                'Standard Deviation of Amount Exchanged': std_dev_amount_exchanged,
                'Highest Amount Exchanged': highest_amount_exchanged,
                'Highest Amount Type': highest_amount_type,
                'First Transaction Date': start_date,
                'Last Transaction Date': end_date
            }

        # Original analysis: Iterate through unique pairs of people
        for i, person in enumerate(people):
            for j in range(i + 1, len(people)):
                entity = people[j]
                pair = tuple(sorted([person, entity]))
                if pair in processed_pairs:
                    continue
                processed_pairs.add(pair)
                temp_df = df[((df['Name'] == person) & (df['Entity'] == entity)) | 
                            ((df['Name'] == entity) & (df['Entity'] == person))]
                if temp_df.empty:
                    continue
                # Perform analysis and append results
                bidirectional_results.append(analyze_pair(temp_df, pair))

        # Additional analysis for entities_of_interest
        if entities_of_interest:
            for person in people:
                for entity in entities_of_interest:
                    # Skip pairs that have already been processed
                    pair = tuple(sorted([person, entity]))
                    if pair in processed_pairs:
                        continue
                    processed_pairs.add(pair)

                    # Filter transactions between the person and entity_of_interest
                    temp_df = df[((df['Name'] == person) & (df['Entity'] == entity))]
                    if temp_df.empty:
                        continue

                    # Perform analysis and append results
                    bidirectional_results.append(analyze_pair(temp_df, pair))

        # Convert results to DataFrame
        bidirectional_results_df = pd.DataFrame(bidirectional_results)
        return bidirectional_results_df 
    

    def get_unique_name_acc(self, single_person_output):
        # Create a list of dictionaries with 'Name' and 'Acc Number'
        name_acc_list = [
            {"Name": data["name"], "Acc Number": data["acc_number"]}
            for data in single_person_output.values()
        ]

        # Convert to DataFrame and drop duplicates
        name_acc_df = pd.DataFrame(name_acc_list).drop_duplicates(subset=["Name", "Acc Number"])

        return name_acc_df

    def single_person_sheets(self, dfs, name_dfs):
        result = {}

        # Iterate over the keys of dfs and name_dfs (which are identical)
        for key in dfs:
            # Retrieve the corresponding dataframe
            one_df = dfs[key]

            initial_df = one_df.copy()   
            df = self.commoner.category_add_ca(initial_df)
            df = self.single_name_entity_addition(df)
            df['Value Date'] = pd.to_datetime(df['Value Date'], format='%d-%m-%Y', errors='coerce')
            new_tran_df = self.commoner.another_method(df) 

            fifo = self.single_fifo_analysis(new_tran_df)
            money_trail = self.single_money_trail_analysis(new_tran_df)
            entity_analysis = self.single_analyze_entities(new_tran_df)
            bidirectional_analysis = self.single_bidirectional_analysis(new_tran_df)

            transaction_sheet_df = self.commoner.transaction_sheet(df)
            # transaction_sheet_df['Value Date'] = pd.to_datetime(transaction_sheet_df['Value Date']).dt.strftime('%d-%m-%Y')

            new_tran_df = self.commoner.another_method(transaction_sheet_df)
            eod_sheet_df = self.commoner.eod(df)
            opening_bal, closing_bal = self.commoner.opening_and_closing_bal(eod_sheet_df, df)

            #summary_df_list is a list of dataframes [df1: Paticulars, df2: Income/Receipts, df3: Important Expenses/Payments, df4: Other Expenses/Payments]
            summary_df_list = self.commoner.summary_sheet(df, opening_bal, closing_bal, new_tran_df)
            
            investment_df = self.commoner.total_investment(new_tran_df)
            creditor_df = self.commoner.creditor_list(transaction_sheet_df)
            debtor_df = self.commoner.debtor_list(transaction_sheet_df)
            cash_withdrawal_df = self.commoner.cash_withdraw(new_tran_df)
            cash_deposit_df = self.commoner.cash_depo(new_tran_df)
            dividend_int_df = self.commoner.div_int(new_tran_df)
            emi_df = self.commoner.emi(new_tran_df)
            refund = self.commoner.refund_reversal(new_tran_df)
            suspense_credit_df = self.commoner.suspense_credit(new_tran_df)
            suspense_debit_df = self.commoner.suspense_debit(new_tran_df)

            # Extract the name and account number from name_dfs
            name, acc_number = name_dfs[key]

            # Store the combined information in the result dictionary
            result[key] = {
                "name":name,
                "acc_number":acc_number,
                "data":{
                    "df":df,
                    "new_tran_df":new_tran_df,
                    'summary_df_list': summary_df_list,
                    "fifo":fifo,
                    "money_trail":money_trail,
                    "entity_analysis":entity_analysis,
                    "bidirectional_analysis":bidirectional_analysis,
                    "transaction_sheet_df":transaction_sheet_df,
                    "eod_sheet_df":eod_sheet_df,
                    "investment_df":investment_df,
                    "creditor_df":creditor_df,
                    "debtor_df":debtor_df,
                    "cash_withdrawal_df":cash_withdrawal_df,
                    "cash_deposit_df":cash_deposit_df,
                    "dividend_int_df":dividend_int_df,
                    "emi_df":emi_df,
                    "refund":refund,
                    "suspense_credit_df":suspense_credit_df,
                    "suspense_debit_df":suspense_debit_df
                }
            }
        # Return the resulting dictionary
        return result

    def get_unique_entity_frequency(self, process_df):
        # Group by 'Entity' and calculate frequency
        entity_freq_df = process_df['Entity'].value_counts().reset_index()

        # Rename columns
        entity_freq_df.columns = ['Entity', 'Frequency (total no of times entity occurred in process_df)']

        return entity_freq_df

    
    def extract_unique_names_and_entities(self, df):
        # Get unique values from 'Name' column
        unique_names = df['Name'].dropna().unique()
        
        # Get unique values from 'Entity' column
        unique_entities = df['Entity'].dropna().unique()
        
        # Combine both lists and remove duplicates
        unique_combined = list(set(unique_names) | set(unique_entities))
        
        return unique_combined


    def group_similar_entities(self, entities, low=0.52, high=1.0, weight_semantic=0.6, weight_string=0.4):
        structure_model = SentenceTransformer(os.path.dirname(os.path.abspath(__file__))+"./matching_names_model")

        n = len(entities)
        grouped = defaultdict(list)
        visited = set()

        # Precompute embeddings for semantic similarity
        embeddings = model.encode(entities, convert_to_tensor=True)

        for i in range(n):
            for j in range(i + 1, n):
                # Calculate semantic similarity
                semantic_score = util.cos_sim(embeddings[i], embeddings[j]).item()

                # Calculate string similarity
                string_score = fuzz.ratio(entities[i].lower(), entities[j].lower()) / 100

                # Combine the scores
                combined_score = (weight_semantic * semantic_score) + (weight_string * string_score)

                if low <= combined_score < high and (entities[i], entities[j]) not in visited:
                    grouped[entities[i]].append(entities[j])
                    visited.add((entities[i], entities[j]))
                    visited.add((entities[j], entities[i]))

        # Consolidate groups into lists
        result = []
        for key, value in grouped.items():
            result.append([key] + value)

        # Merge overlapping groups
        merged = []
        while result:
            first, *rest = result
            first = set(first)
            overlapping = [g for g in rest if first & set(g)]
            for g in overlapping:
                first |= set(g)
            rest = [g for g in rest if not (first & set(g))]
            merged.append(list(first))
            result = rest

        return merged


    def replace_entities(self, df, lists_of_names):
        """
        Efficiently replaces values in the 'Entity' column based on a list of lists of names.
        """
        # Create a set of unique names in the 'Name' column for fast lookup
        name_set = set(df['Name'].dropna())

        # Precompute the most frequent value for each group in 'Entity'
        entity_counts = df['Entity'].value_counts()

        # Iterate over each group in the list of lists
        replacements = {}
        for name_group in lists_of_names:
            # Check if any name in the group exists in the 'Name' column
            anchor_name = next((name for name in name_group if name in name_set), None)

            if anchor_name:
                # Use the anchor name for replacements
                replacements.update({name: anchor_name for name in name_group})
            else:
                # Find the most frequent name in 'Entity' from the group
                most_frequent_name = (
                    entity_counts[entity_counts.index.isin(name_group)].idxmax()
                    if not entity_counts[entity_counts.index.isin(name_group)].empty
                    else name_group[0]  # Default to the first name
                )
                replacements.update({name: most_frequent_name for name in name_group})

        # Apply the replacements to the 'Entity' column
        df['Entity'] = df['Entity'].replace(replacements)

        return df

    def cummalative_person_sheets(self, process_df, name_acc_df):

        del_folder_path = os.path.join(BASE_DIR, "del")
        if not os.path.exists(del_folder_path):
            os.makedirs(del_folder_path)



        #@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@

        #unique_values = self.extract_unique_names_and_entities(process_df)

        #group_of_similar_entities = self.group_similar_entities(unique_values)

        #new_df = self.replace_entities(process_df, group_of_similar_entities)

        #@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@


        # save the process_df to excel file in del folder of current directory make sure to include base dir
        process_df.to_excel(BASE_DIR + "/del/process_df.xlsx", index=False)

        
        name_acc_df.to_excel(BASE_DIR + f"/del/name_acc_df.xlsx", index=False)
        

        #entity table
        entity_df = self.get_unique_entity_frequency(process_df)
        entity_df.to_excel(BASE_DIR + f"/del/entity_df.xlsx", index=False)

        # link analysis data
        link_analysis_df = process_df.groupby(['Name', 'Entity']).agg(Total_Credit=('Credit', 'sum'), Total_Debit=('Debit', 'sum')).reset_index()
        link_analysis_df = link_analysis_df[['Name', 'Total_Credit', 'Total_Debit', 'Entity']]
        link_analysis_df = link_analysis_df[~((link_analysis_df['Total_Credit'] == 0) & (link_analysis_df['Total_Debit'] == 0))]
        link_analysis_df['highlight'] = 0
        unique_names = link_analysis_df['Name'].unique()
        link_analysis_df.loc[link_analysis_df['Entity'].isin(unique_names), 'highlight'] = 1

        #fifo analysis {utizialion sentence: {lifo: lifo_df, fifo: fifo_df}}
        fifo_dictionary = self.lifo_fifo(process_df)

        # fifo_daily.to_excel(BASE_DIR + f"/del/fifo_daily.xlsx", index=False)
        # fifo_weekly.to_excel(BASE_DIR + f"/del/fifo_weekly.xlsx", index=False)
        # fifo_monthly.to_excel(BASE_DIR + f"/del/fifo_monthly.xlsx", index=False)
        # fifo_half_yearly.to_excel(BASE_DIR + f"/del/fifo_half_yearly.xlsx", index=False)
        # fifo_yearly.to_excel(BASE_DIR + f"/del/fifo_yearly.xlsx", index=False)


        #fund flow/money trail
        # ff_daily_analysis = self.analyze_period(process_df, 'D')
        # ff_weekly_analysis = self.analyze_period(process_df, 'W')
        # ff_monthly_analysis = self.analyze_period(process_df, 'M')
        # ff_half_yearly_analysis = self.analyze_period(process_df, '6M')
        # ff_yearly_analysis = self.analyze_period(process_df, 'Y')
        # ff_monthly_analysis.to_excel(BASE_DIR + f"/del/ff_monthly_analysis.xlsx", index=False)
        # ff_half_yearly_analysis.to_excel(BASE_DIR + f"/del/ff_half_yearly_analysis.xlsx", index=False)
        # ff_yearly_analysis.to_excel(BASE_DIR + f"/del/ff_yearly_analysis.xlsx", index=False)

        #bidirectional_analysis
        bda_all_analysis = self.cumulative_bidirectional_analysis(process_df)

        print("********************************************************************************************")

        # Return the concatenated DataFrame
        return {
            "process_df":process_df,
            "name_acc_df":name_acc_df,
            "entity_df":entity_df,
            "link_analysis_df": link_analysis_df,
            "fifo": fifo_dictionary,
            "bidirectional_analysis": bda_all_analysis
        }
