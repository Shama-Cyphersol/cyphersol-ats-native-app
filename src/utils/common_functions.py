import os
from dateutil import parser
# from django.conf import settings
from openpyxl.styles import Font
from pypdf import PdfReader
import shutil
import io
import openpyxl
from openpyxl.styles import Alignment
import time
import pandas as pd
import numpy as np
import PyPDF2
from PyPDF2 import PdfReader, PdfWriter
import pdfplumber
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import regex as re
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
import datefinder
from calendar import monthrange
import calendar
from openpyxl.styles import Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
from openpyxl.styles import Border, Side
from openpyxl.styles import Font
import logging
import openpyxl
from openpyxl.styles import Alignment
import re

bold_font = Font(bold=True)
pd.options.display.float_format = "{:,.2f}".format
pd.set_option("display.max_columns", None)
pd.set_option("display.max_rows", None)
pd.set_option("display.width", None)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
#from old_bank_extractions import CustomStatement
from ..utils.code_for_extraction import ExtractionOnly
from ..utils.model_loader import model_loader


class CommonFunctions:
    def __init__(
        self, bank_names, pdf_paths, pdf_passwords, start_date, end_date, CA_ID
    ):
        self.writer = None
        self.bank_names = bank_names
        self.pdf_paths = pdf_paths
        self.pdf_passwords = pdf_passwords
        self.start_date = start_date
        self.end_date = end_date
        self.account_number = ""
        self.file_name = None
        self.CA_ID = CA_ID
        self.extractor = ExtractionOnly(bank_names, pdf_paths, pdf_passwords, CA_ID)
        self._model = model_loader.get_model()

    ##EXTRACTION PROCESS
    def extract_text_from_file(self, file_path):

        file_extension = os.path.splitext(file_path)[1].lower()

        if file_extension == ".csv":
            df = pd.read_csv(file_path)
            text = df.head(20).to_string(index=False)

        else:
            df = pd.read_excel(file_path)
            text = df.head(20).to_string(index=False)

        return text

    def add_start_n_end_date(self, df, start_date, end_date, bank):
        # df["Balance"] = pd.to_numeric(df["Balance"], errors="coerce")
        # df["Debit"] = pd.to_numeric(df["Debit"], errors="coerce")
        # df["Credit"] = pd.to_numeric(df["Credit"], errors="coerce")
        #
        # # Check if the period falls within the start and end dates
        # start_date_sd = pd.to_datetime(start_date, format="%d-%m-%Y", errors="coerce")
        # end_date_ed = pd.to_datetime(end_date, format="%d-%m-%Y", errors="coerce")
        # period_start = pd.to_datetime(
        #     df["Value Date"].iloc[0], format="%d-%m-%Y", errors="coerce"
        # )
        # period_end = pd.to_datetime(
        #     df["Value Date"].iloc[-1], format="%d-%m-%Y", errors="coerce"
        # )
        #
        # if (start_date_sd - timedelta(days=1)) <= period_start <= (
        #     end_date_ed + timedelta(days=1)
        # ) and (start_date_sd - timedelta(days=1)) <= period_end <= (
        #     end_date_ed + timedelta(days=1)
        # ):
        #     print("The period falls within the start and end dates.")
        # else:
        #     raise ExtractionError(
        #         f"The period for Bank: {bank} does not fall within the start and end dates."
        #     )
        #
        # # add opening and closing balance
        # start_bal = (
        #     df.iloc[0]["Balance"] - df.iloc[0]["Credit"]
        #     if df.iloc[0]["Credit"] > 0
        #     else df.iloc[0]["Balance"] + df.iloc[0]["Debit"]
        # )
        # end_bal = df.iloc[-1]["Balance"]
        #
        # start_row = pd.DataFrame(
        #     [
        #         {
        #             "Value Date": start_date,
        #             "Description": "Opening Balance",
        #             "Debit": 0,
        #             "Credit": 0,
        #             "Balance": start_bal,
        #         }
        #     ]
        # )
        # end_row = pd.DataFrame(
        #     [
        #         {
        #             "Value Date": end_date,
        #             "Description": "Closing Balance",
        #             "Debit": 0,
        #             "Credit": 0,
        #             "Balance": end_bal,
        #         }
        #     ]
        # )
        #
        # idf = pd.concat([start_row, df, end_row], ignore_index=True)
        df["Bank"] = f"{bank}"
        return df

    def extract_account_details(self, text):
        y = text
        return ["_", "XXXXXXXXXX"]

    def extract_extension(self, path):
        # Get the file extension
        # print("Path:", path)
        _, extension = os.path.splitext(path)
        extension = extension.lower()

        # Dictionary mapping file extensions to pandas read functions
        # read_functions = {
        #     '.xlsx': pd.read_excel,
        #     '.xls': pd.read_excel,
        #     '.xlsm': pd.read_excel,
        #     '.xlsb': pd.read_excel,
        #     '.csv': pd.read_csv,
        #     '.xltx': pd.read_excel,
        #     '.xltm': pd.read_excel
        # }

        # # Check if the extension is supported
        # if extension not in read_functions:
        #     raise ValueError(f"Unsupported file extension: {extension}")

        return extension
    
    def extraction_process(self, bank, pdf_path, pdf_password, start_date, end_date):
        
        # Default return values for error handling
        empty_idf = pd.DataFrame()
        default_name_n_num = ["_", "XXXXXXXXXX"]

        try:
            bank = re.sub(r"\d+", "", bank)
            # timestamp = int(timezone.localtime().timestamp())
            CA_ID = self.CA_ID
            ext = self.extract_extension(pdf_path)

            if ext == ".pdf":
                try:
                    idf, text = self.extractor.extract_with_test_cases(bank, pdf_path, pdf_password, CA_ID)
                    name_n_num = self.extract_account_details(text)
                except Exception as e:
                    # Handle PDF extraction failure
                    return empty_idf, default_name_n_num

            elif ext == ".csv":
                try:
                    df = pd.read_csv(pdf_path)
                    df.loc[0] = df.columns
                    df.columns = range(df.shape[1])
                    start_index = df.apply(lambda row: (row.astype(str).str.contains("date", case=False).any() and
                                                    row.astype(str).str.contains("balance|total amount",
                                                                                    case=False).any()) or
                                                    row.astype(str).str.contains("balance|total amount", case=False).any(),
                                        axis=1).idxmax()
                    df = df.loc[start_index:] if start_index is not None else pd.DataFrame()
                    idf = self.extractor.model_for_pdf(df)
                    name_n_num = self.extract_account_details(self.extract_text_from_file(pdf_path))
                except Exception as e:
                    # Handle CSV extraction failure
                    return empty_idf, default_name_n_num

            else:
                try:
                    df = pd.read_excel(pdf_path)
                    df.loc[0] = df.columns
                    df.columns = range(df.shape[1])
                    start_index = df.apply(lambda row: (row.astype(str).str.contains("date", case=False).any() and
                                                    row.astype(str).str.contains("balance|total amount",
                                                                                    case=False).any()) or
                                                    row.astype(str).str.contains("balance|total amount", case=False).any(),
                                        axis=1).idxmax()
                    df = df.loc[start_index:] if start_index is not None else pd.DataFrame()
                    idf = self.extractor.model_for_pdf(df)
                    name_n_num = self.extract_account_details(self.extract_text_from_file(pdf_path))
                except Exception as e:
                    # Handle Excel extraction failure
                    return empty_idf, default_name_n_num

            # Add start and end date
            idf = self.add_start_n_end_date(idf, start_date, end_date, bank)

        except Exception as e:
            # Catch all other errors
            return empty_idf, default_name_n_num

        return idf, name_n_num
    ##EOD
    def monthly(self, df):
        # add a new row with the average of month values in each column
        e = df.copy()
        e.replace(0, np.nan, inplace=True)
        new_row = pd.DataFrame(e.iloc[0:31].mean(axis=0)).T.round(2)
        monthly_avg = pd.concat([df, new_row], ignore_index=True)
        monthly_avg.iloc[-1, 0] = "Average"
        return monthly_avg

    def eod(self, df):
        df["Value Date"] = pd.to_datetime(
            df["Value Date"], format="%d-%m-%Y", errors="coerce"
        )
        df["Month"] = df["Value Date"].dt.strftime("%b-%Y")
        df["Date"] = df["Value Date"].dt.day
        end_day = df["Date"].iloc[-1]
        df = df[["Value Date", "Balance", "Month", "Date", "Bank"]]
        df = df[df["Balance"] != ""]
        bank_names = df["Bank"].unique().tolist()
        multiple_eods = []

        for bank in bank_names:
            idf = df[df["Bank"] == bank]
            result_eod = pd.DataFrame()
            for month in idf["Month"].unique():
                eod_month_df = idf.loc[idf["Month"] == month].drop_duplicates(
                    subset="Date", keep="last"
                )

                # Loop through each day in the month
                for day in range(1, 32):
                    # Check if there are any rows with the current day
                    day_present = False
                    for index, row in eod_month_df.iterrows():
                        if row["Date"] == day:
                            day_present = True
                            break

                    # If day is not present, add a row with NaN values for all columns except the date
                    if not day_present:
                        new_row = {
                            "Balance": 0,
                            "Month": eod_month_df.iloc[0]["Month"],
                            "Date": day,
                        }
                        eod_month_df = pd.concat(
                            [eod_month_df, pd.DataFrame(new_row, index=[0])],
                            ignore_index=True,
                        )
                        eod_month_df = eod_month_df.sort_values(by="Date")

                result_eod = pd.concat([result_eod, eod_month_df], ignore_index=True)

            # iterate through column and replace zeros with previous value
            previous_eod = 0
            for i, value in enumerate(result_eod["Balance"]):
                if value == 0:
                    result_eod.loc[i, "Balance"] = previous_eod
                else:
                    previous_eod = value

            pivot_df = result_eod.pivot(
                index="Date", columns="Month", values="Balance"
            ).reset_index(drop=True)
            column_order = idf["Month"].unique()  # do not change
            pivot_df = pivot_df.reindex(columns=column_order)
            pivot_df.insert(0, "Day", range(1, 32))

            columns = pivot_df.columns[1:]
            col_values = [
                "Feb",
                "Apr",
                "Jun",
                "Sep",
                "Nov",
            ]  # no hard code now :: these are the months in every year not having 31 days

            for i, row in pivot_df.iterrows():
                for month in columns:
                    if any(col in month for col in col_values):
                        y = month.split("-")[1]
                        is_leap = int(y) % 4 == 0 and (
                            int(y) % 100 != 0 or int(y) % 400 == 0
                        )

                        if "Feb" in month and is_leap and row["Day"] > 29:
                            pivot_df.loc[i, month] = 0.0
                        elif "Feb" in month and not is_leap and row["Day"] > 28:
                            pivot_df.loc[i, month] = 0.0
                        elif row["Day"] > 30:
                            pivot_df.loc[i, month] = 0.0

            if end_day != 31.0:
                end_day = int(end_day)
                last_column_list = pivot_df.iloc[:, -1].tolist()
                new_column = last_column_list.copy()
                new_column[end_day:] = [0] * (len(new_column) - end_day)
                pivot_df.iloc[:, -1] = new_column

            multiple_eods.append(pivot_df)

            if len(multiple_eods) < 1:
                adf = multiple_eods[0]
                # add a new row with the sum of values in each column
                new_row = pd.DataFrame(adf.iloc[0:31].sum(axis=0)).T
                total_df = pd.concat([adf, new_row], ignore_index=True)
                total_df.iloc[-1, 0] = "Total"
                all_df = self.monthly(total_df)
            else:
                adf = self.process_repeating_columns(multiple_eods)
                # add a new row with the sum of values in each column
                new_row = pd.DataFrame(adf.iloc[0:31].sum(axis=0)).T
                total_df = pd.concat([adf, new_row], ignore_index=True)
                total_df.iloc[-1, 0] = "Total"
                all_df = self.monthly(total_df)
                # qur = self.QuaterlyAvg(total_df)
                # half = self.HalflyAvg(total_df)
                # YearlyAvg = self.YearlyAvg(total_df)
                # # qur.to_excel(self.writer, sheet_name='Qtrly AVG Bal', index=False)
                # # half.to_excel(self.writer, sheet_name='Half Yrly AVG Bal', index=False)
                # # YearlyAvg.to_excel(self.writer, sheet_name='Yrly AVG Bal', index=False)

                # print(total_df)
                all_df = self.monthly(total_df)
                # HalflyAvg = self.HalflyAvg(total_df)

        # all_df.iloc[:-4]["Day"] = all_df.iloc[:-4]["Day"].astype(int)

        return all_df

    def opening_and_closing_bal(self, edf, df):
        closing_bal = {}
        for column in edf.columns[1:]:
            non_zero_rows = edf.loc[edf[column] != 0].iloc[:-2]
            if not non_zero_rows.empty:
                last_non_zero_row = non_zero_rows.iloc[-1]
                closing_bal[column] = last_non_zero_row[column]

        opening_bal_value_1 = df.iloc[0]["Balance"]
        keys_list = list(closing_bal.keys())
        values_list = list(closing_bal.values())
        values_list.insert(0, opening_bal_value_1)
        values_list.pop()
        opening_bal = dict(zip(keys_list, values_list))

        return opening_bal, closing_bal

    def avgs_df(self, df):
        # quarterly_avg
        if df.shape[1] > 3:
            df_chi_list_1 = []
            # Iterate through every three columns in the original DataFrame
            for i in range(1, df.shape[1], 3):
                # Get the current three columns
                subset = df.iloc[:, i : i + 3]
                if subset.shape[1] < 3:
                    new_row = 0.0
                else:
                    new_row = subset.iloc[-2].sum() / 3
                subset.loc[len(subset)] = new_row
                df_chi_list_1.append(subset)
            result_df = pd.concat(df_chi_list_1, axis=1)
            new_row = pd.Series({"Day": "Quarterly Average"})
            df = df._append(new_row, ignore_index=True)
            result_df.insert(0, "Day", df["Day"])
            df = result_df

            # half - yearly avg
            if df.shape[1] > 6:
                df_chi_list_2 = []
                # Iterate through every three columns in the original DataFrame
                for i in range(1, df.shape[1], 6):
                    # Get the current three columns
                    subset = df.iloc[:, i : i + 6]
                    if subset.shape[1] < 6:
                        new_row = 0.0
                    else:
                        new_row = subset.iloc[-3].sum() / 6
                    subset.loc[len(subset)] = new_row
                    df_chi_list_2.append(subset)
                result_df = pd.concat(df_chi_list_2, axis=1)
                new_row = pd.Series({"Day": "Half-Yearly Average"})
                df = df._append(new_row, ignore_index=True)
                result_df.insert(0, "Day", df["Day"])
                df = result_df

                # yearly avg
                if df.shape[1] > 12:
                    df_chi_list_3 = []
                    # Iterate through every three columns in the original DataFrame
                    for i in range(1, df.shape[1], 12):
                        # Get the current three columns
                        subset = df.iloc[:, i : i + 12]
                        if subset.shape[1] < 12:
                            new_row = 0.0
                        else:
                            new_row = subset.iloc[-4].sum() / 12
                        subset.loc[len(subset)] = new_row
                        df_chi_list_3.append(subset)
                    result_df = pd.concat(df_chi_list_3, axis=1)
                    new_row = pd.Series({"Day": "Yearly Average"})
                    df = df._append(new_row, ignore_index=True)
                    result_df.insert(0, "Day", df["Day"])
                    df = result_df

                else:
                    new_row = pd.Series({"Day": "Yearly Average"})
                    df = df._append(new_row, ignore_index=True)

            else:
                new_row = pd.Series({"Day": "Half-Yearly Average"})
                df = df._append(new_row, ignore_index=True)

        else:
            new_row = pd.Series({"Day": "Quarterly Average"})
            df = df._append(new_row, ignore_index=True)

        return df

    def calculate_fixed_day_average(self, data):
        e = data.copy()
        e.replace(0, np.nan, inplace=True)

        # Check if the last column (assumed to be January of the next year) contains any zeros
        last_column = data.columns[-1]
        if (data[last_column] == 0).any():
            data.drop(
                last_column, axis=1, inplace=True
            )  # This removes the entire column

        # Calculate an overall average for the first 31 rows and format it
        new_row = pd.DataFrame(data.iloc[0:31].mean(axis=0)).T.round(2)
        new_row_string = new_row.applymap("{:.2f}".format)
        Average = new_row_string.applymap(lambda x: float(x))

        # Adding a label for the overall average
        Average["Day"] = ["Daily_Avg"]
        Average = Average[["Day"] + [col for col in Average.columns if col != "Day"]]

        # Define sets of days for which to calculate averages
        sets_of_days = [
            {"days": [5, 15, 25], "label": "Avg_Days_5_15_25"},
            {"days": [5, 10, 15, 25], "label": "Avg_Days_5_10_15_25"},
            {"days": [1, 5, 10, 15, 20, 25], "label": "Avg_Days_1_5_10_15_20_25"},
            {"days": [8, 10, 15, 20, 25], "label": "Avg_Days_8_10_15_20_25"},
            {"days": [1, 5, 10, 15, 20], "label": "Avg_Days_1_5_10_15_20"},
            {"days": [5, 10, 15, 20, 25], "label": "Avg_Days_5_10_15_20_25"},
            {"days": [1, 7, 14, 21, 28], "label": "Avg_Days_1_7_14_21_28"},
            {"days": [1, 5, 10, 15, 25], "label": "Avg_Days_1_5_10_15_25"},
            {"days": [5, 15, 25, 30], "label": "Avg_Days_5_15_25_30"},
            {"days": [2, 4, 10, 17, 21], "label": "Avg_Days_2_4_10_17_25"},
            {"days": [5, 15, 25, 30], "label": "Avg_Days_5_15_25_30"},
        ]

        avg_balance_df_list = []
        for day_set in sets_of_days:
            selected_days_df = data[data["Day"].isin(day_set["days"])]
            # Ensure columns are numeric
            for col in selected_days_df.columns[
                1:
            ]:  # Assuming the first column is 'Day'
                selected_days_df[col] = pd.to_numeric(
                    selected_days_df[col], errors="coerce"
                )
            selected_days_df = selected_days_df.iloc[:, 1:].fillna(
                0
            )  # Handle NaN values if any
            # Calculate mean and round off
            average_balance = selected_days_df.mean(axis=0).round(2)
            average_balance_df = pd.DataFrame(average_balance).T
            average_balance_df["Day"] = [day_set["label"]]
            average_balance_df = average_balance_df[
                ["Day"] + [col for col in average_balance_df.columns if col != "Day"]
            ]
            avg_balance_df_list.append(average_balance_df)
        # Concatenate all averages into a single DataFrame, adding the overall average to the top
        all_avg_balances = pd.concat([Average] + avg_balance_df_list, ignore_index=True)
        # Now you can call the calculate_monthly_averages function with the all_avg_balances DataFrame
        averages_with_monthly = self.calculate_monthly_averages(all_avg_balances)
        # You should decide whether you want to return all_avg_balances or averages_with_monthly
        # Here we assume you want to return the averages with monthly data included
        return averages_with_monthly

    def process_avg_last_6_months(self, filename, data, eod):
        error_df = pd.DataFrame()

        # Check if 'Avg_Last_6_Months' column exists
        if "Avg_Last_6_Months" not in data.columns:
            print("'Avg_Last_6_Months' column not found in DataFrame.")
            return error_df  # Or return an empty DataFrame as needed

        # Extract the second value from 'Avg_Last_6_Months' and perform calculations
        if "Avg_Last_6_Months" in data.columns and len(data["Avg_Last_6_Months"]) > 1:
            if not pd.isna(data["Avg_Last_6_Months"].iloc[1]):
                avg_divided_by_1_5 = data["Avg_Last_6_Months"].iloc[1] / 1.5
            else:
                avg_divided_by_1_5 = np.nan
        else:
            print(
                "'Avg_Last_6_Months' column does not exist or does not have enough data."
            )
            avg_divided_by_1_5 = np.nan

        # For 'Avg_Last_12_Months' at index 0
        if "Avg_Last_12_Months" in data.columns:
            if not pd.isna(data["Avg_Last_12_Months"].iloc[0]):
                avg_divided_by_2_idfc = data["Avg_Last_12_Months"].iloc[0] / 2
            else:
                avg_divided_by_2_idfc = np.nan
        else:
            print("'Avg_Last_12_Months' column does not exist in the DataFrame.")
            avg_divided_by_2_idfc = np.nan

        # For 'Avg_Last_6_Months' at index 2
        if "Avg_Last_6_Months" in data.columns and len(data["Avg_Last_6_Months"]) > 2:
            if not pd.isna(data["Avg_Last_6_Months"].iloc[2]):
                avg_divided_by_2_indus = data["Avg_Last_6_Months"].iloc[2] / 1.5
            else:
                avg_divided_by_2_indus = np.nan
        else:
            print(
                "'Avg_Last_6_Months' column does not exist or does not have enough data."
            )
            avg_divided_by_2_indus = np.nan

        # For 'Avg_Last_12_Months' at index 0 again
        if "Avg_Last_12_Months" in data.columns:
            if not pd.isna(data["Avg_Last_12_Months"].iloc[0]):
                avg_divided_by_2_L_T = data["Avg_Last_12_Months"].iloc[0] / 2
            else:
                avg_divided_by_2_L_T = np.nan
        else:
            print("'Avg_Last_12_Months' column does not exist in the DataFrame.")
            avg_divided_by_2_L_T = np.nan

        annual_interest_rate = 0.0870
        periods = 20 * 12
        principal = 100000
        payment_value = self.pmt(principal, annual_interest_rate, periods)
        payment_for_lap = self.pmt_lap()
        payment_for_bl = self.pmt_bl()

        # Calculating Loan value for axis
        axis_home_loan_value = None
        if payment_value != 0:
            axis_home_loan_value = avg_divided_by_1_5 / payment_value
            axis_home_loan_value = axis_home_loan_value * 100000
            axis_home_loan_value = round(axis_home_loan_value, 2)

        else:
            print("PMT calculation resulted in zero, cannot divide by zero.")

        axis_LAP_value = None
        if payment_for_lap != 0:
            axis_LAP_value = avg_divided_by_1_5 / payment_for_lap
            axis_LAP_value = axis_LAP_value * 100000
        else:
            print("PMT calculation resulted in zero, cannot divide by zero.")

        axis_bl_value = None
        if payment_for_bl != 0:
            axis_bl_value = avg_divided_by_1_5 / payment_for_bl
            axis_bl_value = axis_bl_value / payment_for_lap
            axis_bl_value = axis_bl_value * 100000
        else:
            print("PMT calculation resulted in zero, cannot divide by zero.")

        # Calculating loan value for Idfc
        Idfc_home_loan_value = None
        if payment_value != 0:
            Idfc_home_loan_value = avg_divided_by_2_idfc / payment_value
            Idfc_home_loan_value = Idfc_home_loan_value * 100000
            Idfc_home_loan_value = round(Idfc_home_loan_value, 2)

        else:
            print("PMT calculation resulted in zero, cannot divide by zero.")

        Idfc_LAP_value = None
        if payment_for_lap != 0:
            Idfc_LAP_value = avg_divided_by_2_idfc / payment_for_lap
            Idfc_LAP_value = Idfc_LAP_value * 100000
            Idfc_LAP_value = round(Idfc_LAP_value, 2)

        else:
            print("PMT calculation resulted in zero, cannot divide by zero.")

        Idfc_bl_value = None
        if payment_for_bl != 0:
            Idfc_bl_value = avg_divided_by_2_idfc / payment_for_bl
            Idfc_bl_value = Idfc_bl_value * 100000
            Idfc_bl_value = round(Idfc_bl_value, 2)
        else:
            print("PMT calculation resulted in zero, cannot divide by zero.")

        # Calculating loan value for Indus
        indus_home_loan_value = None
        if payment_value != 0:
            indus_home_loan_value = avg_divided_by_2_indus / payment_value
            indus_home_loan_value = indus_home_loan_value * 100000
            indus_home_loan_value = round(indus_home_loan_value, 2)

        else:
            print("PMT calculation resulted in zero, cannot divide by zero.")

        indus_LAP_value = None
        if payment_for_lap != 0:
            indus_LAP_value = avg_divided_by_2_indus / payment_for_lap
            indus_LAP_value = indus_LAP_value * 100000
            indus_LAP_value = round(indus_LAP_value, 2)
        else:
            print("PMT calculation resulted in zero, cannot divide by zero.")

        indus_bl_value = None
        if payment_for_bl != 0:
            indus_bl_value = avg_divided_by_2_indus / payment_for_bl
            indus_bl_value = indus_bl_value * 100000
            indus_bl_value = round(indus_bl_value, 2)
        else:
            print("PMT calculation resulted in zero, cannot divide by zero.")

        L_T_home_loan_value = None
        if payment_value != 0:
            L_T_home_loan_value = avg_divided_by_2_L_T / payment_value
            L_T_home_loan_value = L_T_home_loan_value * 100000
            L_T_home_loan_value = round(L_T_home_loan_value, 2)
        else:
            print("PMT calculation resulted in zero, cannot divide by zero.")
        L_T_LAP_value = None
        if payment_for_lap != 0:
            L_T_LAP_value = avg_divided_by_2_L_T / payment_for_lap
            L_T_LAP_value = L_T_LAP_value * 100000
            L_T_LAP_value = round(L_T_LAP_value, 2)
        else:
            print("PMT calculation resulted in zero, cannot divide by zero.")

        L_T_bl_value = None
        if payment_for_lap != 0:
            L_T_bl_value = avg_divided_by_2_L_T / payment_for_bl
            L_T_bl_value = L_T_bl_value * 100000
            L_T_bl_value = round(L_T_bl_value, 2)
        else:
            print("PMT calculation resulted in zero, cannot divide by zero.")

        home_loan_values = []

        # Append home loan values to the list
        if axis_home_loan_value is not None:
            home_loan_values.append(axis_home_loan_value)

        if Idfc_home_loan_value is not None:
            home_loan_values.append(Idfc_home_loan_value)

        if indus_home_loan_value is not None:
            home_loan_values.append(indus_home_loan_value)

        if L_T_home_loan_value is not None:
            home_loan_values.append(L_T_home_loan_value)

        # Calculate the maximum home loan value
        max_home_loan_value = (
            max(home_loan_values, default=None) if home_loan_values else None
        )

        lap_values = []
        bl_values = []

        if axis_LAP_value is not None:
            lap_values.append(axis_LAP_value)

        if Idfc_LAP_value is not None:
            lap_values.append(Idfc_LAP_value)

        if indus_LAP_value is not None:
            lap_values.append(indus_LAP_value)

        if L_T_LAP_value is not None:
            lap_values.append(L_T_LAP_value)

        if axis_bl_value is not None:
            bl_values.append(axis_bl_value)

        if Idfc_bl_value is not None:
            bl_values.append(Idfc_bl_value)

        if indus_bl_value is not None:
            bl_values.append(indus_bl_value)

        if L_T_bl_value is not None:
            bl_values.append(L_T_bl_value)

        # Calculate the maximum values
        max_lap_value = max(lap_values, default=None) if lap_values else None
        max_bl_value = max(bl_values, default=None) if bl_values else None

        # Create DataFrame with maximum values
        max_values_df = pd.DataFrame(
            {
                "Maximum Home Loan Value": [max_home_loan_value],
                "Maximum LAP Value": [max_lap_value],
                "Maximum BL Value": [max_bl_value],
            }
        )

        return max_values_df

    def process_repeating_columns(self, oy):
        df = pd.concat(oy, axis=1)
        df = df.loc[:, ~df.columns.duplicated(keep="first") | (df.columns != "Day")]
        repeating_columns = [
            col for col in df.columns if df.columns.tolist().count(col) > 1
        ]

        idf = pd.DataFrame(
            {col: df[col].sum(axis=1).round(4) for col in repeating_columns}
        )
        df = df.drop(columns=repeating_columns)
        concatenated_df = pd.concat([df, idf], axis=1)

        sorted_columns = sorted(
            [col for col in concatenated_df.columns if col != "Day"],
            key=lambda x: pd.to_datetime(x, format="%b-%Y"),
        )
        sorted_columns_formatted = [
            col.strftime("%b-%Y") if isinstance(col, pd.Timestamp) else col
            for col in sorted_columns
        ]
        concatenated_df = concatenated_df[["Day"] + sorted_columns_formatted]
        return concatenated_df

    def calculate_monthly_averages(self, data):
        original_columns = data.columns.tolist()

        # Calculate the average for the last 12 months
        if len(original_columns) >= 12:
            last_12_months = original_columns[-12:]
            if data[last_12_months].applymap(np.isreal).all().all():
                data["Avg_Last_12_Months"] = data[last_12_months].mean(axis=1).round(2)
            else:
                print(
                    "Non-numeric data found in last 12 months columns, 'Avg_Last_12_Months' will not be added."
                )

        # Update the columns list to exclude the new average column
        updated_columns = data.columns.tolist()

        # Calculate the average for the last 6 months
        if len(original_columns) >= 6:
            last_6_months = original_columns[-6:]
            if data[last_6_months].applymap(np.isreal).all().all():
                data["Avg_Last_6_Months"] = data[last_6_months].mean(axis=1).round(2)
            else:
                print(
                    "Non-numeric data found in last 6 months columns, 'Avg_Last_6_Months' will not be added."
                )

        # Similarly, update the columns list again to exclude the new average column
        updated_columns = data.columns.tolist()

        # Calculate the average for the last 18 months
        if len(original_columns) >= 18:
            last_18_months = original_columns[-18:]
            if data[last_18_months].applymap(np.isreal).all().all():
                data["Avg_Last_18_Months"] = data[last_18_months].mean(axis=1).round(2)
            else:
                print(
                    "Non-numeric data found in last 18 months columns, 'Avg_Last_18_Months' will not be added."
                )
        return data

    def pmt(self, principal, annual_interest_rate, periods):
        r = annual_interest_rate / 12  # monthly interest rate
        n = periods
        payment = principal * r / (1 - (1 + r) ** -n)
        # print("pmt", payment)
        return payment

    def pmt_lap(self):
        annual_interest_rate = 0.0950
        periods = 15 * 12
        principal = 100000
        r = annual_interest_rate / 12  # monthly interest rate
        n = periods
        payment = principal * r / (1 - (1 + r) ** -n)
        # print("pmt_lap", payment)
        return payment

    def pmt_bl(self):
        annual_interest_rate = 0.2200
        periods = 3 * 12
        principal = 100000
        r = annual_interest_rate / 12  # monthly interest rate
        n = periods
        payment = principal * r / (1 - (1 + r) ** -n)
        # print("pmt_bl", payment)
        return payment
        
    def preprocess_df(self,df):
        # Convert columns to lowercase
        for col in df.columns:
            if df[col].dtype == "object":
                df[col] = df[col].str.lower()
        # Remove spaces from 'Description'
        df["Description"] = df["Description"].str.replace(" ", "")
        # Define keywords to exclude
        exclude_keywords = [
            # Major Indian Banks
            "sbi", "statebankofindia", "statebank",
            "pnb", "punjabnationalbank",
            "bob", "bankofbaroda",
            "canara", "canarabank",
            "unionbank", "unionbankofindia",
            "boi", "bankofindia", "indianbank",
            "centralbank", "centralbankofindia",
            "iob", "indianoverseasbank", "indianoverseas",
            "uco", "ucobank",
            "bankofmaharashtra", "maharashtrabank",
            "psb", "punjabandsindbank",
            "hdfc", "hdfcbank", "hdfcbankltd", "hdfcbanklimited",
            "icici", "icic", "icicibank", "icicibankltd", "icicibanklimited",
            "axis", "axisbank", "axisbankltd", "axisbanklimited",
            "kotak", "kotakbank", "kotakmahindrabank",
            "yesbank", "yesbankltd", "yesbanklimited",
            "idbi", "idbibank", "idbibankltd", "idbibanklimited",
            "idfc", "idfcfirstbank", "idfcfirst",
            "indusind", "indusindbank",
            "bandhan", "bandhanbank",
            "rblbank",
            "dhanlaxmi", "dhanlaxmibank",
            "federal", "federalbank",
            "southindianbank",
            "karurvysyabank", "karurvysya",
            "cityunionbank",
            "jkbank", "jammukashmirbank",
            "dcbbank",
            "lakshmivilasbank",
            "dbsbank",
            # Payment Processors and Wallets
            "upi",
            "upi-"
            "paytm", "okpaytm","pytm",
            # "phonepe",
            "gpay", "googlepay","amazonupi"
            "amazonpay",
            "mobikwik",
            "freecharge",
            "cred",
            "billdesk",
            "citruspay", "citrus",
            "cashfree",
            "instamojo",
            "payoneer",
            "razorpay", "razor",
            "bhim",
            "nsdl",
            "traces","gpay","paytmqr","@axl"
            # Tax and Financial Terms
            "incometax",
            "itrfees",
            "cbdt",
            "dtax",
            # UPI Handles and Prefixes
            "okaxis", "okicici", "okhdfc", "okhdfcbank",
            "oksbi", "okyesbank", "okkotak", "okbob", "okpnb",
            "okunionbank", "okcanara", "okboi", "okiob",
            "idfcbank", "idfc",
            "ach/",
            "gib/",
            "upi/p2m", "paytmqr", "paytm","p2a",'p2m',"paytmpay","p m"
            # Common Misspellings or Typos
            "icide",  # Typo for ICICI
            "axie",  # Typo for Axis
            "hdffc",  # Typo for HDFC
            "okic", "oki", "okhdf", "okaxi", "okax",
            "rev-", "okbizaxis", "yesupi", "yblupi",
            "idfcfirstbanklimi", "utib", "barb", "eaw-",
            "okbizicici", "sbin", "-utib", "ybl", "yesb"
            # Other Related Terms
            "paymentoncred", "paymentoncredit", "creditpayment"
            #creditor and detors
            "toib-", "brn-clg-chq", "mmt/imps", "neftdr", "neft/mb/", "nft/", "mob/tpft",
            "nlcindialtd", "neft/mb/ax", "tortgs", "rtgsdr", "mob/tpft/", "imb/", "imps",
            "imps/p2a", "mob/selfft/", "inb/", "inb-", "chqpaid", "fundtrf", "iconn",
            "imps-cib", "imps-inet", "imps-rib", "imps-mob", "inft", "mbk/xfer", "neft",
            "payc", "r-utr", "vmt-icon", "chqpaid", "byclg", "rtgs", "neftn", "inb-",
            "neft-barb", "ecs/", "googleindiadigital", "one97communicationslimited",
            "toib-", "neft", "mmt/imps", "neftcr", "imps", "tortgs", "rtgs", "rtgscr",
            "ecs/", "mob/tpft/", "imb/", "mob/selfft/", "inb/", "imps-mob", "nft/",
            "byclg", "inb-", "neft-", "googleindiadigital", "gsttaxpayment","fromgoogleindiadigital"
        ]
        # Preprocessing function for 'Description'
        def preprocess_description(description):
            # Handle missing or empty descriptions
            if pd.isnull(description) or description.strip() == '':
                return ''
            # Remove special characters but keep slashes (/)
            clean_description = re.sub(r'[^a-zA-Z/\s]', ' ', description)
            # Exclude keywords
            for keyword in exclude_keywords:
                pattern = rf'\b{re.escape(keyword)}\b'  # Ensure full word matching and escape special characters
                clean_description = re.sub(pattern, '', clean_description, flags=re.IGNORECASE)
            # Clean up extra spaces after removing keywords
            clean_description = re.sub(r'\s+', ' ', clean_description).strip()
            def remove_keywords(text, keywords):
                for keyword in keywords:
                    # If the keyword contains non-word characters, don't use word boundaries
                    if re.search(r'\W', keyword):
                        pattern = re.escape(keyword)
                    else:
                        pattern = r'\b' + re.escape(keyword) + r'\b'
                    text = re.sub(pattern, '', text, flags=re.IGNORECASE)
                return text
            clean_description = remove_keywords(clean_description, exclude_keywords)
            # Convert to lowercase
            clean_description = clean_description.lower()
            return clean_description
        # Apply the preprocessing to 'Description' column
        df['Description_processed'] = df['Description'].apply(preprocess_description)
        return df

    def category_add_ca(self, df):
        x = df["Balance"]
        des = df['Description']
        df["Debit"] = pd.to_numeric(df["Debit"], errors="coerce")
        df["Credit"] = pd.to_numeric(df["Credit"], errors="coerce")
        for col in df.columns:
            if df[col].dtype == "object":
                df[col] = df[col].str.lower()
        df["Description"] = df["Description"].str.replace(" ", "")
        excel_file_path = os.path.join(BASE_DIR, "CA_Category_sheet.xlsx")
        df2 = pd.read_excel(excel_file_path)
        # Initialize the 'Category' column with "Suspense" for all rows
        df["Category"] = "Suspense"
        df["Entity"] = ""
        pos_pattern = r"^pos.*"
        df.loc[
            (df["Description"].str.contains(pos_pattern, regex=True))
            & (df["Debit"] > 0),
            "Category",
        ] = "POS-Dr"
        df.loc[
            (df["Description"].str.contains(pos_pattern, regex=True))
            & (df["Credit"] > 0),
            "Category",
        ] = "POS-Cr"
        pos_pattern_2 = r"^(vps|ips|ecom|pur|pcd|edc|ecompur)"
        df.loc[
            (df["Description"].str.contains(pos_pattern_2, regex=True))
            & (df["Debit"] > 0),
            "Category",
        ] = "POS-Dr"
        df.loc[
            (df["Description"].str.contains(pos_pattern_2, regex=True))
            & (df["Credit"] > 0),
            "Category",
        ] = "POS-Cr"
        def categorize_bank_charges(df):
            Bank_charges = r"^(bctt|nchg|tChg|tip/scg|rate\.diff|owchquereturncharges|inwardchqreturncharge|chrg|incidentalcharges|iwchq|smschrg|chrg:sms|\*chrg:sms|nachreturncharges|fundtransfercharges|cashwithdrawalchgs|impschg|monthlysmscha|amcatmcharges|monthlyservicechrgs)"
            df = df[
                ~df["Description"].str.contains("POS-Cr|POS-Dr", regex=True, na=False)
            ]
            df.loc[
                (df["Description"].str.contains(Bank_charges, regex=True))
                & (df["Debit"] > 0),
                "Category",
            ] = "Bank Charges"
            return df
        df = categorize_bank_charges(df)
        Bank_charges = r"(wchrgs)"
        df.loc[
            (df["Description"].str.contains(Bank_charges, case=False, regex=True))
            & (df["Debit"] > 0),
            "Category",
        ] = "Bank Charges"
        Bank_Interest_Recieved = r"^(int)"
        df.loc[
            (df["Description"].str.contains(Bank_Interest_Recieved, regex=True))
            & (df["Credit"] > 0),
            "Category",
        ] = "Bank Interest Received"
        Bounce = r"^(r-ret-utr)"
        df.loc[
            (df["Description"].str.contains(Bounce, regex=True)) & (df["Debit"] > 0),
            "Category",
        ] = "Bounce"
        Cash_Withdrawal = r"^(ccwd|vat|mat|nfs|atm|atm-cash-axis|atm-cash|atw|csw|atd|ati|vmt|inf|cwdr|self|cash-atm|atl/|cashpmt|chqpaid)"
        df.loc[
            (df["Description"].str.contains(Cash_Withdrawal, case=False, regex=True))
            & (df["Debit"] > 0),
            "Category",
        ] = "Cash Withdrawal"
        General_insurance = r"^(pac)"
        df.loc[
            (df["Description"].str.contains(General_insurance, case=False, regex=True))
            & (df["Debit"] > 0),
            "Category",
        ] = "General insurance"
        Indirect_tax = r"^(idtx)"
        df.loc[
            (df["Description"].str.contains(Indirect_tax, case=False, regex=True))
            & (df["Debit"] > 0),
            "Category",
        ] = "Indirect tax"
        interest_paid = r"^(int.coll)"
        df.loc[
            (df["Description"].str.contains(interest_paid, case=False, regex=True))
            & (df["Debit"] > 0),
            "Category",
        ] = "interest paid"
        investment = r"^(eba|autosweep|growwpay)"
        df.loc[
            (df["Description"].str.contains(investment, case=False, regex=True))
            & (df["Debit"] > 0),
            "Category",
        ] = "Investment"
        investment = r"(growwpay)"
        df.loc[
            (df["Description"].str.contains(investment, case=False, regex=True))
            & (df["Debit"] > 0),
            "Category",
        ] = "Investment"
        Local_cheque_collection = r"^(lccbrncms)"
        df.loc[
            (
                df["Description"].str.contains(
                    Local_cheque_collection, case=False, regex=True
                )
            )
            & (df["Debit"] > 0),
            "Category",
        ] = "Local cheque collection"
        emi = r"^(emi|lnpy)"
        df.loc[
            (df["Description"].str.contains(emi, case=False, regex=True))
            & (df["Debit"] > 0),
            "Category",
        ] = "Probable EMI"
        # Utility_Bills = r'^(bbps|bpay|axmob-mer|axmob-dthr)'
        # df.loc[(df['Description'].str.contains(Utility_Bills, case=False, regex=True)) & (
        #         df['Debit'] > 0), 'Category'] = 'Utility Bills'
        Tax_Payment = r"^(gib)"
        df.loc[
            (df["Description"].str.contains(Tax_Payment, case=False, regex=True))
            & (df["Debit"] > 0),
            "Category",
        ] = "Tax Payment"
        Tax_Payment = r"(gsttaxpayment|gst@)"
        df.loc[
            (df["Description"].str.contains(Tax_Payment, case=False, regex=True))
            & (df["Debit"] > 0),
            "Category",
        ] = "GST Paid"
        Refund_Reversal = r"^(ft-rev|revchrg|rev:imps|imps:rec|imps_ret)"
        df.loc[
            (df["Description"].str.contains(Refund_Reversal, case=False, regex=True))
            & (df["Credit"] > 0),
            "Category",
        ] = "Refund/Reversal"
        Refund = r"^(imps:rec|ref-tr)"
        df.loc[
            (df["Description"].str.contains(Refund, case=False, regex=True))
            & (df["Credit"] > 0),
            "Category",
        ] = "Refund/Reversal"
        Redemption = r"^(revsweep|sewwptrf)"
        df.loc[
            (df["Description"].str.contains(Redemption, case=False, regex=True))
            & (df["Credit"] > 0),
            "Category",
        ] = "Redemption,Dividend & Interest"
        Recharge = r"^(rchg)"
        df.loc[
            (df["Description"].str.contains(Recharge, case=False, regex=True))
            & (df["Credit"] > 0),
            "Category",
        ] = "Recharge"
        # SBI
        NEFT_SBI = df[df["Description"].str.contains("bytransfer-neft", na=False)]
        if not NEFT_SBI.empty:
            NEFT_1 = NEFT_SBI[
                ~NEFT_SBI["Category"].str.contains("Redemption, Dividend & Interest")
            ]
            def extract_category(description):
                try:
                    # Extract the part between the third and fourth '*' delimiters
                    parts = description.split("*")
                    if len(parts) > 3:
                        name_part = parts[3]  # Extract the part after the third '*'
                        # Remove any trailing content like '*neft/-'
                        return name_part.split("*")[0].strip()
                    else:
                        return "Suspense"  # Fallback for descriptions without enough parts
                except IndexError:
                    return "Suspense"
            neft_names = NEFT_1["Description"].apply(extract_category)
            NEFT_1["Category"] = neft_names
            NEFT_1["Entity"] = neft_names
            df.update(NEFT_1)
        def extract_text_after_numeric(s):
            matches = re.findall(
                r"\d+([^\d]+)", s
            )  # Regex to find digits followed by non-digits
            return (
                matches[-1] if matches else None
            )  # Returning the text after the last numeric character
        # Select all NEFT entries
        NEFT = df[df["Description"].str.contains("neft", na=False)]
        # Process for NEFT_1
        NEFT_1 = NEFT[NEFT["Description"].str.contains("neft/", na=False)]
        if not NEFT_1.empty:
            NEFT_1 = NEFT_1[NEFT_1["Category"].str.contains("Suspense")]
            NEFT_1 = NEFT_1[
                ~NEFT_1["Category"].str.contains(
                    "Redemption, Dividend & Interest,Salary Paid,Salary Received"
                )
            ]
            # neft_names = NEFT_1['Description'].apply(lambda x: x.split('/')[2])
            try:
                neft_names = NEFT_1["Description"].apply(lambda x: x.split("/")[2])
                NEFT_1["Category"] = neft_names
                NEFT_1["Entity"] = neft_names
            except Exception as e:
                print("Error in splitting NEFT names: ", e)
            # NEFT_1['Category'] = neft_names
            df.update(NEFT_1)
        # Process for NEFT_KOTAK
        NEFT_KOTAK = NEFT[
            ~NEFT["Category"].str.contains(
                "Debtor|Creditor|Suspense|Redemption,Dividend & Interest|Salary Paid|Salary Received",
                na=False,
            )
        ]
        if not NEFT_KOTAK.empty:
            extracted_text = NEFT_KOTAK["Description"].apply(extract_text_after_numeric)
            NEFT_KOTAK["Category"] = extracted_text
            df.update(NEFT_KOTAK)
        def extract_neft_colon_category(description):
            try:
                name_part = description.split("neft:")[1]
                name_without_numbers = re.sub(r"\d+$", "", name_part)
                return name_without_numbers.strip()
            except IndexError:
                return "Suspense"  # Or any other default value you prefer
        NEFT_colon = df[df["Description"].str.contains("neft:", na=False)]
        if not NEFT_colon.empty:
            NEFT_colon = NEFT_colon[
                ~NEFT_colon["Category"].str.contains(
                    "Debtor|Creditor|Redemption, Dividend & Interest,Salary Paid,Salary Received",
                    na=False,
                )
            ]
            NEFT_colon["Category"] = NEFT_colon["Description"].apply(
                extract_neft_colon_category
            )
            df.update(NEFT_colon)
        def extract_nefto_union_category(description):
            try:
                name_part = description.split("nefto-")[1]
                name_without_numbers = re.sub(r"\d+$", "", name_part)
                return name_without_numbers.strip()
            except IndexError:
                return "Suspense"  # Or any other default value you prefer
        NEFTO = df[df["Description"].str.contains("nefto-", na=False)]
        if not NEFTO.empty:
            NEFTO = NEFTO[
                ~NEFTO["Category"].str.contains(
                    "Debtor|Creditor|Redemption, Dividend & Interest,Salary Paid,Salary Received",
                    na=False,
                )
            ]
            NEFTO["Category"] = NEFTO["Description"].apply(extract_nefto_union_category)
            df.update(NEFTO)
        def extract_rtgso_union_category(description):
            try:
                name_part = description.split("rtgso-")[1]
                name_without_numbers = re.sub(r"\d+$", "", name_part)
                return name_without_numbers.strip()
            except IndexError:
                return "Suspense"
        RTGSO = df[df["Description"].str.contains("rtgso-", na=False)]
        if not RTGSO.empty:
            RTGSO = RTGSO[
                ~RTGSO["Category"].str.contains(
                    "Debtor|Creditor|Redemption, Dividend & Interest,Salary Paid,Salary Received",
                    na=False,
                )
            ]
            RTGSO["Category"] = RTGSO["Description"].apply(extract_rtgso_union_category)
            df.update(RTGSO)
        def extract_rtgsfr_category(description):
            try:
                category_part = description.split("rtgsfr:")[1].split("/")[0]
                return category_part
            except IndexError:
                return "Suspense"
        RTGSFR = df[df["Description"].str.contains("rtgsfr:", na=False)]
        if not RTGSFR.empty:
            RTGSFR = RTGSFR[
                ~RTGSFR["Category"].str.contains(
                    "Debtor|Creditor|Redemption, Dividend & Interest,Salary Paid,Salary Received",
                    na=False,
                )
            ]
            RTGSFR["Category"] = RTGSFR["Description"].apply(extract_rtgsfr_category)
            df.update(RTGSFR)
        def extract_rtgso_union_category(description):
            try:
                name_part = description.split("rtgs:")[1]
                name_without_numbers = re.sub(r"\d+$", "", name_part)
                return name_without_numbers.strip()
            except IndexError:
                return "Suspense"
        RTGSO = df[df["Description"].str.contains("rtgs:", na=False)]
        if not RTGSO.empty:
            RTGSO = RTGSO[
                ~RTGSO["Category"].str.contains(
                    "Debtor|Creditor|Redemption, Dividend & Interest,Salary Paid,Salary Received",
                    na=False,
                )
            ]
            RTGSO["Category"] = RTGSO["Description"].apply(extract_rtgso_union_category)
            df.update(RTGSO)
        rtgs_ib = df[df["Description"].str.contains("ib/rtgs/", na=False)]
        rtgs_ib = rtgs_ib[
            ~rtgs_ib["Category"].str.contains(
                "Debtor|Creditor|Redemption, Dividend & Interest,Salary Paid,Salary Received",
                na=False,
            )
        ]
        if not rtgs_ib.empty:
            def extract_category(description):
                parts = description.split("/")
                return (
                    parts[3] if len(parts) > 3 else "Suspense"
                )  # Default value for descriptions without enough parts
            rtgs_names = rtgs_ib["Description"].apply(extract_category)
            rtgs_ib["Category"] = rtgs_names
            df.update(rtgs_ib)
        NEFT_ib = df[df["Description"].str.contains("ib/neft/", na=False)]
        if not NEFT_ib.empty:
            NEFT_ib = NEFT_ib[
                ~NEFT_ib["Category"].str.contains(
                    "Redemption, Dividend & Interest,Salary Paid,Salary Received"
                )
            ]
            def extract_category(description):
                parts = description.split("/")
                return (
                    parts[3] if len(parts) > 3 else "Suspense"
                )  # Default value for descriptions without enough parts
            neft_names = NEFT_ib["Description"].apply(extract_category)
            NEFT_ib["Category"] = neft_names
            df.update(NEFT_ib)
        NEFT = df[df["Description"].str.contains("neft/mb/ax", na=False)]
        if not NEFT.empty:
            NEFT = NEFT[
                ~NEFT["Category"].str.contains(
                    "Redemption, Dividend & Interest,Salary Paid,Salary Received"
                )
            ]
            def extract_category(description):
                parts = description.split("/")
                return (
                    parts[3] if len(parts) > 3 else "Suspense"
                )
            neft_names = NEFT["Description"].apply(extract_category)
            NEFT["Category"] = neft_names
            NEFT['Entity'] = neft_names
            df.update(NEFT)
            print(df)
        def extract_keyword_from_description(description):
            parts = description.split("--")
            if len(parts) < 2:
                return None
            words = parts[0].split("-")
            return max(words, key=len, default=None)
        NEFT_entries = df[df["Description"].str.contains("neft-", na=False)]
        NEFT_entries = NEFT_entries[
            ~NEFT_entries["Category"].str.contains(
                "Redemption, Dividend & Interest,Salary Paid,Salary Received"
            )
        ]
        if not NEFT_entries.empty:
            for idx, row in NEFT_entries.iterrows():
                keyword = extract_keyword_from_description(row["Description"])
                if keyword and not keyword.isdigit():
                    df.at[idx, "Category"] = keyword
                else:
                    df.at[idx, "Category"] = "Suspense"
        def extract_neft_io_category(description):
            try:
                category_part = description.split("-")[3]
                if any(char.isdigit() for char in category_part):
                    return "Suspense"
                return category_part
            except IndexError:
                return "Suspense"
        NEFT_IO = df[
            df["Description"].str.contains("neft-", na=False)
            & df["Category"].str.contains("Suspense", na=False)
        ]
        if not NEFT_IO.empty:
            NEFT_IO = NEFT_IO[
                ~NEFT_IO["Category"].str.contains(
                    "Redemption, Dividend & Interest|Bank Interest Received,Salary Paid,Salary Received"
                )
            ]
            NEFT_IO = NEFT_IO[~NEFT_IO["Category"].str.contains("Debtor")]
            NEFT_IO = NEFT_IO[~NEFT_IO["Category"].str.contains("Creditor")]
            NEFT_IO["Category"] = NEFT_IO["Description"].apply(extract_neft_io_category)
            df.update(NEFT_IO)
        def extract_neft_hdfc_cr_category(description):
            parts = description.split("-")
            if len(parts) > 2:
                # Remove trailing numbers from the extracted part
                return re.sub(r'\d+$', '', parts[2]).strip()
            else:
                return "Suspense"  # Default value for descriptions without enough parts
        NEFT_HDFC_CR = df[df["Description"].str.contains("neftcr", na=False)]
        if not NEFT_HDFC_CR.empty:
            NEFT_HDFC_CR = NEFT_HDFC_CR[
                ~NEFT_HDFC_CR["Category"].str.contains(
                    "Redemption, Dividend & Interest|Bank Interest Received,Salary Paid,Salary Received"
                )
            ]
            neft_names = NEFT_HDFC_CR["Description"].apply(
                extract_neft_hdfc_cr_category
            )
            NEFT_HDFC_CR["Category"] = neft_names
            NEFT_HDFC_CR["Entity"] = neft_names
            df.update(NEFT_HDFC_CR)
        NEFT_HDFC_DR = df[df["Description"].str.contains("neftdr", na=False)]
        if not NEFT_HDFC_DR.empty:
            NEFT_HDFC_DR = NEFT_HDFC_DR[
                ~NEFT_HDFC_DR["Category"].str.contains(
                    "Redemption, Dividend & Interest,Salary Paid,Salary Received"
                )
            ]
            def extract_category(description):
                try:
                    if "-" in description:
                        # Extract the third part and remove trailing numbers
                        return re.sub(r'\d+$', '', description.split("-")[2]).strip()
                    else:
                        return description
                except IndexError:
                    return "Suspense"
            neft_names = NEFT_HDFC_DR["Description"].apply(extract_category)
            NEFT_HDFC_DR["Category"] = neft_names
            NEFT_HDFC_DR["Entity"] = neft_names
            df.update(NEFT_HDFC_DR)
        NEFT_thane = df[df["Description"].str.contains("toneft", na=False)]
        if not NEFT_thane.empty:
            NEFT_thane = NEFT_thane[
                ~NEFT_thane["Category"].str.contains(
                    "Redemption, Dividend & Interest,Salary Paid,Salary Received"
                )
            ]
            def extract_category(description):
                try:
                    return description.split("/")[1]
                except IndexError:
                    return "Suspense"  # Default to 'Suspense' in case of IndexError
            neft_thane_names = NEFT_thane["Description"].apply(extract_category)
            NEFT_thane["Category"] = neft_thane_names
            df.update(NEFT_thane)
        NEFT_UCO = df[
            (df["Description"].str.contains("neft/", na=False))
            & (df["Category"] == "Suspense")
        ]
        if not NEFT_UCO.empty:
            NEFT_1 = NEFT_UCO[
                ~NEFT_UCO["Category"].str.contains(
                    "Redemption, Dividend & Interest","Salary Paid","Salary Received"
                )
            ]
            NEFT_1 = NEFT_1[~NEFT_1["Category"].str.contains("Debtor")]
            NEFT_1 = NEFT_1[~NEFT_1["Category"].str.contains("Creditor")]
            def extract_category(description):
                try:
                    return description.split("/")[1]
                except IndexError:
                    return "Suspense"  # Default to 'Suspense' in case of IndexError
            neft_uco_names = NEFT_1["Description"].apply(extract_category)
            NEFT_1["Category"] = neft_uco_names
            df.update(NEFT_1)
        def extract_net_neft_category(description):
            try:
                # Split by 'rtgsfr:' and take the first part after it
                category_part = description.split("net/neft/")[1].split("/")[0]
                return category_part
            except IndexError:
                # In case of an IndexError, return 'Suspense'
                return "Suspense"
        net_neft = df[df["Description"].str.contains("net/neft/", na=False)]
        net_neft = net_neft[
            ~net_neft["Category"].str.contains(
                "Redemption, Dividend & Interest,Salary Paid,Salary Received"
            )
        ]
        if not net_neft.empty:
            net_neft["Category"] = net_neft["Description"].apply(
                extract_net_neft_category
            )
            df.update(net_neft)
        def extract_nft_category(description):
            try:
                category_part = description.split("nft/")[1].split("/")[0]
                return category_part
            except IndexError:
                return "Suspense"
        nft_neft = df[df["Description"].str.contains("nft/", na=False)]
        nft_neft = nft_neft[
            ~nft_neft["Category"].str.contains(
                "Redemption, Dividend & Interest,Salary Paid,Salary Received"
            )
        ]
        if not nft_neft.empty:
            nft_neft["Category"] = nft_neft["Description"].apply(extract_nft_category)
            df.update(nft_neft)
        def extract_bob_neft_category(description):
            try:
                category_part = description.split("-")[2]
                if any(char.isdigit() for char in category_part):
                    return "Suspense"
                return category_part
            except IndexError:
                return "Suspense"
        NEFT_BOB = df[df["Description"].str.contains("neft-", na=False)]
        if not NEFT_BOB.empty:
            NEFT_BOB = NEFT_BOB[
                ~NEFT_BOB["Category"].str.contains(
                    "Redemption, Dividend & Interest|Bank Interest Received,Salary Paid,Salary Received"
                )
            ]
            NEFT_BOB = NEFT_BOB[~NEFT_BOB["Category"].str.contains("Debtor")]
            NEFT_BOB = NEFT_BOB[~NEFT_BOB["Category"].str.contains("Creditor")]
            NEFT_BOB["Category"] = NEFT_BOB["Description"].apply(
                extract_bob_neft_category
            )
            df.update(NEFT_BOB)
        def extract_neft_category(description):
            try:
                name_part = description.split("-")[3]
                if any(char.isdigit() for char in name_part):
                    return "Suspense"
                return name_part
            except IndexError:
                return "Suspense"
        NEFT_IO = df[
            df["Description"].str.contains("neft-", na=False)
            & df["Category"].str.contains("Suspense", na=False)
        ]
        if not NEFT_IO.empty:
            NEFT_IO = NEFT_IO[
                ~NEFT_IO["Category"].str.contains(
                    "Redemption, Dividend & Interest|Bank Interest Received,Salary Paid,Salary Received"
                )
            ]
            NEFT_IO = NEFT_IO[~NEFT_IO["Category"].str.contains("Debtor")]
            NEFT_IO = NEFT_IO[~NEFT_IO["Category"].str.contains("Creditor")]
            NEFT_IO["Category"] = NEFT_IO["Description"].apply(extract_neft_category)
            df.update(NEFT_IO)
        def extract_net_neft_category(description):
            try:
                category_part = description.split("net-neft-")[1].split("-")[3]
                return category_part
            except IndexError:
                return "Suspense"
        NEFT = df[df["Description"].str.contains("net-neft-", na=False)]
        NEFT = NEFT[
            ~NEFT["Category"].str.contains(
                "Redemption, Dividend & Interest|Bank Interest Received,Salary Paid,Salary Received"
            )
        ]
        if not NEFT.empty:
            NEFT["Category"] = NEFT["Description"].apply(extract_net_neft_category)
            df.update(NEFT)
        def extract_neft_name(description):
            try:
                name_part = description.split("neft-")[1].split("/")[0]
                return name_part
            except IndexError:
                return "Suspense"
        NEFT_Kar = df[
            df["Description"].str.contains("neft-", na=False)
            & df["Category"].str.contains("Suspense", na=False)
        ]
        NEFT_Kar = NEFT_Kar[
            ~NEFT_Kar["Category"].str.contains(
                "Redemption, Dividend & Interest|Bank Interest Received,Salary Paid,Salary Received"
            )
        ]
        if not NEFT_Kar.empty:
            NEFT_Kar["Category"] = NEFT_Kar["Description"].apply(extract_neft_name)
            df.update(NEFT_Kar)
        def extract_category_neft_sbi(description):
            if "totransfer-neft" in description:
                parts = description.split("-")
                name_part = parts[-1]
                return name_part
            else:
                try:
                    parts = description.split("/")
                    category_part = parts[2]
                    if any(char.isdigit() for char in category_part):
                        return "Suspense"
                    else:
                        return category_part
                except IndexError:
                    return "Suspense"
        # neft_SBI = df[df["Description"].str.contains("totransfer-neft", na=False)]
        # neft_SBI = neft_SBI[
        #     ~neft_SBI["Category"].str.contains(
        #         "Redemption, Dividend & Interest|Bank Interest Received,Salary Paid,Salary Received"
        #     )
        # ]
        # if not neft_SBI.empty:
        #     neft_SBI["Category"] = neft_SBI["Description"].apply(
        #         extract_category_neft_sbi
        #     )
        #     df.update(neft_SBI)
        def filter_emi_transactions(df):
            # Convert 'Debit' to numeric, coercing errors to NaN
            df["Debit"] = pd.to_numeric(df["Debit"], errors="coerce")
            # Define the pattern for EMI-related keywords
            keywords = [
                "emi",
                "achidfcfblimited",
                "cholamandalaminvest",
                "lnpy",
                "ach/",
                "ach-",
                "achdr",
                "ecs",
                "achd",
                "bajajfinance",
                "cms",
                "lamum",
                "lcfm",
                "loanreco",
                "lptne",
                "nach",
                "magmafincorpltd",
                "toachdraditybirl",
                "toachdrambitfinv",
                "toachdrclixcapita",
                "toachdrdeutscheb",
                "toachdrdhaniloan",
                "toachdrfedbankfi",
                "toachdrfullerton",
                "toachdrindiabulls",
                "toachdrindinfhouf",
                "toachdrindusind",
                "toachdrlendingkar",
                "toachdrmagmafinco",
                "toachdrmahnimahin",
                "toachdrmoneywisef",
                "toachdrneogrowth",
                "toachdrtatacapita",
                "toachdrtpachmag",
                "toachdrtpachneo",
                "toachdrtpcapfrst",
                "toachdryesbankr",
                "achracpc",
            ]
            pattern = r"(" + "|".join(keywords) + r")"
            # Filter for transactions containing the keywords
            emi_transactions = df[
                df["Description"].str.contains(pattern, case=False, regex=True)
                & (~df["Debit"].isnull())
                & (df["Debit"] > 0)
            ]
            # Find repeated transactions with the same debit value
            repeated_emi = emi_transactions[
                emi_transactions.duplicated(subset=["Debit"], keep=False)
            ]
            # Exclude transactions with values ending in '000' and values <= 1000
            filtered_emi_indices = repeated_emi[
                ~repeated_emi["Debit"].astype(str).str.endswith("000")
                & (repeated_emi["Debit"] > 1000)
            ].index
            # Categorize as 'Probable EMI'
            df.loc[filtered_emi_indices, "Category"] = "Probable EMI"
            return df
        filter_emi_transactions(df)
        def Bounce(df):
            keywords = ["return", "Bounce", "i/wchqreturn", "out-chqreturn"]
            pattern = r"\b(" + "|".join(keywords) + r")\b"
            df.loc[
                df["Description"].str.contains(pattern, regex=True) & (df["Debit"] > 0),
                "Category",
            ] = "Bounce"
            # print(df)
            return df
        Bounce(df)
        # Iterate through the rows of df2
        for _, keyword_row in df2.iterrows():
            mask = df["Description"].str.contains(
                keyword_row["Description"], case=False, na=False
            )
            if keyword_row["Debit / Credit"] == "Debit":
                mask = mask & (df["Debit"] > 0)  # check if Debit is greater than 0
            elif keyword_row["Debit / Credit"] == "Credit":
                mask = mask & (df["Credit"] > 0)  # check if Credit is greater than 0
            # Update the category for matching transactions
            df.loc[mask, "Category"] = keyword_row["Category"]
        # df = df[['Value Date', 'Description', 'Debit', 'Credit', 'Balance', 'Category', 'Bank']]
        #####
        MPS = df[
            df["Description"].str.contains("mps/", na=False)
            & ~df["Description"].str.contains("imps/")
        ]
        if not MPS.empty:
            for idx, row in MPS.iterrows():
                if row["Credit"] > 0:
                    df.at[idx, "Category"] = "UPI-Cr"
                elif row["Debit"] > 0:
                    df.at[idx, "Category"] = "UPI-Dr"
        Salary_credit = (
            (df["Description"].str.contains("imps|neft|rtgs", case=False, na=False))
            & (df["Description"].str.contains("salary", case=False, na=False))
            & (df["Credit"] > 0)
        )
        Salary_debit = (
            (df["Description"].str.contains("imps|neft|rtgs", case=False, na=False))
            & (df["Description"].str.contains("salary", case=False, na=False))
            & (df["Debit"] > 0)
        )
        df.loc[Salary_credit, "Category"] = "Salary Received"
        df.loc[Salary_debit, "Category"] = "Salary Paid"
        mask_withdrawal = (
            df["Description"].str.contains(
                "eaw-|nwd-|atw-|tocash", case=False, na=False
            )
        ) & (df["Debit"] > 0)
        df.loc[mask_withdrawal, "Category"] = "Cash Withdrawal"
        General_insurance = [
            "acko",
            "adityabirlahealth",
            "bajajallianz",
            "bhartiaxa",
            "carehealth",
            "cholamandalam",
            "ecgc",
            "edelweiss",
            "future generali",
            "godigit",
            "hdfcergo",
            "icicilombard",
            "iffcotokio",
            "kotakgeneral",
            "liberty",
            "manipalcigna",
            "maxbupahealth",
            "nationalinsurance",
            "pmsby",
            "rahejaqbe",
            "royalsundaram",
            "sbigeneral",
            "shriram",
            "starhealth",
            "tataaig",
            "thenewindiaassurance",
            "theoriental",
            "unitedindia",
            "universalsompo",
        ]
        df.loc[
            df["Description"].apply(
                lambda x: any(keyword in x for keyword in General_insurance)
            )
            & (df["Debit"] > 0),
            "Category",
        ] = "General insurance"
        online_shopping_keywords = [
            "amazon",
            "bigbasket",
            "ecom",
            "flipkart",
            "mamaearth",
            "myntra",
            "nykaa",
            "meesho",
        ]
        df.loc[
            df["Description"].apply(
                lambda x: any(
                    keyword in x.lower() for keyword in online_shopping_keywords
                )
                and "amazonpay" not in x.lower()
            )
            & (df["Debit"] > 0),
            "Category",
        ] = "Online Shopping"
        INB = df[df["Description"].str.contains("inb/|inb-td/", na=False)]
        INB = INB[~INB["Description"].str.contains("gsttaxpayments", na=False)]
        INB = INB[
            ~INB["Category"].str.contains(
                "Salary Paid|Salary Received|GST Paid", na=False
            )
        ]
        if not INB.empty:
            INB["Category"] = INB["Description"].apply(
                lambda x: (
                    x.split("/")[2]
                    if "inb/" in x and len(x.split("/")) > 2
                    else x.split("/")[1] if len(x.split("/")) > 1 else x
                )
            )
            INB["Entity"] = INB["Description"].apply(
                lambda x: (
                    x.split("/")[2]
                    if (("inb/ift/" in x or "inb/" in x) and len(x.split("/")) > 2)
                    else x.split("/")[1]
                    if len(x.split("/")) > 1
                    else x
                )
            )
            df.update(INB)
        BIL_IMB_entries = df[df["Description"].str.contains("bil/imb/", na=False)]
        BIL_IMB_entries = BIL_IMB_entries[
            ~BIL_IMB_entries["Category"].str.contains("Salary Paid,Salary Received")
        ]
        if not BIL_IMB_entries.empty:
            def extract_name_from_bilimb(description):
                parts = description.split("/")
                if len(parts) >= 4:
                    return parts[3]
                return None
            for idx, row in BIL_IMB_entries.iterrows():
                name = extract_name_from_bilimb(row["Description"])
                if name and not name.isdigit():
                    df.at[idx, "Category"] = name
                else:
                    df.at[idx, "Category"] = "Suspense"
        ECS = df[df["Description"].str.contains("ecs/", na=False)]
        ECS = ECS[~ECS["Category"].str.contains("Salary Paid,Salary Received")]
        if not ECS.empty:
            ECS = ECS[~ECS["Category"].str.contains("Redemption, Dividend & Interest")]
            ECS = ECS[~ECS["Category"].str.contains("Probable EMI|Bank Charges")]
            ECS_names = ECS["Description"].apply(lambda x: x.split("/")[1])
            ECS["Category"] = ECS_names
            df.update(ECS)
        MPS = df[df["Description"].str.contains("MPS/", na=False)]
        if not ECS.empty:
            ECS = ECS[~ECS["Description"].str.contains("imps/")]
        IMPS_HDFC = df[df["Description"].str.contains("imps", na=False)]
        IMPS_HDFC = IMPS_HDFC[
            ~IMPS_HDFC["Category"].str.contains("Salary Paid,Salary Received")
        ]
        if not IMPS_HDFC.empty:
            pattern = r"imps-\d+-(.*?)-"
            Ext = IMPS_HDFC["Description"].str.extract(pattern)
            IMPS_HDFC["Category"] = Ext
            IMPS_HDFC["Entity"] = Ext
            df.update(IMPS_HDFC)
        imps_rib = df[
            df["Description"].str.contains("imps-rib|imps-inet|imps-cib", na=False)
        ]
        imps_rib = imps_rib[
            ~imps_rib["Category"].str.contains("Salary Paid,Salary Received")
        ]
        if not imps_rib.empty:
            imps_rib_name = imps_rib["Description"].apply(lambda x: x.split("/")[3])
            imps_rib["Category"] = imps_rib_name
            imps_rib["Entity"] = imps_rib_name
            df.update(imps_rib)
        imps_mob = df[df["Description"].str.contains("imps-mob/", na=False)]
        imps_mob = imps_mob[
            ~imps_mob["Category"].str.contains("Salary Paid,Salary Received")
        ]
        if not imps_mob.empty:
            imps_mob_name = imps_mob["Description"].apply(lambda x: x.split("/")[3])
            imps_mob["Category"] = imps_mob_name
            imps_mob["Entity"] = imps_mob_name
            df.update(imps_mob)
        # imps_idfc = df[df["Description"].str.contains("imps/", na=False)]
        # if not imps_idfc.empty:
        #     imps_idfc_name = imps_idfc['Description'].apply(lambda x: x.split('/')[2])
        #     imps_idfc['Category'] = imps_idfc_name
        #     df.update(imps_idfc)
        def extract_imps_hdfc_category(description):
            try:
                name_part = description.split("/")[2]
                if any(char.isdigit() for char in name_part):
                    return "Suspense"
                return name_part
            except IndexError:
                return "Suspense"
        imps_idfc = df[df["Description"].str.contains("imps/", na=False)]
        imps_idfc = imps_idfc[
            ~imps_idfc["Category"].str.contains("Salary Paid,Salary Received")
        ]
        if not imps_idfc.empty:
            imps_idfc["Category"] = imps_idfc["Description"].apply(
                extract_imps_hdfc_category
            )
            df.update(imps_idfc)
        def extract_category_axis(x):
            try:
                category_part = x.split("/")[3]
                if any(char.isdigit() for char in category_part):
                    return "Suspense"
                return category_part
            except IndexError:
                return "Suspense"
        def extract_entity_axis(x):
            try:
                category_part = x.split("/")[3]
                if any(char.isdigit() for char in category_part):
                    return ""
                return category_part
            except IndexError:
                return ""
        imps_axis = df[df["Description"].str.contains("imps/p2a", na=False)]
        imps_axis = imps_axis[
            ~imps_axis["Category"].str.contains("Salary Paid","Salary Received")
        ]
        if not imps_axis.empty:
            imps_axis["Category"] = imps_axis["Description"].apply(
                extract_category_axis
            )
            imps_axis["Entity"] = imps_axis["Description"].apply(
                extract_entity_axis
            )
            df.update(imps_axis)
        def extract_category_axis_combined(x):
            try:
                try:
                    category_part = x.split("/")[3]  # Extract the 4th part (index 3)
                    category_cleaned = "".join(
                        filter(lambda char: not char.isdigit(), category_part)
                    )
                    if category_cleaned:
                        return category_cleaned
                except IndexError:
                    pass  # Fall back to new logic if the original logic fails
                parts = x.split("/")
                if len(parts) > 3:
                    category_part = parts[3]
                    category_cleaned = "".join(
                        filter(lambda char: not char.isdigit(), category_part)
                    )
                    if category_cleaned:
                        return category_cleaned
                if len(parts) > 4:
                    possible_name = parts[4]
                    if possible_name.isalpha():
                        return possible_name  # Return the name if it's valid
                return "Suspense"
            except Exception as e:
                print(f"Error processing description: {x}, Error: {str(e)}")
                return "Suspense"
        imps_axis = df[df["Description"].str.contains("mmt/imps", na=False)]
        imps_axis = imps_axis[
            ~imps_axis["Category"].str.contains("Salary Paid,Salary Received")
        ]
        if not imps_axis.empty:
            imps_axis["Category"] = imps_axis["Description"].apply(
                extract_category_axis_combined
            )
            imps_axis["Entity"] = imps_axis["Description"].apply(
                extract_category_axis_combined
            )
            df.update(imps_axis)
        imps_fed = df[df["Description"].str.contains("ftimps", na=False)]
        imps_fed = imps_fed[
            ~imps_fed["Category"].str.contains("Salary Paid,Salary Received")
        ]
        if not imps_fed.empty:
            imps_rib_name = imps_fed["Description"].apply(lambda x: x.split("/")[3])
            imps_fed["Category"] = imps_rib_name
            df.update(imps_fed)
        # imps_svc = df[df["Description"].str.contains("byimps", na=False)]
        # if not imps_svc.empty:
        #     imps_svc_name = imps_svc['Description'].apply(lambda x: x.split('-')[2])
        #     imps_svc['Category'] = imps_svc_name
        #     df.update(imps_svc)
        def safe_extract_category(description):
            try:
                return description.split("-")[2]
            except IndexError:
                return "Suspense"
        imps_svc = df[df["Description"].str.contains("byimps", na=False)]
        imps_svc = imps_svc[
            ~imps_svc["Category"].str.contains("Salary Paid,Salary Received")
        ]
        if not imps_svc.empty:
            # Apply the safe extraction function to the 'Description' column
            imps_svc["Category"] = imps_svc["Description"].apply(safe_extract_category)
            # Update the original DataFrame
            df.update(imps_svc)
        # SBI
        def extract_category_imps_sbi(description):
            try:
                parts = description.split("/")
                category_part = parts[2]
                if any(char.isdigit() for char in category_part):
                    return "Suspense"
                else:
                    return category_part
            except IndexError:
                return "Suspense"
        imps_SBI = df[
            df["Description"].str.contains("totransfer-inbimps/p2a/", na=False)
        ]
        imps_SBI = imps_SBI[
            ~imps_SBI["Category"].str.contains("Salary Paid","Salary Received")
        ]
        if not imps_SBI.empty:
            # Apply the custom function to extract categories
            imps_SBI["Category"] = imps_SBI["Description"].apply(
                extract_category_imps_sbi
            )
            imps_SBI["Entity"] = imps_SBI["Description"].apply(
                extract_category_imps_sbi
            )
            df.update(imps_SBI)
        MOB = df[df["Description"].str.contains("mob/tpft/", na=False)]
        MOB = MOB[~MOB["Category"].str.contains("Salary Paid","Salary Received")]
        if not MOB.empty:
            MOB_names = MOB["Description"].apply(
                lambda x: (
                    x.split("/")[-2] if not x.split("/")[2].isdigit() else "Suspense"
                )
            )
            MOB["Category"] = MOB_names
            df.update(MOB)
        MOB_1 = df[df["Description"].str.contains("mob/selfft/", na=False)]
        if "Category" in MOB_1.columns:  # Ensure 'Category' column exists
            MOB_1 = MOB_1[~MOB_1["Category"].str.contains("Salary Paid|Salary Received", na=False)]
        if not MOB_1.empty:
            MOB_names1 = MOB_1["Description"].apply(
                lambda x: (
                    x.split("/")[-2] if not x.split("/")[2].isdigit() else "Suspense"
                )
            )
            MOB_1["Category"] = MOB_names1
            df.update(MOB_1)
        # BRN_clg = df[df["Description"].str.contains("brn-clg-chqpaidto", na=False)]
        # if not BRN_clg.empty:
        #     BRN_clg_names = BRN_clg['Description'].apply(lambda x: x.split('to ')[-1].split('/')[0].strip())
        #     BRN_clg['Category'] = BRN_clg_names
        #     df.update(BRN_clg)
        BRN_clg = df[df["Description"].str.contains("brn-clg-chqpaidto", na=False)]
        BRN_clg = BRN_clg[
            ~BRN_clg["Category"].str.contains("Salary Paid","Salary Received")
        ]
        if not BRN_clg.empty:
            # This lambda function extracts the category name based on your statement structure
            BRN_clg["Category"] = BRN_clg["Description"].apply(
                lambda x: x.split("brn-clg-chqpaidto")[-1].split("/")[0].strip()
            )
            BRN_clg["Entity"] = BRN_clg["Description"].apply(
                lambda x: x.split("brn-clg-chqpaidto")[-1].split("/")[0].strip()
            )
            df.update(BRN_clg)
        def extract_category_ben(x):
            try:
                category_part = x.split("/")[3]
                if any(char.isdigit() for char in category_part):
                    return "Suspense"
                return category_part
            except IndexError:
                return "Suspense"
        imps_ben = df[df["Description"].str.contains("imps/ben/", na=False)]
        imps_ben = imps_ben[
            ~imps_ben["Category"].str.contains("Salary Paid,Salary Received")
        ]
        if not imps_ben.empty:
            imps_ben["Category"] = imps_ben["Description"].apply(extract_category_ben)
            df.update(imps_ben)
        def extract_sentimps_text_after_numeric(description):
            try:
                # Find the position of the "sentimps" prefix
                prefix_position = description.find("sentimps")
                if prefix_position != -1:
                    # Extract the part after the prefix
                    after_prefix = description[prefix_position + len("sentimps") :]
                    # Find the first non-digit character after the numeric part
                    first_non_digit_index = len(after_prefix)
                    for i, char in enumerate(after_prefix):
                        if not char.isdigit():
                            first_non_digit_index = i
                            break
                    # Extract the text part after the numeric sequence
                    text_after_numeric = after_prefix[first_non_digit_index:]
                    # Extract the part before the next '/'
                    text_part = text_after_numeric.split("/")[0]
                    return text_part if text_part else "Suspense"
                return "Suspense"
            except Exception as e:
                return "Suspense"
        imps_new = df[df["Description"].str.contains("sentimps", na=False)]
        imps_new = imps_new[
            ~imps_new["Category"].str.contains("Salary Paid,Salary Received")
        ]
        if not imps_new.empty:
            imps_new["Category"] = imps_new["Description"].apply(
                extract_sentimps_text_after_numeric
            )
            df.update(imps_new)
        # Filter rows containing "idfcfirstb" in the Description
        Loan = df[df["Description"].str.contains("idfcfirstb|krazybees", na=False)]
        if not Loan.empty:
            # Update the Category for the filtered rows
            df.loc[
                (df["Description"].str.contains("idfcfirstb|krazybees", regex=True))
                & (df["Credit"] > 0),
                "Category",
            ] = "Loan"
        BRN = df[
            (
                df["Description"].str.contains("brn-flexi")
                | df["Description"].str.contains("/SBI Funds/STATE BAN//ATTN//")
            )
            & df["Credit"].notnull()
        ]
        if not BRN.empty:
            df.loc[BRN.index, "Category"] = "Redemption of Investment"
        RTGS = df[df["Description"].str.contains("rtgs/", na=False)]
        RTGS = RTGS[~RTGS["Category"].str.contains("Salary Paid,Salary Received")]
        if not RTGS.empty:
            RTGS_names = RTGS["Description"].apply(lambda x: x.split("/")[2])
            RTGS["Category"] = RTGS_names
            df.update(RTGS)
        RTGS_HDFC_CR = df[df["Description"].str.contains("rtgscr", na=False)]
        RTGS_HDFC_CR = RTGS_HDFC_CR[
            ~RTGS_HDFC_CR["Category"].str.contains("Salary Paid,Salary Received")
        ]
        if not RTGS_HDFC_CR.empty:
            RTGS_names = RTGS_HDFC_CR["Description"].apply(lambda x: x.split("-")[2])
            RTGS_HDFC_CR["Category"] = RTGS_names
            RTGS_HDFC_CR["Entity"] = RTGS_names
            df.update(RTGS_HDFC_CR)
        def extract_rtgs_category(description):
            try:
                return description.split("-")[2]
            except IndexError:
                return "Suspense"
        RTGS_HDFC_DR = df[
            df["Description"].str.contains("rtgsdr", na=False)
            | df["Description"].str.contains("rtgs-", na=False)
        ]
        RTGS_HDFC_DR = RTGS_HDFC_DR[
            ~RTGS_HDFC_DR["Category"].str.contains("Salary Paid,Salary Received")
        ]
        if not RTGS_HDFC_DR.empty:
            # Apply the custom function to extract RTGS names
            RTGS_HDFC_DR["Category"] = RTGS_HDFC_DR["Description"].apply(
                extract_rtgs_category
            )
            RTGS_HDFC_DR["Entity"] = RTGS_HDFC_DR["Description"].apply(
                extract_rtgs_category
            )
            df.update(RTGS_HDFC_DR)
        RTGS_DCB = df[df["Description"].str.contains("inrtgs", na=False)]
        RTGS_DCB = RTGS_DCB[
            ~RTGS_DCB["Category"].str.contains("Salary Paid,Salary Received")
        ]
        if not RTGS_DCB.empty:
            RTGS_name = RTGS_DCB["Description"].apply(lambda x: x.split("/")[1])
            RTGS_DCB["Category"] = RTGS_name
            df.update(RTGS_DCB)
        RTGS_DCB = df[df["Description"].str.contains("tortgs", na=False)]
        RTGS_DCB = RTGS_DCB[
            ~RTGS_DCB["Category"].str.contains("Salary Paid,Salary Received")
        ]
        if not RTGS_DCB.empty:
            def extract_category(description):
                try:
                    return description.split("/")[1]
                except IndexError:
                    return "Suspense"
            RTGS_name = RTGS_DCB["Description"].apply(extract_category)
            RTGS_DCB["Category"] = RTGS_name
            df.update(RTGS_DCB)
        toib_svc = df[df["Description"].str.contains("toib", na=False)]
        toib_svc = toib_svc[
            ~toib_svc["Category"].str.contains("Salary Paid,Salary Received")
        ]
        if not toib_svc.empty:
            def extract_category_svc(description):
                try:
                    return description.split("-")[2]
                except IndexError:
                    return "Suspense"
            toib_svc_names = toib_svc["Description"].apply(extract_category_svc)
            toib_svc["Category"] = toib_svc_names
            df.update(toib_svc)
        BYNEFTID_DCB = df[df["Description"].str.contains("byneftid", na=False)]
        BYNEFTID_DCB = BYNEFTID_DCB[
            ~BYNEFTID_DCB["Category"].str.contains("Salary Paid,Salary Received")
        ]
        if not BYNEFTID_DCB.empty:
            def extract_category(description):
                try:
                    return description.split("from")[1].split("]")[0]
                except IndexError:
                    return "Suspense"  # We return "Unknown" if the format is incorrect
            BYNEFTID_name = BYNEFTID_DCB["Description"].apply(extract_category)
            BYNEFTID_DCB["Category"] = BYNEFTID_name
            df.update(BYNEFTID_DCB)
        BYRTGSID_DCB = df[df["Description"].str.contains("byrtgsid", na=False)]
        BYRTGSID_DCB = BYRTGSID_DCB[
            ~BYRTGSID_DCB["Category"].str.contains("Salary Paid,Salary Received")
        ]
        if not BYRTGSID_DCB.empty:
            def extract_name_from_description(description):
                try:
                    return "".join(
                        filter(str.isalpha, description.split("[")[-1].split("]")[0])
                    )
                except IndexError:
                    return "Suspense"  # We return "Unknown" if the format is incorrect
            BYRTGSID_name = BYRTGSID_DCB["Description"].apply(
                extract_name_from_description
            )
            BYRTGSID_DCB["Category"] = BYRTGSID_name
            df.update(BYRTGSID_DCB)
        BYCLG_DCB = df[df["Description"].str.contains("byclg:", na=False)]
        BYCLG_DCB = BYCLG_DCB[
            ~BYCLG_DCB["Category"].str.contains("Salary Paid,Salary Received")
        ]
        if not BYCLG_DCB.empty:
            def extract_name_from_description(description):
                try:
                    name_part = description.split("byclg:")[1]
                    name = "".join(filter(lambda x: not x.isdigit(), name_part))
                    return name
                except IndexError:
                    return "Suspense"  # We return "Unknown" if the format is incorrect
            BYCLG_name = BYCLG_DCB["Description"].apply(extract_name_from_description)
            BYCLG_DCB["Category"] = BYCLG_name
            df.update(BYCLG_DCB)
        CASHDEP = df[
            df["Description"].str.contains(
                "cashdep|byclearing/|atr/", na=False
            )
        ]
        if not CASHDEP.empty:
            extracted_names = [
                re.search(r"CASHDEP(.*?)-", s).group(1)
                for s in CASHDEP["Description"]
                if re.search(r"CASHDEP(.*?)-", s)
            ]
            # Update the Category column of CASHDEP_ZAHIRDAHISAR with the extracted names
            CASHDEP["Category"] = "Cash Deposits"
            # Update the original dataframe df with the modified rows from CASHDEP_ZAHIRDAHISAR
            df.update(CASHDEP)
        def extract_rtgs_name_jankalyan(description):
            if "rtgs" in description.lower():
                try:
                    name_part = description.split("rtgs")[1]
                    name_without_numbers = re.match(r"[a-zA-Z]+", name_part)
                    if name_without_numbers:
                        return name_without_numbers.group(0)
                    else:
                        return "Suspense"
                except IndexError:
                    return "Suspense"
            # Removed the else condition
        # Applying the function to the DataFrame
        RTGS_jankalyan = df[
            (df["Description"].str.contains("rtgs", case=False, na=False))
            & (df["Category"] == "Suspense")
        ]
        RTGS_jankalyan = RTGS_jankalyan[
            ~RTGS_jankalyan["Category"].str.contains("Salary Paid,Salary Received")
        ]
        if not RTGS_jankalyan.empty:
            RTGS_jankalyan["Category"] = RTGS_jankalyan["Description"].apply(
                extract_rtgs_name_jankalyan
            )
            df.update(RTGS_jankalyan)
        def extract_neft_jankalyan(description):
            if "neft" in description.lower():
                try:
                    name_part = description.split("neft")[1]
                    name_without_numbers = re.match(r"[a-zA-Z]+", name_part)
                    if name_without_numbers:
                        return name_without_numbers.group(0)
                    else:
                        return "Suspense"
                except IndexError:
                    return "Suspense"
        NEFT = df[
            (df["Description"].str.contains("neft", case=False, na=False))
            & (df["Category"] == "Suspense")
        ]
        NEFT = NEFT[~NEFT["Category"].str.contains("Salary Paid,Salary Received")]
        if not NEFT.empty:
            NEFT["Category"] = NEFT["Description"].apply(extract_neft_jankalyan)
            df.update(NEFT)
        def extract_imps_category(description):
            if "imps[" in description.lower():
                try:
                    # Splitting at 'imps[' and extracting the part after it
                    parts = description.split("imps[")[1]
                    # Further splitting by ']' to isolate the segments
                    segments = parts.split("]")
                    # The desired name is likely the second segment (after the first ']')
                    if len(segments) > 1:
                        extracted_name = segments[1].strip("[").strip("]")
                        return extracted_name
                    else:
                        return "Suspense"
                except IndexError:
                    return "Suspense"
        IMPS = df[
            (df["Description"].str.contains("imps\[", case=False, na=False))
            & (df["Category"] == "Suspense")
        ]
        IMPS = IMPS[~IMPS["Category"].str.contains("Salary Paid,Salary Received")]
        if not IMPS.empty:
            IMPS["Category"] = IMPS["Description"].apply(extract_imps_category)
            df.update(IMPS)
        # df = categorize_name_transactions(df)
        PF = df[df["Description"].str.contains("providentfund", na=False)]
        if not PF.empty:
            # Update the Category for the filtered rows
            df.loc[
                (df["Description"].str.contains("providentfund", regex=True))
                & (df["Credit"] > 0),
                "Category",
            ] = "Provident Fund"
        Salary_credit = (
            (df["Description"].str.contains("mmt/imps", case=False, na=False))
            & (df["Description"].str.contains("salary", case=False, na=False))
            & (df["Credit"] > 0)
        )
        Salary_debit = (
            (df["Description"].str.contains("mmt/imps", case=False, na=False))
            & (df["Description"].str.contains("salary", case=False, na=False))
            & (df["Debit"] > 0)
        )
        df.loc[Salary_credit, "Category"] = "Salary Received"
        df.loc[Salary_debit, "Category"] = "Salary Paid"
        # last_move = r"(imps|neft|rtgs)"
        # df.loc[
        #     (df["Description"].str.contains(last_move, regex=True))
        #     & (df["Debit"] > 0)
        #     & (df["Category"] == "Suspense"),
        #     "Category",
        # ] = "Creditor"
        # df.loc[
        #     (df["Description"].str.contains(last_move, regex=True))
        #     & (df["Credit"] > 0)
        #     & (df["Category"] == "Suspense"),
        #     "Category",
        # ] = "Debtor"
        # Preprocess the DataFrame
        df = self.preprocess_df(df)
        print(df)
        labels = ["person", "org"]
        confidence_threshold = 0.2  # Adjust as needed
        # Define categories to include
        categories_to_include = ["upi-cr", "upi-dr", 'departmental stores', 'donation', 'loan',
                                'loan given', 'rent paid']
        # Keywords to include in the logic
        keywords_to_include = [
            "imps-cib", "imps-inet", "imps-rib", "imps-mob",
            "neft-barb","imps", "tortgs", "rtgs", "imps-mob","inb-", "neft-"
        ]
        categories_to_exclude = ["suspense"]
        def extract_entities(row):
            # Skip rows where Entity is already populated
            if row['Entity']:
                return row['Entity']
            if row['Category'] in categories_to_exclude:
                return row['Entity']  # Skip processing and return existing value
            # Check if the category is relevant or if keywords are present in the description
            is_category_included = row['Category'] in categories_to_include
            is_keyword_present = any(keyword.lower() in row['Description'].lower() for keyword in keywords_to_include)
            # If neither condition is met, return the existing Entity value (even if it's empty)
            if not is_category_included and not is_keyword_present:
                return row['Entity']  # Return the existing value
            text = row['Description_processed']
            if text == '':
                return ''  # Return an empty string for Entity if the description is empty
            print(text)
            # Predict entities using the model
            entities = model.predict_entities(text, labels)
            # Filter entities based on confidence threshold
            filtered_entities = [entity for entity in entities if entity["score"] >= confidence_threshold]
            # Return the entity text if any entities are found
            if filtered_entities:
                highest_score_entity = max(filtered_entities, key=lambda x: x['score'])
                return highest_score_entity['text']  # Return only the Entity text
            else:
                return ''  # Return an empty string if no entities are found
        def apply_regex_to_empty_entities_axis(row):
            if row['Entity'] == '' and row['Category'] in categories_to_include:
                if "upi/p2a" in row['Description'] or "upi/p2m" in row['Description']:
                    match = re.search(r'upi/p2[am]/\d+/(\w+)/', row['Description'])
                    if match:
                        return match.group(1)
            return row['Entity']
        def apply_regex_to_empty_entities_hdfc(row):
            if row['Entity'] == '' and row['Category'] in categories_to_include:
                if "upi-" in row['Description']:
                    match = re.search(r'(?<=upi-)([a-zA-Z]+)', row['Description'])
                    if match:
                        return match.group(1)
            return row['Entity']
        def apply_regex_to_empty_entities_sbi(row):
            if row['Entity'] == '' and row['Category'] in categories_to_include:
                if "totransfer-upi" in row['Description'] or "bytransfer-upi" in row['Description']:
                    match = re.search(r'(totransfer|bytransfer)-upi/[cd]r/\d+/([a-zA-Z]+)/', row['Description'])
                    if match:
                        return match.group(2)
            return row['Entity']
        def apply_regex_to_empty_entities_kotak(row):
            if row['Entity'] == '' and row['Category'] in categories_to_include:
                # Check if 'upi/' is in the description
                if "upi/" in row['Description']:
                    # Match the name immediately after 'upi/'
                    match = re.search(r'upi/([a-zA-Z.]+)', row['Description'])
                    if match:
                        # Clean the name by removing special characters and numbers
                        name = re.sub(r'[^a-zA-Z]', '', match.group(1))  # Keep only alphabetic characters
                        return name if name else "Suspense"  # Return 'Suspense' if the name is empty
            return row['Entity']
        def extract_name_from_neft(description):
            try:
                matches = re.findall(r'[a-zA-Z]+', description)
                if matches:
                    longest_match = max(matches, key=len)
                    return longest_match
                return "Suspense"  # Fallback if no name is found
            except Exception:
                return "Suspense"  # Fallback for errors
        NEFT_STATEMENTS = df[df["Description"].str.contains("neft", na=False)]
        if not NEFT_STATEMENTS.empty:
            NEFT_STATEMENTS["Entity"] = NEFT_STATEMENTS["Description"].apply(extract_name_from_neft)
            df.update(NEFT_STATEMENTS)
        def apply_regex_to_empty_entities_rbl(row):
            if row['Entity'] == '' and row['Category'] in categories_to_include:
                if "upi/" in row['Description']:
                    match = re.search(r'upi/\d+/\w+/([a-zA-Z0-9@.]+)', row['Description'])
                    if match:
                        # Extract the raw name
                        raw_name = match.group(1)
                        # Remove numeric and '@' characters
                        cleaned_name = re.sub(r'[0-9@]', '', raw_name)
                        return cleaned_name
            return row['Entity']
        def apply_regex_to_empty_entities_idfc(row):
            if row['Entity'] == '' and row['Category'] in categories_to_include:
                print(f"Processing Row: {row['Description']} with Category: {row['Category']}")
                if "upi/mob" in row['Description']:
                    match = re.search(r'upi/mob/\d+/([\w]+)', row['Description'])
                    if match:
                        raw_name = match.group(1)
                        cleaned_name = re.sub(r'[0-9@]', '', raw_name)
                        print(f"Extracted Name: {cleaned_name}")
                        return cleaned_name
                    else:
                        print("No Match Found")
            return row['Entity']
        # Step 1: Apply regex logic first
        df['Entity'] = df.apply(apply_regex_to_empty_entities_axis, axis=1)
        df['Entity'] = df.apply(apply_regex_to_empty_entities_hdfc, axis=1)
        df['Entity'] = df.apply(apply_regex_to_empty_entities_sbi, axis=1)
        df['Entity'] = df.apply(apply_regex_to_empty_entities_kotak, axis=1)
        df['Entity'] = df.apply(apply_regex_to_empty_entities_rbl, axis=1)
        df['Entity'] = df.apply(apply_regex_to_empty_entities_idfc, axis=1)
        print("apply",df)
        # Step 2: Apply model-based extraction only to rows with empty Entity
        df['Entity'] = df.apply(extract_entities, axis=1)
        # Drop the intermediate column if not needed
        df.drop(columns=['Description_processed'], inplace=True)
        df["Balance"] = x  # Manish
        return df


 
    ##SHEETS
    def process_name_n_num_df(self, data):
        name_n_num_df = pd.DataFrame(data, columns=['Account Number', 'Account Name', 'Bank'])
        name_n_num_df = name_n_num_df.iloc[[0]]
        df_transposed = name_n_num_df.transpose()
        df_transposed.reset_index(inplace=True)
        df = pd.DataFrame(df_transposed)
        return df

    def summary_sheet(self, idf, open_bal, close_bal, new_tran_df):

        idf["Value Date"] = pd.to_datetime(
            idf["Value Date"], format="%d-%m-%Y", errors="coerce"
        )
        idf["Month"] = idf["Value Date"].dt.strftime("%b-%Y")
        idf["Date"] = idf["Value Date"].dt.day
        new_tran_df["Value Date"] = pd.to_datetime(
            new_tran_df["Value Date"], format="%d-%m-%Y", errors="coerce"
        )
        new_tran_df["Month"] = new_tran_df["Value Date"].dt.strftime("%b-%Y")
        new_tran_df["Date"] = new_tran_df["Value Date"].dt.day
        opening_bal = open_bal
        closing_bal = close_bal

        def total_amount_cr(df):
            sum = df["Credit"].sum(axis=0)
            return sum

        def total_amount_dr(df):
            sum = df["Debit"].sum(axis=0)
            return sum

        def total_amount_cd(df):
            amount = 0
            for index, row in df.iterrows():
                credit_amount = row["Credit"]
                if row["Category"] == "Cash Deposits" and credit_amount > 0:
                    amount += credit_amount
            return amount

        def total_amount_cw(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Debit"]
                if row["Category"] == "Cash Withdrawal" and debit_amount > 0:
                    amount += debit_amount
            return amount

        def total_amount_pos_cr(df):
            amount = 0
            for index, row in df.iterrows():
                credit_amount = row["Credit"]
                if row["Category"] == "POS-Cr" and credit_amount > 0:
                    amount += credit_amount
            return amount

        def total_amount_pos_dr(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Debit"]
                if row["Category"] == "POS-Dr" and debit_amount > 0:
                    amount += debit_amount
            return amount

        def Upi_cr(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Credit"]
                if row["Category"] == "UPI-Cr" and debit_amount > 0:
                    amount += debit_amount
            return amount

        def Upi_dr(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Debit"]
                if row["Category"] == "UPI-Dr" and debit_amount > 0:
                    amount += debit_amount
            return amount

        def Debtors(df):
            amount = 0
            for index, row in df.iterrows():
                credit_amount = row["Credit"]
                if row["Category"] == "Debtor" and credit_amount > 0:
                    amount += credit_amount
            return amount

        def Creditor(df):
            amount = 0
            for index, row in df.iterrows():
                credit_amount = row["Debit"]
                if row["Category"] == "Creditor" and credit_amount > 0:
                    amount += credit_amount
            return amount

        def Bounce(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Debit"]
                if row["Category"] == "Bounce" and debit_amount > 0:
                    amount += debit_amount
            return amount

        def Cash_Reversal(df):
            amount = 0
            for index, row in df.iterrows():
                credit_amount = row["Credit"]
                if row["Category"] == "Cash Reversal" and credit_amount > 0:
                    amount += credit_amount
            return amount

        def Claim_set(df):
            amount = 0
            for index, row in df.iterrows():
                credit_amount = row["Credit"]
                if row["Category"] == "Probable Claim Settlement" and credit_amount > 0:
                    amount += credit_amount
            return amount

        def Refund_Ref(df):
            amount = 0
            for index, row in df.iterrows():
                credit_amount = row["Credit"]
                if row["Category"] == "Refund/Reversal" and credit_amount > 0:
                    amount += credit_amount
            return amount

        def Donation(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Debit"]
                if row["Category"] == "Donation" and debit_amount > 0:
                    amount += debit_amount
            return amount

        def Interest(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Debit"]
                if row["Category"] == "Interest Debit" and debit_amount > 0:
                    amount += debit_amount
            return amount

        def TDS_ON_FOREX(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Debit"]
                if row["Category"] == "tds_on_forex" and debit_amount > 0:
                    amount += debit_amount
            return amount

        def Forex(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Debit"]
                if row["Category"] == "forex" and debit_amount > 0:
                    amount += debit_amount
            return amount

        def Other_Exp(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Debit"]
                if row["Category"] == "Other Expenses" and debit_amount > 0:
                    amount += debit_amount
            return amount

        def LOAN_GR(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Debit"]
                if row["Category"] == "Loan given" and debit_amount > 0:
                    amount += debit_amount
            return amount

        def CHEQUE_PAID(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Debit"]
                if row["Category"] == "Cheque" and debit_amount > 0:
                    amount += debit_amount
            return amount

        def DEPART(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Debit"]
                if row["Category"] == "Departmental Stores" and debit_amount > 0:
                    amount += debit_amount
            return amount

        def PAY_SELF(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Debit"]
                if row["Category"] == "payToSelf" and debit_amount > 0:
                    amount += debit_amount
            return amount

        def total_investment(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Debit"]
                if row["Category"] == "Investment" and debit_amount > 0:
                    amount += debit_amount
            return amount

        def recieved_interest(df):
            amount = 0
            for index, row in df.iterrows():
                credit_amount = row["Credit"]
                if row["Category"] == "Bank Interest Received" and credit_amount > 0:
                    amount += credit_amount
            return amount

        def recieved_salary(df):
            amount = 0
            for index, row in df.iterrows():
                credit_amount = row["Credit"]
                if row["Category"] == "Salary Received" and credit_amount > 0:
                    amount += credit_amount
            return amount

        def loan_recieved(df):
            amount = 0
            for index, row in df.iterrows():
                credit_amount = row["Credit"]
                if row["Category"] == "Loan" and credit_amount > 0:
                    amount += credit_amount
            return amount

        def nach_reciept(df):
            count = 0
            for index, row in df.iterrows():
                description = row["Description"]
                if "nach" in description.lower():
                    count += 1
            return count

        def recieved_tax(df):
            amount = 0
            for index, row in df.iterrows():
                credit_amount = row["Credit"]
                if row["Category"] == "Income Tax Refund" and credit_amount > 0:
                    amount += credit_amount
            return amount

        def recieved_rent(df):
            amount = 0
            for index, row in df.iterrows():
                credit_amount = row["Credit"]
                if row["Category"] == "Rent Received" and credit_amount > 0:
                    amount += credit_amount
            return amount

        def dividend_i(df):
            amount = 0
            for index, row in df.iterrows():
                credit_amount = row["Credit"]
                if (
                    row["Category"] == "Redemption, Dividend & Interest"
                    and credit_amount > 0
                ):
                    amount += credit_amount
            return amount

        def paid_interest(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Debit"]
                if row["Category"] == "Interest Debit" and debit_amount > 0:
                    amount += debit_amount
            return amount

        def paid_salary(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Debit"]
                if row["Category"] == "Salary paid" and debit_amount > 0:
                    amount += debit_amount
            return amount

        def paid_bank(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Debit"]
                if row["Category"] == "Bank Charges" and debit_amount > 0:
                    amount += debit_amount
            return amount

        def paid_emi(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Debit"]
                if row["Category"] == "Probable EMI" and debit_amount > 0:
                    amount += debit_amount
            return amount

        def paid_tds(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Debit"]
                if row["Category"] == "TDS Deducted" and debit_amount > 0:
                    amount += debit_amount
            return amount

        def paid_tax(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Debit"]
                if row["Category"] == "Income Tax Paid" and debit_amount > 0:
                    amount += debit_amount
            return amount

        def GST(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Debit"]
                if row["Category"] == "GST Paid" and debit_amount > 0:
                    amount += debit_amount
            return amount

        def utility_bills_i(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Debit"]
                if row["Category"] == "Utility Bills" and debit_amount > 0:
                    amount += debit_amount
            return amount

        def travelling_bills(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Debit"]
                if row["Category"] == "Travelling Expense" and debit_amount > 0:
                    amount += debit_amount
            return amount

        def paid_rent(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Debit"]
                if row["Category"] == "Rent Paid" and debit_amount > 0:
                    amount += debit_amount
            return amount

        def paid_general_insurance(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Debit"]
                if row["Category"] == "General insurance" and debit_amount > 0:
                    amount += debit_amount
            return amount

        def paid_life_insurance(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Debit"]
                if row["Category"] == "Life insurance" and debit_amount > 0:
                    amount += debit_amount
            return amount

        def food_expense(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Debit"]
                if row["Category"] == "Food Expense/Hotel" and debit_amount > 0:
                    amount += debit_amount
            return amount

        def credit_card_payment(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Debit"]
                if row["Category"] == "Credit Card Payment" and debit_amount > 0:
                    amount += debit_amount
            return amount

        def paid_online_shopping(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Debit"]
                if row["Category"] == "Online Shopping" and debit_amount > 0:
                    amount += debit_amount
            return amount

        def paid_property_tax(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Debit"]
                if row["Category"] == "Property Tax" and debit_amount > 0:
                    amount += debit_amount
            return amount

        def paid_gas_payment(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Debit"]
                if row["Category"] == "Gas Payments" and debit_amount > 0:
                    amount += debit_amount
            return amount

        def paid_gold_loan(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Debit"]
                if row["Category"] == "Gold Loan" and debit_amount > 0:
                    amount += debit_amount
            # print("amount",amount)
            return amount

        def Enter(df):
            Enter = 0
            for index, row in df.iterrows():
                debit_amount = row["Debit"]
                if (
                    row["Category"] == "Subscription / Entertainment"
                    and debit_amount > 0
                ):
                    Enter += debit_amount
            return Enter

        def suspenses_dr(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Debit"]
                if row["Category"] == "Suspense" and debit_amount > 0:
                    amount += debit_amount
            return amount

        def suspenses_cr(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Credit"]
                if row["Category"] == "Suspense" and debit_amount > 0:
                    amount += debit_amount
            return amount

        # -----------------------X----------------------------------#
        new_tran_df["Credit"] = pd.to_numeric(new_tran_df["Credit"], errors="coerce")
        new_tran_df["Debit"] = pd.to_numeric(new_tran_df["Debit"], errors="coerce")
        # idf['Month'] = idf['Value Date'].dt.strftime('%b-%Y')
        months = idf["Month"].unique()

        amount_cr = {}
        amount_dr = {}
        amount_cd = {}
        amount_cw = {}
        amount_pos_cr = {}
        amount_pos_dr = {}
        upi_cr = {}
        upi_dr = {}
        debitor_list = {}
        creditor_list = {}
        bounce = {}
        cash_rev = {}
        claim_settlement = {}
        ref_ref = {}
        donation = {}

        investment = {}
        interest_debit = {}
        tds_on_forex = {}

        forex = {}
        other_expense = {}
        loan_gr = {}
        cheque = {}
        departmental = {}
        pay_self = {}

        interest_recieved = {}
        salary_recieved = {}
        nach_reciepts = {}
        loans_recieved = {}
        income_tax_refund = {}
        dividend = {}
        rent_recieved = {}

        interest_paid = {}
        salary_paid = {}
        bank_charges = {}
        emi = {}
        tds_deducted = {}
        gst = {}
        income_tax_paid = {}
        utility_bills = {}
        travelling_expense = {}
        rent_paid = {}
        total_expense = {}
        Entertainment = {}

        general_insurance = {}
        life_insurance = {}
        food_expenses = {}
        credit_card_payments = {}
        online_shopping = {}
        property_tax = {}
        gas_payment = {}
        gold_loan = {}
        rent_paid = {}
        total_amount = {}
        suspense_cr = {}
        suspense_dr = {}

        for month in months:
            new_df = new_tran_df[new_tran_df["Month"] == month].reset_index(drop=True)
            amount_cr.update({month: total_amount_cr(new_df)})
            amount_dr.update({month: total_amount_dr(new_df)})
            amount_cd.update({month: total_amount_cd(new_df)})
            amount_cw.update({month: total_amount_cw(new_df)})
            amount_pos_cr.update({month: total_amount_pos_cr(new_df)})
            amount_pos_dr.update({month: total_amount_pos_dr(new_df)})
            upi_cr.update({month: Upi_cr(new_df)})
            upi_dr.update({month: Upi_dr(new_df)})
            debitor_list.update({month: Debtors(new_df)})
            creditor_list.update({month: Creditor(new_df)})
            bounce.update({month: Bounce(new_df)})
            cash_rev.update({month: Cash_Reversal(new_df)})
            claim_settlement.update({month: Claim_set(new_df)})
            ref_ref.update({month: Refund_Ref(new_df)})
            donation.update({month: Donation(new_df)})
            investment.update({month: total_investment(new_df)})
            interest_debit.update({month: Interest(new_df)})
            tds_on_forex.update({month: TDS_ON_FOREX(new_df)})
            forex.update({month: Forex(new_df)})
            other_expense.update({month: Other_Exp(new_df)})
            loan_gr.update({month: LOAN_GR(new_df)})
            cheque.update({month: CHEQUE_PAID(new_df)})
            departmental.update({month: DEPART(new_df)})
            pay_self.update({month: PAY_SELF(new_df)})

            interest_recieved.update({month: recieved_interest(new_df)})
            salary_recieved.update({month: recieved_salary(new_df)})
            nach_reciepts.update({month: nach_reciept(new_df)})
            loans_recieved.update({month: loan_recieved(new_df)})
            income_tax_refund.update({month: recieved_tax(new_df)})
            dividend.update({month: dividend_i(new_df)})
            rent_recieved.update({month: recieved_rent(new_df)})

            interest_paid.update({month: paid_interest(new_df)})
            salary_paid.update({month: paid_salary(new_df)})
            bank_charges.update({month: paid_bank(new_df)})
            emi.update({month: paid_emi(new_df)})
            tds_deducted.update({month: paid_tds(new_df)})
            gst.update({month: GST(new_df)})
            income_tax_paid.update({month: paid_tax(new_df)})
            utility_bills.update({month: utility_bills_i(new_df)})
            travelling_expense.update({month: travelling_bills(new_df)})
            rent_paid.update({month: paid_rent(new_df)})
            Entertainment.update({month: Enter(new_df)})

            general_insurance.update({month: paid_general_insurance(new_df)})
            life_insurance.update({month: paid_life_insurance(new_df)})
            food_expenses.update({month: food_expense(new_df)})
            credit_card_payments.update({month: credit_card_payment(new_df)})
            online_shopping.update({month: paid_online_shopping(new_df)})
            property_tax.update({month: paid_property_tax(new_df)})
            gas_payment.update({month: paid_gas_payment(new_df)})
            gold_loan.update({month: paid_gold_loan(new_df)})
            rent_paid.update({month: paid_rent(new_df)})
            suspense_cr.update({month: suspenses_cr(new_df)})
            suspense_dr.update({month: suspenses_dr(new_df)})

            ###now we make sheets
            sheet_1 = pd.DataFrame([opening_bal, closing_bal, amount_cr, amount_dr])
            sheet_1.insert(
                0,
                "Particulars",
                [
                    "Opening Balance",
                    "Closing Balance",
                    "Total Amount of Credit Transactions",
                    "Total Amount of Debit Transactions",
                ],
            )
            # sheet_1['Total'] = sheet_1.iloc[:, 1:].sum(axis=1)
            # sheet_1['Total'] = sheet_1.iloc[:-2, 1:].sum(axis=1)
            # Sum for "Total Amount of Credit Transactions"
            sheet_1.loc[sheet_1.index[-2], "Total"] = sheet_1.iloc[-2, 1:].sum()
            sheet_1.loc[sheet_1.index[-1], "Total"] = sheet_1.iloc[-1, 1:].sum()

            sheet_2 = pd.DataFrame(
                [
                    salary_recieved,
                    debitor_list,
                    interest_recieved,
                    rent_recieved,
                    dividend,
                    amount_cd,
                    loans_recieved,
                    income_tax_refund,
                    amount_pos_cr,
                    upi_cr,
                    bounce,
                    cash_rev,
                    claim_settlement,
                    ref_ref,
                    suspense_cr,
                ]
            )
            sheet_2.insert(
                0,
                "Income / Receipts",
                [
                    "Salary Received",
                    "Debtors Amount",
                    "Bank Interest Received",
                    "Rent Received",
                    "Redemption, Dividend & Interest",
                    "Cash Deposits",
                    "Loans Received",
                    "Income Tax Refund",
                    "POS Txns - Cr",
                    "UPI-Cr",
                    "Bounce Transaction",
                    "Cash Reversal",
                    "Probable Claim Settlement",
                    "Refund/Reversal",
                    "Suspense - Cr",
                ],
            )
            sheet_2 = sheet_2._append(sheet_2.iloc[0:, :].sum(), ignore_index=True)
            sheet_2.iloc[-1, 0] = "Total Credit"
            sheet_2["Total"] = sheet_2.iloc[:, 1:].sum(axis=1)

            sheet_3 = pd.DataFrame(
                [
                    creditor_list,
                    salary_paid,
                    emi,
                    investment,
                    interest_debit,
                    gold_loan,
                    rent_paid,
                    travelling_expense,
                    donation,
                    tds_deducted,
                    tds_on_forex,
                    gst,
                    income_tax_paid,
                    property_tax,
                    general_insurance,
                    life_insurance,
                ]
            )
            sheet_3.insert(
                0,
                "Important Expenses / Payments",
                [
                    "Creditor Amount",
                    "Salaries Paid",
                    "Probable EMI",
                    "Investment Details",
                    "Interest Debit",
                    "Gold Loan (Only Interest)",
                    "Rent Paid",
                    "Travelling Expense",
                    "Donation",
                    "TDS Deducted",
                    "TDS on Forex",
                    "Total GST",
                    "Total Income Tax Paid",
                    "Property Tax",
                    "General Insurance",
                    "Life Insurance",
                ],
            )
            sheet_3 = sheet_3._append(sheet_3.iloc[0:, :].sum(), ignore_index=True)
            sheet_3.iloc[-1, 0] = "Total"
            sheet_3["Total"] = sheet_3.iloc[:, 1:].sum(axis=1)

            sheet_4 = pd.DataFrame(
                [
                    credit_card_payments,
                    forex,
                    bank_charges,
                    other_expense,
                    utility_bills,
                    Entertainment,
                    food_expenses,
                    online_shopping,
                    amount_cw,
                    amount_pos_dr,
                    upi_dr,
                    loan_gr,
                    cheque,
                    departmental,
                    pay_self,
                    suspense_dr,
                ]
            )
            sheet_4.insert(
                0,
                "Other Expenses / Payments",
                [
                    "Credit Card Payment",
                    "Forex Charges",
                    "Bank Charges",
                    "Other Expenses",
                    "Utility Bills",
                    "Subscription / Entertainment",
                    "Food Expenses",
                    "Online Shopping",
                    "Withdrawal",
                    "POS Txns - Dr",
                    "UPI-Dr",
                    "Loan Given",
                    "Cheque Paid",
                    "Departmental Store",
                    "Payment to Self",
                    "Suspense - Dr",
                ],
            )
            sheet_4 = sheet_4._append(sheet_4.iloc[0:, :].sum(), ignore_index=True)
            sheet_4.iloc[-1, 0] = "Total Debit"
            sheet_4["Total"] = sheet_4.iloc[:, 1:].sum(axis=1)

            df_list = [sheet_1, sheet_2, sheet_3, sheet_4]

        return df_list

    def summary_sheet_dsa(self, idf, open_bal, close_bal, eod):
        opening_bal = open_bal
        closing_bal = close_bal
        eod_avg_df = self.avgs_df(eod)

        def total_number_cr(df):
            number = df[df["Credit"] != 0.00]["Credit"].count()
            return number

        def total_amount_cr(df):
            sum = df["Credit"].sum(axis=0)
            return sum

        def total_number_dr(df):
            number = df[df["Debit"] != 0.00]["Debit"].count()
            return number

        def total_amount_dr(df):
            sum = df["Debit"].sum(axis=0)
            return sum

        def total_number_cd(df):
            cd = df[df["Category"] == "Cash Deposits"]
            cd_count = cd["Category"].count()
            return cd_count

        def total_amount_cd(df):
            amount = 0
            for index, row in df.iterrows():
                credit_amount = row["Credit"]
                if row["Category"] == "Cash Deposits" and credit_amount > 0:
                    amount += credit_amount
            return amount

        def total_number_cw(df):
            cw = df[df["Category"] == "Cash Withdrawal"]
            cw_count = cw["Category"].count()
            return cw_count

        def total_amount_cw(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Debit"]
                if row["Category"] == "Cash Withdrawal" and debit_amount > 0:
                    amount += debit_amount
            return amount

        def emi1(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Debit"]
                if row["Category"] == "Probable EMI" and debit_amount > 0:
                    amount += debit_amount
            return amount

        def total_amount_pos_cr(df):
            amount = 0
            for index, row in df.iterrows():
                credit_amount = row["Credit"]
                if row["Category"] == "POS-Cr" and credit_amount > 0:
                    amount += credit_amount
            return amount

        def total_amount_pos_dr(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Debit"]
                if row["Category"] == "POS-Dr" and debit_amount > 0:
                    amount += debit_amount
            return amount

        def total_investment(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Debit"]
                if row["Category"] == "Investment" and debit_amount > 0:
                    amount += debit_amount
            return amount

        def received_interest(df):
            amount = 0
            for index, row in df.iterrows():
                credit_amount = row["Credit"]
                if row["Category"] == "Interest Credit" and credit_amount > 0:
                    amount += credit_amount
            return amount

        def received_salary(df):
            amount = 0
            for index, row in df.iterrows():
                credit_amount = row["Credit"]
                if row["Category"] == "Salary Received" and credit_amount > 0:
                    amount += credit_amount
            return amount

        def interest_rece(df):
            amount = 0
            for index, row in df.iterrows():
                credit_amount = row["Credit"]
                if row["Category"] == "Bank Interest Received" and credit_amount > 0:
                    amount += credit_amount
            return amount

        def paid_interest1(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Debit"]
                if row["Category"] == "Bank Charges" and debit_amount > 0:
                    amount += debit_amount
            return amount

            # salary paid

        def paid_salary(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Debit"]
                if row["Category"] == "Salary paid" and debit_amount > 0:
                    amount += debit_amount
            return amount

        def salary_rec(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Credit"]
                if row["Category"] == "Salary Received" and debit_amount > 0:
                    amount += debit_amount
            return amount

        def paid_tds1(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Debit"]
                if row["Category"] == "TDS Deducted" and debit_amount > 0:
                    amount += debit_amount
            return amount

        def GST(df):
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Debit"]
                if row["Category"] == "GST Paid" and debit_amount > 0:
                    amount += debit_amount
            return amount

        def suspenses(df):
            amount = 0
            amount = 0
            for index, row in df.iterrows():
                debit_amount = row["Debit"]
                if row["Category"] == "Suspense" and debit_amount > 0:
                    amount += debit_amount
            return amount

        idf["Credit"] = pd.to_numeric(idf["Credit"], errors="coerce")
        idf["Debit"] = pd.to_numeric(idf["Debit"], errors="coerce")
        # idf['Month'] = idf['Value Date'].dt.strftime('%b-%Y')
        months = idf["Month"].unique()

        number_cr = {}
        amount_cr = {}

        number_dr = {}
        amount_dr = {}

        number_cd = {}
        amount_cd = {}

        number_cw = {}
        amount_cw = {}

        # no_cheque_depo = {}
        # amt_cheque_depo = {}

        # number_cash_issued = {}
        # amount_cash_issued = {}
        #
        # inward_bounce = {}
        # outward_bounce = {}
        #
        #
        # avg_eod_bal = {}
        # qtrlu_bal = {}
        # half_bal = {}
        # yrly_bal = {}
        # all_bank_avg_bal = {}
        #
        # top_5_funds = {}
        # top_5_redemptions = {}

        # bounced = {}
        emi = {}

        # amount_pos_cr = {}
        # amount_pos_dr = {}

        # datewise_bal = {}
        investment_dr = {}

        # diff_bank_ab = {}
        interest_rec = {}

        paid_interest = {}
        paid_sal = {}

        received_salary = {}
        paid_tds = {}
        # paid_gst = {}

        investment_cr = {}
        suspense = {}

        for month in months:
            new_df = idf[idf["Month"] == month].reset_index(drop=True)
            number_cr.update({month: total_number_cr(new_df)})
            amount_cr.update({month: total_amount_cr(new_df)})
            number_dr.update({month: total_number_dr(new_df)})
            amount_dr.update({month: total_amount_dr(new_df)})
            number_cd.update({month: total_number_cd(new_df)})
            amount_cd.update({month: total_amount_cd(new_df)})
            number_cw.update({month: total_number_cw(new_df)})
            amount_cw.update({month: total_amount_cw(new_df)})
            emi.update({month: emi1(new_df)})

            investment_dr.update({month: total_investment(new_df)})
            interest_rec.update({month: interest_rece(new_df)})
            paid_interest.update({month: paid_interest1(new_df)})
            paid_sal.update({month: paid_salary(new_df)})
            received_salary.update({month: salary_rec(new_df)})
            paid_tds.update({month: paid_tds1(new_df)})
            ###now we make sheets
            sheet_1 = pd.DataFrame(
                [
                    number_cr,
                    amount_cr,
                    number_dr,
                    amount_dr,
                    number_cd,
                    amount_cd,
                    number_cw,
                    amount_cw,
                    emi,
                    investment_dr,
                    interest_rec,
                    paid_interest,
                    paid_sal,
                    received_salary,
                    paid_tds,
                    opening_bal,
                    closing_bal,
                ]
            )

            sheet_1.insert(
                0,
                "Particulars",
                [
                    "Total No. of Credit Transactions",
                    "Total Amount of Credit Transactions",
                    "Total No. of Debit Transactions",
                    "Total Amount of Debit Transactions",
                    "Total No. of Cash Deposits",
                    "Total Amount of Cash Deposits",
                    "Total No. of Cash Withdrawals",
                    "Total Amount of Cash Withdrawals",
                    "EMI",
                    "Investment Details",
                    "Bank Interest Received",
                    "Bank Interest Paid (Only in OD/CC A/c)",
                    "Salaries Paid",
                    "Salary Received",
                    "TDS Deducted",
                    "Opening Balance",
                    "Closing Balance",
                ],
            )
            sheet_1["Total"] = sheet_1.apply(
                lambda row: (
                    row[1:].sum()
                    if row["Particulars"] not in ["Opening Balance", "Closing Balance"]
                    else None
                ),
                axis=1,
            )  # or `else None` if you want it to be blank

            df_list = [sheet_1]

        return df_list

    def transaction_sheet(self, df):
        if len(df["Bank"].unique()) > 1:
            tdf = df[
                [
                    "Value Date",
                    "Description",
                    "Debit",
                    "Credit",
                    "Balance",
                    "Category",
                    "Bank",
                ]
            ]
        else:
            tdf = df[
                ["Value Date", "Description", "Debit", "Credit", "Balance", "Category"]
            ]
        return tdf

    def total_investment(self, df):
        invest_df = pd.DataFrame()
        for index, row in df.iterrows():
            arow = row["Category"]
            if arow == "Investment":
                invest_df = invest_df._append(row, ignore_index=True)

        if invest_df.empty:
            invest_df = pd.DataFrame(
                columns=[
                    "Value Date",
                    "Description",
                    "Debit",
                    "Credit",
                    "Balance",
                    "Category",
                    "Bank",
                ]
            )
        return invest_df

    def redemption_investment(self, df):
        red_df = pd.DataFrame()
        for index, row in df.iterrows():
            arow = row["Category"]
            if arow == "Redemption of Investment":
                red_df = red_df._append(row, ignore_index=True)

        if red_df.empty:
            red_df = pd.DataFrame(
                columns=[
                    "Value Date",
                    "Description",
                    "Debit",
                    "Credit",
                    "Balance",
                    "Category",
                    "Bank",
                ]
            )
        return red_df

    def cash_withdraw(self, df):
        cashw_df = pd.DataFrame()
        for index, row in df.iterrows():
            arow = row["Category"]
            if arow == "Cash Withdrawal":
                cashw_df = cashw_df._append(row, ignore_index=True)

        if cashw_df.empty:
            cashw_df = pd.DataFrame(
                columns=[
                    "Value Date",
                    "Description",
                    "Debit",
                    "Credit",
                    "Balance",
                    "Category",
                    "Bank",
                ]
            )
        return cashw_df

    def cash_depo(self, df):
        cashd_df = pd.DataFrame()
        for index, row in df.iterrows():
            arow = row["Category"]
            if arow == "Cash Deposits":
                cashd_df = cashd_df._append(row, ignore_index=True)

        if cashd_df.empty:
            cashd_df = pd.DataFrame(
                columns=[
                    "Value Date",
                    "Description",
                    "Debit",
                    "Credit",
                    "Balance",
                    "Category",
                    "Bank",
                ]
            )
        return cashd_df

    def Bank_charges(self, df):
        Bank_df = pd.DataFrame()
        for index, row in df.iterrows():
            arow = row["Category"]
            if arow == "Bank Charges":
                Bank_df = Bank_df._append(row, ignore_index=True)

        if Bank_df.empty:
            Bank_df = pd.DataFrame(
                columns=[
                    "Value Date",
                    "Description",
                    "Debit",
                    "Credit",
                    "Balance",
                    "Category",
                    "Bank",
                ]
            )
        return Bank_df

    def Entertainment(self, df):
        Entertainment = pd.DataFrame()
        for index, row in df.iterrows():
            arow = row["Category"]
            if arow == "Entertainment":
                Entertainment = Entertainment._append(row, ignore_index=True)

        if Entertainment.empty:
            Entertainment = pd.DataFrame(
                columns=[
                    "Value Date",
                    "Description",
                    "Debit",
                    "Credit",
                    "Balance",
                    "Category",
                    "Bank",
                ]
            )
        return Entertainment

    def div_int(self, df):
        iii = pd.DataFrame()
        for index, row in df.iterrows():
            arow = row["Category"]
            if arow == "Redemption, Dividend & Interest":
                iii = iii._append(row, ignore_index=True)

        if iii.empty:
            iii = pd.DataFrame(
                columns=[
                    "Value Date",
                    "Description",
                    "Debit",
                    "Credit",
                    "Balance",
                    "Category",
                    "Bank",
                ]
            )
        return iii

    def emi(self, df):
        em_i = pd.DataFrame()
        for index, row in df.iterrows():
            arow = row["Category"]
            if arow == "Probable EMI":
                em_i = em_i._append(row, ignore_index=True)

        if em_i.empty:
            em_i = pd.DataFrame(
                columns=[
                    "Value Date",
                    "Description",
                    "Debit",
                    "Credit",
                    "Balance",
                    "Category",
                    "Bank",
                ]
            )
        return em_i

    def refund_reversal(self, df):
        refund = df[df["Category"].str.contains("Refund/Reversal", na=False)]

        if refund.empty:
            refund = pd.DataFrame(
                columns=[
                    "Value Date",
                    "Description",
                    "Debit",
                    "Credit",
                    "Balance",
                    "Category",
                    "Bank",
                ]
            )
        return refund

    def creditor_list(self, df):
        df["Debit"] = pd.to_numeric(df["Debit"], errors="coerce")
        df["Credit"] = pd.to_numeric(df["Credit"], errors="coerce")
        debit = df[
            (~df["Debit"].isnull()) & ((df["Credit"].isnull()) | (df["Credit"] == 0))
        ]
        patterns = [
            "toib-",
            "brn-clg-chq",
            "mmt/imps",
            "neftdr",
            "neft/mb/",
            "nft/",
            "mob/tpft",
            "nlcindialtd",
            "neft/mb/ax",
            "tortgs",
            "rtgsdr",
            "mob/tpft/",
            "imb/",
            "imps",
            "imps/p2a",
            "mob/selfft/",
            "inb/",
            "inb-",
            "chqpaid",
            "fundtrf",
            "iconn",
            "imps-cib",
            "imps-inet",
            "imps-rib",
            "imps-mob",
            "inft",
            "mbk/xfer",
            "neft",
            "payc",
            "r-utr",
            "vmt-icon",
            "chqpaid",
            "byclg",
            "rtgs",
            "neftn",
            "inb-",
            "neft-barb",
            "ecs/",
        ]
        regex_pattern = "|".join(patterns)
        Creditor_list = debit[
            debit["Description"].str.contains(regex_pattern, case=False)
        ]

        def extract_name(description):
            match = re.match(r"^[a-zA-Z\s]+$", description)
            return match.group(0) if match else None

        name_transactions = df[
            (df["Category"] == "Creditor") & (df["Description"].apply(self.is_name))
        ]
        name_transactions["Category"] = name_transactions["Description"].apply(
            extract_name
        )
        Creditor_list = pd.concat([Creditor_list, name_transactions])

        # Additional code to exclude specified keywords
        exclude_keywords = [
            "ach/",
            "ach-",
            "achdr",
            "achd",
            "bajajfinance",
            "cms/",
            "lamum",
            "lcfm",
            "lnpy",
            "loanreco",
            "lptne",
            "nach",
            "magmafincorpltd",
            "toachdraditybirl",
            "toachdrambitfinv",
            "toachdrclixcapita",
            "toachdrdeutscheb",
            "toachdrdhaniloan",
            "toachdrfedbankfi",
            "toachdrfullerton",
            "toachdrindiabulls",
            "toachdrindinfhouf",
            "toachdrindusind",
            "toachdrlendingkar",
            "toachdrmagmafinco",
            "toachdrmahnimahin",
            "toachdrmoneywisef",
            "toachdrneogrowth",
            "toachdrtatacapita",
            "toachdrtpachmag",
            "toachdrtpachneo",
            "toachdrtpcapfrst",
            "toachdryesbankr",
            "gsttaxpayment",
        ]
        exclude_pattern = "|".join(exclude_keywords)
        Creditor_list = Creditor_list[
            ~Creditor_list["Description"].str.contains(exclude_pattern, case=False)
        ]
        exclude_descriptions = ["billdesk", "gsttaxpayment", "atomstockbroker"]
        for exclude in exclude_descriptions:
            Creditor_list = Creditor_list[
                ~Creditor_list["Description"].str.contains(exclude, case=False)
            ]
        exclude_categories = [
            "Payment Received",
            "Payment Made",
            "Suspense",
            "fastag",
            "Refund/Reversal",
            "Salary Paid",
            "Loan given",
            "Credit Card Payment",
            "Food Expense/Hotel",
            "Income Tax Paid",
            "Rent Paid",
            "Utility Bills",
            "Reimbursement",
            "Travelling Expense",
            "Bank Charges",
            "POS-Cr",
            "POS-Dr",
            "Payment Made",
            "Payment Received",
            "Cash Withdrawal",
            "Bonus Paid",
            "General insurance",
            "Investment",
            "Online Shopping",
            "Probable EMI",
            "TDS Deducted",
            "GST Paid",
        ]
        for exclude in exclude_categories:
            Creditor_list = Creditor_list[
                ~Creditor_list["Category"].str.contains(exclude, case=False)
            ]
        Creditor_list = Creditor_list.sort_values(by="Category")
        return Creditor_list

    def debtor_list(self, df):
        df["Debit"] = pd.to_numeric(df["Debit"], errors="coerce")
        df["Credit"] = pd.to_numeric(df["Credit"], errors="coerce")
        credit = df[
            (~df["Credit"].isnull()) & ((df["Debit"].isnull()) | (df["Debit"] == 0))
        ]
        patterns = [
            "toib-",
            "neft",
            "mmt/imps",
            "neftcr",
            "imps",
            "tortgs",
            "rtgs",
            "rtgscr",
            "ecs/",
            "mob/tpft/",
            "imb/",
            "mob/selfft/",
            "inb/",
            "imps-mob",
            "nft/",
            "byclg",
            "inb-",
            "neft-",
            "googleindiadigital",
            "gsttaxpayment",
        ]
        regex_pattern = "|".join(patterns)
        Debtor_list = credit[
            credit["Description"].str.contains(regex_pattern, case=False)
        ]

        exclude_categories = [
            "Redemption, Dividend & Interest",
            "Suspense",
            "Redemption of Investment",
            "Refund/Reversal",
            "Loan",
            "Provident Fund",
            "Payment Made",
            "Payment Received",
            "Bounce",
            "Reimbursement",
            "GST Paid",
        ]
        Debtor_list = Debtor_list[
            ~Debtor_list["Category"].str.contains(
                "|".join(exclude_categories), case=False
            )
        ]

        def extract_name(description):
            match = re.match(r"^[a-zA-Z\s]+$", description)
            return match.group(0) if match else None

        name_transactions = df[
            (df["Category"] == "Debtor") & (df["Description"].apply(self.is_name))
        ]
        name_transactions["Category"] = name_transactions["Description"].apply(
            extract_name
        )
        Debtor_list = pd.concat([Debtor_list, name_transactions])

        Debtor_list = Debtor_list.sort_values(by="Category")
        return Debtor_list

    def is_name(self, description):
        return bool(re.match(r"^[a-zA-Z\s]+$", description))

    def categorize_name_transactions(self, df):
        # Apply 'NameTransaction' category for descriptions that appear to be names
        df.loc[
            (
                df["Category"].isin(["Suspense", "Payment Made", "Payment Received"])
            )  # Check if Category is either 'Suspense' or 'Payment Made'
            & df["Description"].apply(
                self.is_name
            )  # Descriptions that appear to be names
            & (
                df["Category"] != "NameTransaction"
            ),  # Ensure it's not already categorized as 'NameTransaction'
            "Category",
        ] = "NameTransaction"
        df.loc[
            (
                df["Category"] == "NameTransaction"
            )  # Consider only 'NameTransaction' category
            & df["Credit"].notna(),  # Credit transactions
            "Category",
        ] = "Debtor"
        df.loc[
            (
                df["Category"] == "NameTransaction"
            )  # Consider only 'NameTransaction' category
            & df["Debit"].notna(),  # Debit transactions
            "Category",
        ] = "Creditor"

        return df

    def another_method(self, df):
        # df = self.categorize_name_transactions(df)
        Creditor_list = self.creditor_list(df)
        Debtor_list = self.debtor_list(df)
        NEW_DF = df.copy()
        Creditor_list["Category"] = "Creditor"
        Debtor_list["Category"] = "Debtor"
        NEW_DF.update(Creditor_list)
        NEW_DF.update(Debtor_list)

        return NEW_DF

    def suspense_credit(self, df):
        c_df = pd.DataFrame()
        for index, row in df.iterrows():
            credit_amount = pd.to_numeric(row["Credit"], errors="coerce")
            arow = row["Category"]
            if arow == "Suspense" and credit_amount > 0:
                c_df = c_df._append(row, ignore_index=True)
        if c_df.empty:
            c_df = pd.DataFrame(columns=["Value Date", "Description", "Credit"])
        else:
            c_df = c_df[["Value Date", "Description", "Credit"]]
        return c_df

    def suspense_debit(self, df):
        d_df = pd.DataFrame()
        for index, row in df.iterrows():
            debit_amount = pd.to_numeric(row["Debit"], errors="coerce")
            arow = row["Category"]
            if arow == "Suspense" and debit_amount > 0:
                d_df = d_df._append(row, ignore_index=True)
        if d_df.empty:
            d_df = pd.DataFrame(columns=["Value Date", "Description", "Debit"])
        else:
            d_df = d_df[["Value Date", "Description", "Debit"]]
        return d_df

    def payment(self, df):
        df_debit = df[df["Debit"].notnull()]
        iii = pd.DataFrame()
        for index, row in df_debit.iterrows():
            new_row = {
                "Date": row["Value Date"],
                "Effective Date": "",
                "Bill Ref": "",
                "Dr Ledger": row["Category"],
                "Cr Ledger": "",
                "Amount": row["Debit"],
                "Narration": row[
                    "Description"
                ],  # Assuming you want to keep Narration as the description
            }
            iii = iii._append(new_row, ignore_index=True)
        if iii.empty:
            iii = pd.DataFrame(
                columns=[
                    "Date",
                    "Effective Date",
                    "Bill Ref",
                    "Dr Ledger",
                    "Cr Ledger",
                    "Amount",
                    "Narration",
                ]
            )
        return iii

    def receipt(self, df):
        df_credit = df[df["Credit"].notnull()]
        iii = pd.DataFrame()
        for index, row in df_credit.iterrows():
            new_row = {
                "Date": row["Value Date"],
                "Effective Date": "",
                "Cr Ledger": row["Category"],
                "Dr Ledger": "",
                "Amount": row["Credit"],
                "Narration": row[
                    "Description"
                ],  # Assuming you want to keep Narration as the description
            }
            iii = iii._append(new_row, ignore_index=True)
        if iii.empty:
            iii = pd.DataFrame(
                columns=[
                    "Date",
                    "Effective Date",
                    "Cr Ledger",
                    "Dr Ledger",
                    "Amount",
                    "Narration",
                ]
            )

        return iii

    def BOUNCE(self, df):
        d_df = pd.DataFrame()
        for index, row in df.iterrows():
            debit_amount = pd.to_numeric(row["Debit"], errors="coerce")
            arow = row["Category"]
            if arow == "Bounce" and debit_amount > 0:
                d_df = d_df._append(row, ignore_index=True)
        if d_df.empty:
            d_df = pd.DataFrame(
                columns=[
                    "Value Date",
                    "Description",
                    "Debit",
                    "Credit",
                    "Balance",
                    "Category",
                    "Bank",
                ]
            )
        return d_df

    def BankwiseEli(self, data, eod):

        error_df = pd.DataFrame()
        # Check if 'Avg_Last_6_Months' column exists
        # if 'Avg_Last_6_Months' not in data.columns:
        #     print("'Avg_Last_6_Months' column not found in DataFrame.")
        #     return error_df  # Or return an empty DataFrame as needed

        processed_data = self.calculate_fixed_day_average(eod)

        # Check for the existence of 'Avg_Last_6_Months' column
        if "Avg_Last_6_Months" not in processed_data.columns:
            print("'Avg_Last_6_Months' column not found in processed_data.")
            return error_df  # Or handle the absence of this column as needed

        # Conditional handling for 'Avg_Last_12_Months'
        if "Avg_Last_12_Months" in processed_data.columns:
            selected_columns = processed_data[
                ["Day", "Avg_Last_12_Months", "Avg_Last_6_Months"]
            ]
        else:
            print(
                "'Avg_Last_12_Months' column not found, proceeding with 'Avg_Last_6_Months' only."
            )
            selected_columns = processed_data[["Day", "Avg_Last_6_Months"]]

        # For Axis
        if "Avg_Last_6_Months" in data.columns and len(data["Avg_Last_6_Months"]) > 1:
            if not pd.isna(data["Avg_Last_6_Months"].iloc[1]):
                avg_divided_by_1_5 = data["Avg_Last_6_Months"].iloc[1] / 1.5
            else:
                avg_divided_by_1_5 = np.nan
        else:
            print(
                "'Avg_Last_6_Months' column does not exist or does not have enough data."
            )
            avg_divided_by_1_5 = np.nan

        # For 'Avg_Last_12_Months' at index 0
        if "Avg_Last_12_Months" in data.columns:
            if not pd.isna(data["Avg_Last_12_Months"].iloc[0]):
                avg_divided_by_2_idfc = data["Avg_Last_12_Months"].iloc[0] / 2
            else:
                avg_divided_by_2_idfc = np.nan
        else:
            print("'Avg_Last_12_Months' column does not exist in the DataFrame.")
            avg_divided_by_2_idfc = np.nan

        # For 'Avg_Last_6_Months' at index 2
        if "Avg_Last_6_Months" in data.columns and len(data["Avg_Last_6_Months"]) > 2:
            if not pd.isna(data["Avg_Last_6_Months"].iloc[2]):
                avg_divided_by_2_indus = data["Avg_Last_6_Months"].iloc[2] / 1.5
            else:
                avg_divided_by_2_indus = np.nan
        else:
            print(
                "'Avg_Last_6_Months' column does not exist or does not have enough data."
            )
            avg_divided_by_2_indus = np.nan

        # For 'Avg_Last_12_Months' at index 0 again
        if "Avg_Last_12_Months" in data.columns:
            if not pd.isna(data["Avg_Last_12_Months"].iloc[0]):
                avg_divided_by_2_L_T = data["Avg_Last_12_Months"].iloc[0] / 2
            else:
                avg_divided_by_2_L_T = np.nan
        else:
            print("'Avg_Last_12_Months' column does not exist in the DataFrame.")
            avg_divided_by_2_L_T = np.nan

        annual_interest_rate = 0.0870
        periods = 20 * 12
        principal = 100000
        payment_value = self.pmt(principal, annual_interest_rate, periods)
        payment_for_lap = self.pmt_lap()
        payment_for_bl = self.pmt_bl()

        # Calculating Loan value for axis
        axis_home_loan_value = None
        if payment_value != 0:
            axis_home_loan_value = avg_divided_by_1_5 / payment_value
            axis_home_loan_value = axis_home_loan_value * 100000
            axis_home_loan_value = round(axis_home_loan_value, 2)

        else:
            print("PMT calculation resulted in zero, cannot divide by zero.")

        axis_LAP_value = None
        if payment_for_lap != 0:
            axis_LAP_value = avg_divided_by_1_5 / payment_for_lap
            axis_LAP_value = axis_LAP_value * 100000
        else:
            print("PMT calculation resulted in zero, cannot divide by zero.")

        axis_bl_value = None
        if payment_for_bl != 0:
            axis_bl_value = avg_divided_by_1_5 / payment_for_bl
            axis_bl_value = axis_bl_value / payment_for_lap
            axis_bl_value = axis_bl_value * 100000
        else:
            print("PMT calculation resulted in zero, cannot divide by zero.")

        # Calculating loan value for Idfc
        Idfc_home_loan_value = None
        if payment_value != 0:
            Idfc_home_loan_value = avg_divided_by_2_idfc / payment_value
            Idfc_home_loan_value = Idfc_home_loan_value * 100000
            Idfc_home_loan_value = round(Idfc_home_loan_value, 2)

        else:
            print("PMT calculation resulted in zero, cannot divide by zero.")

        Idfc_LAP_value = None
        if payment_for_lap != 0:
            Idfc_LAP_value = avg_divided_by_2_idfc / payment_for_lap
            Idfc_LAP_value = Idfc_LAP_value * 100000
            Idfc_LAP_value = round(Idfc_LAP_value, 2)

        else:
            print("PMT calculation resulted in zero, cannot divide by zero.")

        Idfc_bl_value = None
        if payment_for_bl != 0:
            Idfc_bl_value = avg_divided_by_2_idfc / payment_for_bl
            Idfc_bl_value = Idfc_bl_value * 100000
            Idfc_bl_value = round(Idfc_bl_value, 2)
        else:
            print("PMT calculation resulted in zero, cannot divide by zero.")

        # Calculating loan value for Indus
        indus_home_loan_value = None
        if payment_value != 0:
            indus_home_loan_value = avg_divided_by_2_indus / payment_value
            indus_home_loan_value = indus_home_loan_value * 100000
            indus_home_loan_value = round(indus_home_loan_value, 2)

        else:
            print("PMT calculation resulted in zero, cannot divide by zero.")

        indus_LAP_value = None
        if payment_for_lap != 0:
            indus_LAP_value = avg_divided_by_2_indus / payment_for_lap
            indus_LAP_value = indus_LAP_value * 100000
            indus_LAP_value = round(indus_LAP_value, 2)
        else:
            print("PMT calculation resulted in zero, cannot divide by zero.")

        indus_bl_value = None
        if payment_for_bl != 0:
            indus_bl_value = avg_divided_by_2_indus / payment_for_bl
            indus_bl_value = indus_bl_value * 100000
            indus_bl_value = round(indus_bl_value, 2)
        else:
            print("PMT calculation resulted in zero, cannot divide by zero.")

        L_T_home_loan_value = None
        if payment_value != 0:
            L_T_home_loan_value = avg_divided_by_2_L_T / payment_value
            L_T_home_loan_value = L_T_home_loan_value * 100000
            L_T_home_loan_value = round(L_T_home_loan_value, 2)
        else:
            print("PMT calculation resulted in zero, cannot divide by zero.")
        L_T_LAP_value = None
        if payment_for_lap != 0:
            L_T_LAP_value = avg_divided_by_2_L_T / payment_for_lap
            L_T_LAP_value = L_T_LAP_value * 100000
            L_T_LAP_value = round(L_T_LAP_value, 2)
        else:
            print("PMT calculation resulted in zero, cannot divide by zero.")

        L_T_bl_value = None
        if payment_for_lap != 0:
            L_T_bl_value = avg_divided_by_2_L_T / payment_for_bl
            L_T_bl_value = L_T_bl_value * 100000
            L_T_bl_value = round(L_T_bl_value, 2)
        else:
            print("PMT calculation resulted in zero, cannot divide by zero.")

        # Adding new columns in the specific order, with differentiating spaces for the repeated names
        selected_columns["Bank / NBFC_1"] = " "  # Added _1 for differentiation
        selected_columns["Home Loan Eligibility"] = " "
        selected_columns["Bank / NBFC_2"] = " "  # Added _2 for differentiation
        selected_columns["LAP Loan Eligibility"] = " "
        selected_columns["Bank / NBFC_3"] = " "  # Added _3 for differentiation
        selected_columns["Loan Eligibility Business loan"] = " "

        # Axis and L&T
        selected_columns.at[1, "Bank / NBFC_1"] = "AXIS , L&t"
        selected_columns.at[1, "Home Loan Eligibility"] = (
            f"{axis_home_loan_value} , {L_T_home_loan_value}"
        )

        selected_columns.at[1, "Bank / NBFC_2"] = "AXIS"
        selected_columns.at[1, "LAP Loan Eligibility"] = axis_LAP_value

        selected_columns.at[1, "Bank / NBFC_3"] = "AXIS"
        selected_columns.at[1, "Loan Eligibility Business loan"] = axis_bl_value

        # IDFC
        selected_columns.at[0, "Bank / NBFC_1"] = "IDFC"
        selected_columns.at[0, "Home Loan Eligibility"] = Idfc_home_loan_value

        selected_columns.at[0, "Bank / NBFC_2"] = "IDFC , L&T"
        selected_columns.at[0, "LAP Loan Eligibility"] = (
            f"{Idfc_LAP_value} , {L_T_LAP_value}"
        )

        selected_columns.at[0, "Bank / NBFC_3"] = "IDFC , L&T"
        selected_columns.at[0, "Loan Eligibility Business loan"] = (
            f"{Idfc_bl_value} , {L_T_bl_value}"
        )

        # Indus
        selected_columns.at[2, "Bank / NBFC_1"] = "INDUS"
        selected_columns.at[2, "Home Loan Eligibility"] = indus_home_loan_value

        selected_columns.at[2, "Bank / NBFC_2"] = "INDUS"
        selected_columns.at[2, "LAP Loan Eligibility"] = indus_LAP_value

        selected_columns.at[2, "Bank / NBFC_2"] = "INDUS"
        selected_columns.at[2, "LAP Loan Eligibility"] = indus_bl_value

        return selected_columns

    def Pos_cr(self, df):
        d_df = pd.DataFrame()
        for index, row in df.iterrows():
            debit_amount = pd.to_numeric(row["Credit"], errors="coerce")
            arow = row["Category"]
            if arow == "POS-Cr" and debit_amount > 0:
                d_df = d_df._append(row, ignore_index=True)
        if d_df.empty:
            d_df = pd.DataFrame(
                columns=[
                    "Value Date",
                    "Description",
                    "Debit",
                    "Credit",
                    "Balance",
                    "Category",
                    "Bank",
                ]
            )
        return d_df

    def Pos_dr(self, df):
        d_df = pd.DataFrame()
        for index, row in df.iterrows():
            debit_amount = pd.to_numeric(row["Debit"], errors="coerce")
            arow = row["Category"]
            if arow == "POS-Dr" and debit_amount > 0:
                d_df = d_df._append(row, ignore_index=True)
        if d_df.empty:
            d_df = pd.DataFrame(
                columns=[
                    "Value Date",
                    "Description",
                    "Debit",
                    "Credit",
                    "Balance",
                    "Category",
                    "Bank",
                ]
            )
        return d_df

    def UPI_cr(self, df):
        d_df = pd.DataFrame()
        for index, row in df.iterrows():
            debit_amount = pd.to_numeric(row["Credit"], errors="coerce")
            arow = row["Category"]
            if arow == "UPI-Cr" and debit_amount > 0:
                d_df = d_df._append(row, ignore_index=True)
        if d_df.empty:
            d_df = pd.DataFrame(
                columns=[
                    "Value Date",
                    "Description",
                    "Debit",
                    "Credit",
                    "Balance",
                    "Category",
                    "Bank",
                ]
            )
        return d_df

    def UPI_dr(self, df):
        d_df = pd.DataFrame()
        for index, row in df.iterrows():
            debit_amount = pd.to_numeric(row["Credit"], errors="coerce")
            arow = row["Category"]
            if arow == "UPI-Dr" and debit_amount > 0:
                d_df = d_df._append(row, ignore_index=True)
        if d_df.empty:
            d_df = pd.DataFrame(
                columns=[
                    "Value Date",
                    "Description",
                    "Debit",
                    "Credit",
                    "Balance",
                    "Category",
                    "Bank",
                ]
            )
        return d_df

    ##EXCEL SHEETS MANIPULATION
    def adjust_excel_column_widths(self, filename):
        # Load the workbook
        workbook = load_workbook(filename)

        for sheetname in workbook.sheetnames:
            worksheet = workbook[sheetname]
            if sheetname == "Summary":
                self.adjust_column_width(worksheet, multiplier=1)
            else:
                self.adjust_column_width(worksheet, multiplier=1)

        # Save the modified workbook back
        workbook.save(filename)

    def adjust_column_width(self, worksheet, multiplier=1.1):
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = max_length * multiplier
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

    def adjust_column_widths_for_varied_sheets(self, filename, sheet_specs):
        # Load the workbook
        workbook = load_workbook(filename)

        # Iterate through each sheet
        for sheet_name in workbook.sheetnames:
            for phrase, column_widths in sheet_specs.items():
                if phrase in sheet_name:
                    worksheet = workbook[sheet_name]

                    # Adjust column widths
                    for column_letter, width in column_widths.items():
                        worksheet.column_dimensions[column_letter].width = width
                    break  # Stop checking other phrases once a match is found

        # Save the modified workbook back
        workbook.save(filename)

    def clear_cells_fixed(self, filename):
        """
        Clear the values from cells O22 and O23 in an Excel file.

        Parameters:
        - filename: Path to the Excel file
        """
        # Load the workbook
        workbook = openpyxl.load_workbook(filename)

        # Assuming the modifications are for the first sheet (as per previous context)
        worksheet = workbook.active

        # Fixed cell addresses to clear
        cell_addresses = ["O22", "O23"]

        # Clear the values from specified cells
        for cell_address in cell_addresses:
            worksheet[cell_address].value = None

        # Save the modified workbook back to the same file
        workbook.save(filename)

    def color_excel_tabs_inplace(self, filename):
        color_order = [
            "CCC0DA",
            "FFFF99",
            "00B0F0",
            "C4BD97",
            "CCC0DA",
            "DA9694",
            "E6B8B7",
            "FCD5B4",
            "C4D79B",
            "FCD5B4",
            "BFBFBF",
            "92D050",
        ]
        summary_sheet_color = "16365C"

        # Open the workbook
        workbook = openpyxl.load_workbook(filename)

        # Iterate through each sheet and set the tab color
        for idx, sheet_name in enumerate(workbook.sheetnames):
            if sheet_name == "Summary":
                workbook[sheet_name].sheet_properties.tabColor = summary_sheet_color
            else:
                color = color_order[idx % len(color_order)]
                workbook[sheet_name].sheet_properties.tabColor = color

        # Save the workbook
        workbook.save(filename)

    def center_text_in_bankwise_sheets(
        self, filename, bankwise_sheet_prefix="BankWise Eligibility"
    ):
        wb = openpyxl.load_workbook(filename)

        # Create an Alignment style object for centering text
        center_aligned_text = Alignment(horizontal="center", vertical="center")

        # Define a border style for the bottom border
        bottom_border = Border(bottom=Side(border_style="thick", color="000000"))
        left_border = Border(left=Side(border_style="thick", color="000000"))
        left_bottom_border = Border(
            left=Side(border_style="thick", color="000000"),
            bottom=Side(border_style="thick", color="000000"),
        )

        # Loop through each sheet in the workbook
        for sheet_name in wb.sheetnames:
            # Check if the sheet name contains the specified prefix
            if bankwise_sheet_prefix in sheet_name:
                ws = wb[sheet_name]

                # Apply center alignment to all cells
                for row in ws.iter_rows():
                    for cell in row:
                        cell.alignment = center_aligned_text

                # Apply left border to cells from D1 to D18
                for cell in ws["D1:D18"]:
                    cell[0].border = Border(
                        left=left_border.left
                    )  # Apply only the left border

                for cell in ws["F1:F18"]:
                    cell[0].border = Border(
                        left=left_border.left
                    )  # Apply only the left border

                for cell in ws["H1:H18"]:
                    cell[0].border = Border(
                        left=left_border.left
                    )  # Apply only the left border

                for cell in ws["J1:J18"]:
                    cell[0].border = Border(
                        left=left_border.left
                    )  # Apply only the left border

                # Apply bottom borders to cells from D18 to I18
                for cell in ws["D18":"I18"][0]:  # Access the first row in the slice
                    cell.border = bottom_border

                # Apply bottom borders to cells from D18 to I18
                for cell in ws["A18":"C18"][0]:  # Access the first row in the slice
                    cell.border = bottom_border

                # Apply left and bottom borders to cells D18, F18, and H18
                for cell_ref in ["D18", "F18", "H18"]:
                    ws[cell_ref].border = left_bottom_border

        # Save the workbook after making changes
        wb.save(filename)

    def color_summary_sheet(self, filename):
        # Load the workbook
        wb = openpyxl.load_workbook(filename)
        ws = wb["Summary"]  # Access the "Summary" sheet
        for cell in ws[1]:
            cell.value = None
            cell.border = Border()

        # Define colors
        fill_color = PatternFill(
            start_color="b9b9b9", end_color="b9b9b9", fill_type="solid"
        )
        royal_blue_fill = PatternFill(
            start_color="000058", end_color="000058", fill_type="solid"
        )
        white_bold_font = Font(bold=True, color="FFFFFF")
        light_blue_fill = PatternFill(
            start_color="b5cbe0", end_color="b5cbe0", fill_type="solid"
        )
        white_fill = PatternFill(
            start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
        )
        border_right = Border(
            right=Side(border_style="thin", color="000000")
        )  # thin black line
        border_thick_bottom = Border(
            bottom=Side(border_style="thin", color="000000")
        )  # thick black line

        for row in range(2, 5):  # Rows in openpyxl are 1-indexed
            for cell in ws[row]:
                cell.fill = fill_color
                cell.font = bold_font
        for row in [6, 12, 30, 49]:
            for cell in ws[row]:
                cell.fill = royal_blue_fill
                cell.font = white_bold_font
        for start_row in [6, 12, 30, 49]:
            row = start_row + 1
            while row not in [6, 12, 30, 49, 66]:
                for cell in ws[row]:
                    cell.fill = (
                        white_fill if (row - start_row) % 2 == 1 else light_blue_fill
                    )
                row += 1
        # for row_num in [16, 27, 41, 54]:
        #     for cell in ws[row_num]:
        #         cell.fill = white_fill
        for row in ws["A"]:
            row.border = border_right
        for cell in ws[66]:
            cell.border = border_thick_bottom

        def apply_alternating_fill(sheet):

            for cell in sheet[1]:
                cell.fill = royal_blue_fill
                cell.font = white_bold_font

            for row in sheet.iter_rows(min_row=3):
                fill = light_blue_fill if (row[0].row - 3) % 2 == 0 else white_fill
                for cell in row:
                    cell.fill = fill

        for sheet_name in wb.sheetnames[1:]:
            ws = wb[sheet_name]
            apply_alternating_fill(ws)

        # Save the changes
        wb.save(filename)

    def color_summary_sheet_dsa(self, filename):
        # Load the workbook
        wb = openpyxl.load_workbook(filename)
        ws = wb["Summary"]  # Access the "Summary" sheet
        ws.sheet_properties.tabColor = "1072BA"
        for cell in ws[1]:
            cell.value = None
            cell.border = Border()

        # Define colors
        fill_color = PatternFill(
            start_color="b9b9b9", end_color="b9b9b9", fill_type="solid"
        )
        royal_blue_fill = PatternFill(
            start_color="000058", end_color="000058", fill_type="solid"
        )
        white_bold_font = Font(bold=True, color="FFFFFF")
        light_blue_fill = PatternFill(
            start_color="b5cbe0", end_color="b5cbe0", fill_type="solid"
        )
        white_fill = PatternFill(
            start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
        )
        border_right = Border(
            right=Side(border_style="thin", color="000000")
        )  # thin black line
        border_thick_bottom = Border(
            bottom=Side(border_style="thin", color="000000")
        )  # thick black line

        for cell in ws[1]:
            cell.value = None
            cell.border = Border()

            # Applying grey fill and bold font to rows 2-4
        for row in range(2, 5):
            for cell in ws[row]:
                cell.fill = fill_color
                cell.font = bold_font

            # Applying royal blue fill and white bold font to row 6
        for cell in ws[6]:
            cell.fill = royal_blue_fill
            cell.font = white_bold_font

            # Applying alternate white and light blue fills from row 7 to row 23
        for row in range(7, 24):
            fill_to_use = light_blue_fill if row % 2 == 0 else white_fill
            for cell in ws[row]:
                cell.fill = fill_to_use

            # Applying right border to column A
        for cell in ws["A"]:
            cell.border = border_right

            # Applying thick bottom border to row 23
        for cell in ws[23]:
            cell.border = border_thick_bottom

        wb.save(filename)

        def apply_alternating_fill(sheet):

            for cell in sheet[1]:
                cell.fill = royal_blue_fill
                cell.font = white_bold_font

            for row in sheet.iter_rows(min_row=3):
                fill = light_blue_fill if (row[0].row - 3) % 2 == 0 else white_fill
                for cell in row:
                    cell.fill = fill

        for sheet_name in wb.sheetnames[1:]:
            ws = wb[sheet_name]
            apply_alternating_fill(ws)

        # Save the changes
        wb.save(filename)

    def format_numbers_with_commas(self, filename):
        # Load the workbook
        workbook = openpyxl.load_workbook(filename)

        # Iterate through each sheet
        for sheet in workbook.worksheets:
            # Iterate through each row and column
            for row in sheet.iter_rows(
                min_row=2, max_col=sheet.max_column, max_row=sheet.max_row
            ):
                for cell in row:
                    # Check if the cell contains a number
                    if isinstance(cell.value, (int, float)):
                        # Set the number format
                        cell.number_format = "#,##0.00"

        # Save the workbook
        workbook.save(filename)

    def add_filters_to_excel(self, filename):
        workbook = openpyxl.load_workbook(filename)

        for sheet in workbook.sheetnames[1:]:
            worksheet = workbook[sheet]
            worksheet.auto_filter.ref = worksheet.dimensions

        workbook.save(filename)

    def approximate_width(self, cell_content):
        """
        Approximates the width of a cell's content.

        Parameters:
        cell_content (str): The content of the cell.

        Returns:
        float: An approximate width of the cell content.
        """
        max_length = 0
        for char in cell_content:
            if char.islower():
                max_length += 1.3  # Approx width for lowercase
            else:
                max_length += 1.3  # Approx width for uppercase and numbers
        return max_length

    def auto_adjust_column_width(self, filename):
        """
        Automatically adjusts the column widths in all sheets of an Excel workbook based on the largest cell in each column.

        Parameters:
        filename (str): Path to the Excel file.
        """
        workbook = openpyxl.load_workbook(filename)

        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            column_widths = {}
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value:
                        # Approximate the width of the cell's content
                        estimated_width = self.approximate_width(str(cell.value))
                        column = cell.column_letter
                        column_widths[column] = max(
                            column_widths.get(column, 0), estimated_width
                        )

            for column, width in column_widths.items():
                worksheet.column_dimensions[column].width = width

        workbook.save(filename)

    def Eligibility_note(
        self, filename, empty_rows_between=1, sheet_name_keyword="BankWise Eligibility"
    ):
        note_parts = [
            "Disclaimer:",
            "1. The above loan eligibility calculations applies to self-employed clients and on the basis of Average Bank Balance only ",
            "2. The above eligibility is based on the analysis of the current uploaded bank statement. Kindly upload all bank statements to obtain more accurate eligibility.",
            "3. Final Approval will be dependent on complete thorough process and submission of relevant documents, CIBIL check, etc. by the respective Banks/NBFC's.",
            "4. Nothing contained in this eligibility should be deemed to create any right and/or interest whatsoever in favor of or against any party.",
            "5. Client data is not stored in our portal",
        ]
        note = "\r\n".join(note_parts)

        try:
            workbook = openpyxl.load_workbook(filename)
            target_sheet = None

            # Find a sheet that contains the keyword
            for name in workbook.sheetnames:
                if sheet_name_keyword in name:
                    target_sheet = workbook[name]
                    break

            if not target_sheet:
                print(
                    f"No sheet containing '{sheet_name_keyword}' found in the workbook."
                )
                return

            last_row = target_sheet.max_row
            start_row = last_row + empty_rows_between if last_row else 1

            merge_start = f"A{start_row}"
            merge_end = (
                f"H{start_row + 5}"  # Extend the merge to 5 rows below the start
            )

            target_sheet.merge_cells(f"{merge_start}:{merge_end}")

            # Add the text to the merged cells and enable text wrapping
            cell = target_sheet[merge_start]
            cell.value = note
            cell.alignment = Alignment(wrap_text=True)  # Enable wrap_text for the cell

            workbook.save(filename)
        except Exception as e:
            print(f"An error occurred: {e}")

    def Summary_note(self, filename, empty_rows_between=1, sheet_name="Summary"):
        note_parts = [
            "Disclaimer/Caveat: The entries throughout this file and tables are based on best guess basis and",
            "information filtered under expenses and income. An attempt has been made to reflect the narration as",
            "close as possible to the actuals. However, variations from above are possible based on customer profile",
            "and their transactions with parties. Kindly cross check with your clients for any discrepancies",
        ]
        note = "\r\n".join(note_parts)

        try:
            workbook = openpyxl.load_workbook(filename)
            if sheet_name not in workbook.sheetnames:
                print(f"The sheet {sheet_name} does not exist in the workbook.")
                return
            sheet = workbook[sheet_name]
            last_row = sheet.max_row
            start_row = last_row + empty_rows_between if last_row else 1

            merge_start = f"A{start_row}"
            merge_end = (
                f"H{start_row + 4}"  # Extend the merge to 3 rows below the start
            )

            sheet.merge_cells(f"{merge_start}:{merge_end}")

            # Add the text to the merged cells and enable text wrapping
            cell = sheet[merge_start]
            cell.value = note
            cell.alignment = Alignment(wrap_text=True)  # Enable wrap_text for the cell

            workbook.save(filename)
        except Exception as e:
            print(f"An error occurred: {e}")

    def Eod_note(self, filename, bank_names, note, empty_rows_between=1, column=1):
        workbook = openpyxl.load_workbook(filename)
        for bank_name in bank_names:
            sheet_name = f"{bank_name} EOD Balance"
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                last_row = sheet.max_row
                sheet.cell(row=last_row + empty_rows_between, column=column, value=note)
            else:
                print(f"No sheet named {sheet_name} found in {filename}.")
        workbook.save(filename)

    def Comon_Eod_note(self, filename, note, empty_rows_between=1, column=1):
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook[f"Combined EOD Balance"]
        last_row = sheet.max_row
        sheet.cell(row=last_row + empty_rows_between, column=column, value=note)
        workbook.save(filename)

    def Investment_note(self, filename, empty_rows_between=1, sheet_name="Investment"):
        note_parts = [
            "*This table reflects probable transactions in securities made during the year. \r\nKindly confirm the same from Annual Information Statement (AIS) reflected on the Income Tax Portal and the capital gain reports sent by the respective authorities."
        ]
        note = "\r\n".join(note_parts)

        try:
            workbook = openpyxl.load_workbook(filename)
            if sheet_name not in workbook.sheetnames:
                print(f"The sheet {sheet_name} does not exist in the workbook.")
                return
            sheet = workbook[sheet_name]
            last_row = sheet.max_row
            start_row = last_row + empty_rows_between if last_row else 1

            merge_start = f"A{start_row}"
            merge_end = (
                f"H{start_row + 2}"  # Extend the merge to 3 rows below the start
            )

            sheet.merge_cells(f"{merge_start}:{merge_end}")

            # Add the text to the merged cells and enable text wrapping
            cell = sheet[merge_start]
            cell.value = note
            cell.alignment = Alignment(wrap_text=True)  # Enable wrap_text for the cell

            workbook.save(filename)
        except Exception as e:
            print(f"An error occurred: {e}")

    def CreditorList_note(self, filename, empty_rows_between=1, sheet_name="Creditors"):
        note_parts = [
            "*The entries in this table likely pertain to payments from the parties during the period mentioned. \r\nIn case of payments through online portals, we have mentioned the portal names as reflected in the narration of the bank statement. \r\nWe would like to highlight that in case of contra entries, the name of the client will be reflected as a creditor."
        ]
        note = "\r\n".join(note_parts)

        try:
            workbook = openpyxl.load_workbook(filename)
            if sheet_name not in workbook.sheetnames:
                print(f"The sheet {sheet_name} does not exist in the workbook.")
                return
            sheet = workbook[sheet_name]
            last_row = sheet.max_row
            start_row = last_row + empty_rows_between if last_row else 1

            merge_start = f"A{start_row}"
            merge_end = (
                f"H{start_row + 2}"  # Extend the merge to 3 rows below the start
            )

            sheet.merge_cells(f"{merge_start}:{merge_end}")

            # Add the text to the merged cells and enable text wrapping
            cell = sheet[merge_start]
            cell.value = note
            cell.alignment = Alignment(wrap_text=True)  # Enable wrap_text for the cell

            workbook.save(filename)
        except Exception as e:
            print(f"An error occurred: {e}")

    def DebtorList_note(self, filename, empty_rows_between=1, sheet_name="Debtors"):
        note_parts = [
            "*The entries in this table likely pertains to receipts from the respective parties. \r\nIn case of receipts through online portals, we have mentioned the portal names as reflected in the narration of the bank statement. \r\nWe would like to highlight that in case of contra entries, the name of the client will be reflected as a debtor."
        ]
        note = "\r\n".join(note_parts)

        try:
            workbook = openpyxl.load_workbook(filename)
            if sheet_name not in workbook.sheetnames:
                print(f"The sheet {sheet_name} does not exist in the workbook.")
                return
            sheet = workbook[sheet_name]
            last_row = sheet.max_row
            start_row = last_row + empty_rows_between if last_row else 1

            merge_start = f"A{start_row}"
            merge_end = (
                f"H{start_row + 2}"  # Extend the merge to 3 rows below the start
            )

            sheet.merge_cells(f"{merge_start}:{merge_end}")

            # Add the text to the merged cells and enable text wrapping
            cell = sheet[merge_start]
            cell.value = note
            cell.alignment = Alignment(wrap_text=True)  # Enable wrap_text for the cell

            workbook.save(filename)
        except Exception as e:
            print(f"An error occurred: {e}")

    def CashWithdrawalt_note(self, filename, note, empty_rows_between=1, column=1):
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook["Cash Withdrawal"]
        last_row = sheet.max_row
        sheet.cell(row=last_row + empty_rows_between, column=column, value=note)
        workbook.save(filename)

    def Cash_Deposit_note(self, filename, note, empty_rows_between=1, column=1):
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook["Cash Deposit"]
        last_row = sheet.max_row
        sheet.cell(row=last_row + empty_rows_between, column=column, value=note)
        workbook.save(filename)

    def Emi_note(
        self, filename, note, empty_rows_between=1, column=1, sheet_name="Probable EMI"
    ):
        note = note.replace("\n", "\r\n")
        try:
            workbook = openpyxl.load_workbook(filename)
            if sheet_name not in workbook.sheetnames:
                print(f"The sheet {sheet_name} does not exist in the workbook.")
                return
            sheet = workbook[sheet_name]
            last_row = sheet.max_row
            start_row = last_row + empty_rows_between if last_row else 1
            for i, line in enumerate(note.split("\r\n")):
                sheet.cell(row=start_row + i, column=column, value=line)
            workbook.save(filename)
        except:
            print("statement")

    def create_excel_sheet(self, save_path, loan_value_df):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Color

        # Include any other specific classes or functions you need from openpyxl

        # Check if DataFrame is empty or if specific columns are missing
        if (
            loan_value_df.empty
            or "Maximum Home Loan Value" not in loan_value_df.columns
            or "Maximum LAP Value" not in loan_value_df.columns
            or "Maximum BL Value" not in loan_value_df.columns
        ):
            print("DataFrame is empty or required columns are missing.")
            return  # or handle this situation in a way that fits your application

        # Your existing logic with safety checks
        max_home_loan = (
            0
            if pd.isna(loan_value_df["Maximum Home Loan Value"].iloc[0])
            else round(loan_value_df["Maximum Home Loan Value"].iloc[0] / 1000) * 1000
        )
        max_lap = (
            0
            if pd.isna(loan_value_df["Maximum LAP Value"].iloc[0])
            else round(loan_value_df["Maximum LAP Value"].iloc[0] / 1000) * 1000
        )
        max_bl = (
            0
            if pd.isna(loan_value_df["Maximum BL Value"].iloc[0])
            else round(loan_value_df["Maximum BL Value"].iloc[0] / 1000) * 1000
        )

        commission_percentage = [
            0.45 / 100,
            0.65 / 100,
            1.00 / 100,
        ]  # Convert percentages to fractions

        # Calculate commission in Rs
        commission_home_loan = round(max_home_loan * commission_percentage[0], 2)
        commission_lap = round(max_lap * commission_percentage[1], 2)
        commission_bl = round(max_bl * commission_percentage[2], 2)

        try:
            wb = load_workbook(save_path)
        except FileNotFoundError:
            wb = Workbook()

        ws = wb.create_sheet(title="Opportunity to Earn")
        content_font = Font(size=11)
        thin_side = Side(border_style="thin", color="000000")  # color is optional

        # Create the border using the defined sides
        border = Border(
            top=thin_side, left=thin_side, right=thin_side, bottom=thin_side
        )
        # Set column widths
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 20

        # Define fonts and styles
        header_font = Font(bold=True)
        center_aligned_text = Alignment(horizontal="center")
        thin_border_side = Side(border_style="thin")
        header_border = Border(
            top=thin_border_side,
            bottom=thin_border_side,
            left=thin_border_side,
            right=thin_border_side,
        )

        # Add headers and apply styles
        headers = ["B4", "C4", "D4", "E4"]
        titles = ["Product", "Amount", "Commission %", "Commission (in Rs)"]

        for cell, title in zip(headers, titles):
            ws[cell] = title
            ws[cell].font = header_font
            ws[cell].alignment = center_aligned_text
            ws[cell].border = header_border

        # You may also want to set the width of the columns to make sure the headers fit properly
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 45

        # Create the dataframe
        data = {
            "Product": [
                "Home Loan / Balance Transfer",
                "Loan Against Property / Balance Transfer",
                "Business Loan",
                "Term Plan",
                "General Insurance",
            ],
            "Amount": [max_home_loan, max_lap, max_bl, np.nan, np.nan],
            "Commission %": ["0.45%", "0.65%", "1.00%", "1%-30%", "upto 10%"],
            "Commission (in Rs)": [
                commission_home_loan,
                commission_lap,
                commission_bl,
                np.nan,
                np.nan,
            ],
        }
        df = pd.DataFrame(data)

        # Add data to Excel from dataframe
        for r_idx, row in enumerate(
            dataframe_to_rows(df, index=False, header=False), 5
        ):
            for c_idx, value in enumerate(row, 2):
                ws.cell(row=r_idx, column=c_idx, value=value)
                ws.cell(row=r_idx, column=c_idx).font = content_font
                ws.cell(row=r_idx, column=c_idx).border = border

        # Merge cells and add texts
        from openpyxl.styles import Alignment, Font
        from openpyxl.styles.colors import BLUE

        # Assuming 'ws' is your worksheet and 'header_font' is your predefined font
        ws.merge_cells("B2:E2")
        ws["B2"] = (
            "Congratulations!!! Here is an Opportunity for You to Earn.\n\nYour Client is Eligible for the following products"
        )

        # Define the font with the desired dark blue color. Excel uses ARGB format, 'FF' for solid.
        dark_blue_font = Font(
            color="FF134F63",
            size=header_font.size,
            bold=header_font.bold,
            italic=header_font.italic,
        )
        ws["B2"].alignment = Alignment(wrap_text=True)

        ws.row_dimensions[2].height = 60
        # Set the font to the cell
        ws["B2"].font = dark_blue_font

        # Define the alignment to left-aligned text.
        left_aligned_text = Alignment(
            horizontal="left", vertical="center", wrap_text=True
        )

        # Set the alignment to the cell
        ws["B2"].alignment = left_aligned_text

        title_text = "To Proceed Further:"
        title_cell = ws["B10"]
        title_cell.value = title_text
        title_cell.font = Font(bold=True, size=11)
        title_cell.alignment = Alignment(
            horizontal="left", vertical="top", wrap_text=True
        )

        # Add the detailed message to B11 without bold
        detail_message = (
            "***In case your client is interested in any of the above products, you can contact our trusted "
            "vendor M/s BizPedia Tech Private Limited on 8828824242 and email id support@leadsathi.in. "
            'Kindly use the promo code "CYPHERSOLEARN" to avail the higher commission structure.\n\n'
            "Once the referrals are successfully closed, you will be eligible for payouts based on the above commission structure.\n\n"
            "The respective payments will be released on the 25th of the next month."
        )

        detail_message_cell = ws["B11"]
        detail_message_cell.value = detail_message
        detail_message_cell.font = Font(size=11)
        detail_message_cell.alignment = Alignment(
            horizontal="left", vertical="top", wrap_text=True
        )
        ws.row_dimensions[11].height = 100
        ws.merge_cells("B10:E10")
        ws.merge_cells("B11:E11")

        merge_ranges_to_unmerge = [
            str(merge_range)
            for merge_range in ws.merged_cells.ranges
            if "B13" in merge_range
        ]

        # Unmerge them in a separate loop
        for merge_range in merge_ranges_to_unmerge:
            ws.unmerge_cells(merge_range)

        # Now you can safely set the value of the top-left cell of the range you intend to merge
        ws["B13"] = (
            "Disclaimer:\n\n1. The above loan eligibility calculations apply to self-employed clients only.\n"
            "2. For salaried clients, the vendor will need more details to calculate the eligibility.\n"
            "3. The above eligibility is based on the analysis of the current uploaded bank statement. Kindly upload "
            "all bank statements to obtain more accurate eligibility.\n"
            "4. Final Approval will be dependent on complete thorough process and submission of relevant documents, "
            "CIBIL check, etc.\n"
            "5. Nothing contained in this eligibility should be deemed to create any right and/or interest whatsoever "
            "in favor of or against any party."
        )

        # Set the alignment and font for the cell before merging
        cell = ws["B13"]
        cell.alignment = Alignment(wrap_text=True)
        cell.font = Font(size=11)

        # Merge the cells after setting the value
        ws.merge_cells("B13:E21")

        for row in ws.iter_rows(min_row=4, max_row=9, min_col=4, max_col=4):
            for cell in row:
                cell.alignment = center_aligned_text
        # Save the workbook
        wb.save(save_path)

        return save_path

    def Entertainment_note(self, filename, note, empty_rows_between=1, column=1):
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook["Subscription_Entertainment"]
        last_row = sheet.max_row
        sheet.cell(row=last_row + empty_rows_between, column=column, value=note)
        workbook.save(filename)

    def Refund_note(self, filename, note, empty_rows_between=1, column=1):
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook["Refund-Reversal"]
        last_row = sheet.max_row
        sheet.cell(row=last_row + empty_rows_between, column=column, value=note)
        workbook.save(filename)

    def Suspense_Credit_note(
        self, filename, empty_rows_between=1, sheet_name="Suspense Credit"
    ):
        note_parts = [
            "*This table pertains to transactions unidentified as per the current ledger bifurcation of the software. \r\nIn case of any technical errors, inconvience is highly regretted and feedback is appreciated."
        ]
        note = "\r\n".join(note_parts)

        try:
            workbook = openpyxl.load_workbook(filename)
            if sheet_name not in workbook.sheetnames:
                print(f"The sheet {sheet_name} does not exist in the workbook.")
                return
            sheet = workbook[sheet_name]
            last_row = sheet.max_row
            start_row = last_row + empty_rows_between if last_row else 1

            merge_start = f"A{start_row}"
            merge_end = (
                f"H{start_row + 2}"  # Extend the merge to 3 rows below the start
            )

            sheet.merge_cells(f"{merge_start}:{merge_end}")

            # Add the text to the merged cells and enable text wrapping
            cell = sheet[merge_start]
            cell.value = note
            cell.alignment = Alignment(wrap_text=True)  # Enable wrap_text for the cell

            workbook.save(filename)
        except Exception as e:
            print(f"An error occurred: {e}")

    def Suspense_Debit_note(
        self, filename, empty_rows_between=1, sheet_name="Suspense Debit"
    ):
        note_parts = [
            "*This table likely pertains to transactions unidentified as per the current ledger bifurcation of the software."
        ]
        note = "\r\n".join(note_parts)

        try:
            workbook = openpyxl.load_workbook(filename)
            if sheet_name not in workbook.sheetnames:
                print(f"The sheet {sheet_name} does not exist in the workbook.")
                return
            sheet = workbook[sheet_name]
            last_row = sheet.max_row
            start_row = last_row + empty_rows_between if last_row else 1

            merge_start = f"A{start_row}"
            merge_end = (
                f"H{start_row + 2}"  # Extend the merge to 3 rows below the start
            )

            sheet.merge_cells(f"{merge_start}:{merge_end}")

            # Add the text to the merged cells and enable text wrapping
            cell = sheet[merge_start]
            cell.value = note
            cell.alignment = Alignment(wrap_text=True)  # Enable wrap_text for the cell

            workbook.save(filename)
        except Exception as e:
            print(f"An error occurred: {e}")
